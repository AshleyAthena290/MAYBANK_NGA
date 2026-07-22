#!/usr/bin/env python3
"""Extract ALL API specifications from DEP_App_Dashboard_DDD_API_Design_v1.xlsx"""

import openpyxl
import json
from typing import Dict, List, Any

class ComprehensiveAPIExtractor:
    def __init__(self, excel_file):
        self.wb = openpyxl.load_workbook(excel_file, data_only=True)
        self.apis = []
        
    def extract_all_apis(self) -> List[Dict]:
        """Extract all API specifications from the workbook"""
        # Get all sheets except README and INDEX
        api_sheets = [s for s in self.wb.sheetnames if s not in ['README', 'INDEX']]
        
        print(f"Processing {len(api_sheets)} API sheets...")
        
        for sheet_name in api_sheets:
            api_spec = self.extract_api_spec(sheet_name)
            if api_spec:
                self.apis.append(api_spec)
                print(f"  ✓ {sheet_name}")
        
        return self.apis
    
    def extract_api_spec(self, sheet_name: str) -> Dict:
        """Extract API specification from a single sheet"""
        ws = self.wb[sheet_name]
        spec = {
            'sheet_name': sheet_name,
            'service': None,
            'description': None,
            'business_rule': None,
            'method': None,
            'url': None,
            'message_type': None,
            'headers': [],
            'body_params': [],
            'query_params': [],
            'response_fields': [],
            'request_sample': None,
            'response_sample': None,
        }
        
        rows = list(ws.iter_rows(values_only=True))
        
        for row_idx, row in enumerate(rows):
            if not any(row):
                continue
            
            # Safely get column values
            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            # Extract basic info (top section)
            if col_a == 'Service':
                spec['service'] = col_b
            elif col_a == 'Description':
                spec['description'] = col_b
            elif col_a == 'Business Rule':
                spec['business_rule'] = col_b
            elif col_a == 'Method':
                spec['method'] = col_b
            elif col_a == 'URL':
                spec['url'] = col_b
            elif col_a == 'Message Type':
                spec['message_type'] = col_b
            elif col_a == 'Request Sample':
                spec['request_sample'] = col_b
            elif col_a == 'Response Sample':
                spec['response_sample'] = col_b
            
            # Parse parameter sections
            if col_a == 'Request' and col_b == 'HTTP Header':
                # Next rows are header definitions
                for sub_row_idx in range(row_idx + 2, min(row_idx + 50, len(rows))):
                    sub_row = rows[sub_row_idx]
                    if not any(sub_row):
                        continue
                    
                    # Check if this is the header row or parameter row
                    sub_col_a = str(sub_row[0]).strip() if sub_row[0] else ""
                    sub_col_b = str(sub_row[1]).strip() if len(sub_row) > 1 and sub_row[1] else ""
                    
                    if sub_col_a in ['HTTP Body', 'Request Parameter', 'Response'] or (sub_col_a == 'Request' and sub_col_b):
                        break
                    
                    if sub_col_b and sub_col_b != 'Name' and sub_col_a not in ['Name', None, '']:
                        # This is a header parameter row
                        param = {
                            'name': sub_col_b,
                            'type': str(sub_row[2]).strip() if len(sub_row) > 2 and sub_row[2] else '',
                            'length': str(sub_row[3]).strip() if len(sub_row) > 3 and sub_row[3] else '',
                            'mandatory': str(sub_row[4]).strip() if len(sub_row) > 4 and sub_row[4] else '',
                            'sample': str(sub_row[5]).strip() if len(sub_row) > 5 and sub_row[5] else '',
                            'description': str(sub_row[6]).strip() if len(sub_row) > 6 and sub_row[6] else '',
                        }
                        spec['headers'].append(param)
            
            elif col_a == 'HTTP Body':
                # Parse body parameters
                for sub_row_idx in range(row_idx + 2, min(row_idx + 100, len(rows))):
                    sub_row = rows[sub_row_idx]
                    if not any(sub_row):
                        continue
                    
                    sub_col_a = str(sub_row[0]).strip() if sub_row[0] else ""
                    sub_col_b = str(sub_row[1]).strip() if len(sub_row) > 1 and sub_row[1] else ""
                    
                    if sub_col_a in ['Request Parameter', 'Response'] or (sub_col_a == 'Request' and sub_col_b):
                        break
                    
                    if sub_col_b and sub_col_b != 'Name' and sub_col_a not in ['Name', None, '']:
                        param = {
                            'name': sub_col_b,
                            'type': str(sub_row[2]).strip() if len(sub_row) > 2 and sub_row[2] else '',
                            'length': str(sub_row[3]).strip() if len(sub_row) > 3 and sub_row[3] else '',
                            'mandatory': str(sub_row[4]).strip() if len(sub_row) > 4 and sub_row[4] else '',
                            'sample': str(sub_row[5]).strip() if len(sub_row) > 5 and sub_row[5] else '',
                            'description': str(sub_row[6]).strip() if len(sub_row) > 6 and sub_row[6] else '',
                        }
                        spec['body_params'].append(param)
            
            elif col_a == 'Request Parameter':
                # Parse query parameters
                for sub_row_idx in range(row_idx + 2, min(row_idx + 100, len(rows))):
                    sub_row = rows[sub_row_idx]
                    if not any(sub_row):
                        continue
                    
                    sub_col_a = str(sub_row[0]).strip() if sub_row[0] else ""
                    sub_col_b = str(sub_row[1]).strip() if len(sub_row) > 1 and sub_row[1] else ""
                    
                    if sub_col_a in ['Response', 'Request Sample'] or (sub_col_a == 'Request' and sub_col_b):
                        break
                    
                    if sub_col_b and sub_col_b != 'Name' and sub_col_a not in ['Name', None, '']:
                        param = {
                            'name': sub_col_b,
                            'type': str(sub_row[2]).strip() if len(sub_row) > 2 and sub_row[2] else '',
                            'length': str(sub_row[3]).strip() if len(sub_row) > 3 and sub_row[3] else '',
                            'mandatory': str(sub_row[4]).strip() if len(sub_row) > 4 and sub_row[4] else '',
                            'sample': str(sub_row[5]).strip() if len(sub_row) > 5 and sub_row[5] else '',
                            'description': str(sub_row[6]).strip() if len(sub_row) > 6 and sub_row[6] else '',
                        }
                        spec['query_params'].append(param)
            
            elif col_a == 'Response':
                # Parse response fields
                for sub_row_idx in range(row_idx + 1, min(row_idx + 100, len(rows))):
                    sub_row = rows[sub_row_idx]
                    if not any(sub_row):
                        continue
                    
                    sub_col_a = str(sub_row[0]).strip() if sub_row[0] else ""
                    sub_col_b = str(sub_row[1]).strip() if len(sub_row) > 1 and sub_row[1] else ""
                    
                    if sub_col_a in ['Request Sample', 'Response Sample'] or (sub_col_a and sub_col_b and sub_col_a not in ['Name', None, ''] and sub_col_b == 'Name'):
                        break
                    
                    if sub_col_b and sub_col_b != 'Name' and sub_col_a not in ['Name', 'Response', None, '']:
                        param = {
                            'name': sub_col_b,
                            'type': str(sub_row[2]).strip() if len(sub_row) > 2 and sub_row[2] else '',
                            'length': str(sub_row[3]).strip() if len(sub_row) > 3 and sub_row[3] else '',
                            'mandatory': str(sub_row[4]).strip() if len(sub_row) > 4 and sub_row[4] else '',
                            'sample': str(sub_row[5]).strip() if len(sub_row) > 5 and sub_row[5] else '',
                            'description': str(sub_row[6]).strip() if len(sub_row) > 6 and sub_row[6] else '',
                        }
                        spec['response_fields'].append(param)
        
        # Only return if service name is found
        return spec if spec['service'] else None

# Main execution
if __name__ == '__main__':
    print("=" * 100)
    print("EXTRACTING ALL API SPECIFICATIONS FROM DDD FILE")
    print("=" * 100)
    
    extractor = ComprehensiveAPIExtractor(r'input\api\DEP_App_Dashboard_DDD_API_Design_v1.xlsx')
    apis = extractor.extract_all_apis()
    
    print(f"\n✓ Successfully extracted {len(apis)} APIs\n")
    
    # Display summary
    for i, api in enumerate(apis, 1):
        print(f"{i}. {api['service']}")
        print(f"   Method: {api['method']}, URL: {api['url']}")
        print(f"   Headers: {len(api['headers'])}, Body: {len(api['body_params'])}, Query: {len(api['query_params'])}, Response: {len(api['response_fields'])}")
    
    # Save to JSON
    with open('all_apis_metadata.json', 'w') as f:
        json.dump({
            'total_apis': len(apis),
            'apis': apis
        }, f, indent=2)
    
    print(f"\n✓ API metadata saved to all_apis_metadata.json")
