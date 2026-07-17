import openpyxl
import json
import sys
from io import StringIO

# Set encoding to UTF-8
sys.stdout = StringIO()
import os
os.environ['PYTHONIOENCODING'] = 'utf-8'

# Load the workbook
wb = openpyxl.load_workbook(r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx")

def clean_value(val):
    """Clean cell value"""
    if val is None:
        return None
    return str(val).strip()

def extract_section_rows(ws, start_row, section_header):
    """Extract rows from a section"""
    rows_data = []
    found_section = False
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True)):
        # Look for section header
        if not found_section:
            if any(section_header in str(cell if cell else '') for cell in row if cell):
                found_section = True
                continue
        
        if found_section:
            # Check if we hit next section or empty rows
            if row[0] and any(x in str(row[0]) for x in ['Request', 'Response', 'Business', 'Error', 'Validation', 'Sample']):
                if section_header not in str(row[0]):
                    break
            
            # Skip header row
            if 'Name' in str(row[1] if len(row) > 1 else ''):
                continue
            
            # If first column is empty but has data in other columns
            if not row[0] or row[0] == 'None':
                if any(row[1:]):
                    rows_data.append([clean_value(c) for c in row])
            elif section_header in str(row[0]):
                continue
            else:
                break
    
    return rows_data

def extract_api_detailed(sheet_name):
    """Extract detailed API info from a sheet"""
    ws = wb[sheet_name]
    
    api = {}
    
    # Extract basic info (first 6 rows usually)
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] == 'Service':
            api['service'] = clean_value(row[1])
        elif row[0] == 'Method':
            api['method'] = clean_value(row[1])
        elif row[0] == 'URL':
            api['url'] = clean_value(row[1])
        elif row[0] == 'Message Type':
            api['messageType'] = clean_value(row[1])
    
    # Extract sections
    api['headers'] = extract_section_rows(ws, 1, 'HTTP Header')
    api['body'] = extract_section_rows(ws, 1, 'HTTP Body')
    api['parameters'] = extract_section_rows(ws, 1, 'Request Parameter')
    api['response'] = extract_section_rows(ws, 1, 'Response')
    
    # Extract all rows to look for samples, business rules, etc
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([clean_value(c) if c else None for c in row])
    
    # Find Request Sample and Response Sample
    for i, row in enumerate(all_rows):
        if row[0] == 'Request Sample' and row[1]:
            api['requestSample'] = row[1]
        elif row[0] == 'Response Sample' and row[1]:
            api['responseSample'] = row[1][:300]
        elif row[0] == 'Business Rule':
            api['businessRules'] = [r[1] for r in all_rows[i+1:i+20] if r[1] and r[1] not in ['Name', 'Type', 'Description']][:10]
        elif row[0] == 'Error Code':
            api['errorCodes'] = [[r[1], r[2]] for r in all_rows[i+1:i+20] if r[1] and r[1] not in ['Code', 'Description']]
        elif row[0] == 'Validation':
            api['validations'] = [r[1] for r in all_rows[i+1:i+20] if r[1] and r[1] not in ['Rule', 'Description']][:10]
    
    return api

# Get API list from index
def get_api_list():
    """Extract API list from API_Specs_Index"""
    ws = wb['API_Specs_Index']
    apis = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # API Name
            apis.append({
                'no': clean_value(row[0]),
                'sheet': clean_value(row[1]),
                'name': clean_value(row[2]),
                'method': clean_value(row[3]),
                'microservice': clean_value(row[4]),
                'endpoint': clean_value(row[5]),
                'status': clean_value(row[6]),
                'remarks': clean_value(row[7])
            })
    
    return apis

# Extract and save to file for readable output
output = []

output.append("=" * 120)
output.append("ECLIPSE ACCOUNT DASHBOARD API SPECIFICATIONS EXTRACTION")
output.append("=" * 120)
output.append("")

# Get API list
api_list = get_api_list()
output.append(f"\nTOTAL UNIQUE APIs: {len([a for a in api_list if a['name']])}")
output.append("\n" + "=" * 120)
output.append("1. API LIST SUMMARY")
output.append("=" * 120)

api_dict = {}
for api in api_list:
    if api['name']:
        if api['name'] not in api_dict:
            api_dict[api['name']] = {
                'method': api['method'],
                'endpoints': set(),
                'microservice': api['microservice'],
                'status': api['status']
            }
        if api['endpoint']:
            api_dict[api['name']]['endpoints'].add(api['endpoint'])

output.append(f"\n{'No.':<5} {'API Name':<40} {'Method':<10} {'Microservice':<20} {'Endpoint':<60}")
output.append("-" * 140)

for i, (name, info) in enumerate(api_dict.items(), 1):
    endpoints = list(info['endpoints'])
    endpoint = endpoints[0][:55] + "..." if len(endpoints[0]) > 58 else endpoints[0]
    output.append(f"{i:<5} {name:<40} {info['method']:<10} {info['microservice']:<20} {endpoint:<60}")

# Extract detailed specs for key API sheets
output.append("\n\n" + "=" * 120)
output.append("2. DETAILED API SPECIFICATIONS")
output.append("=" * 120)

api_sheets = {s: s for s in wb.sheetnames 
              if any(x in s.lower() for x in ['get', 'set', 'block', 'reactivate', 'link', 'auto', 'remove'])}

for sheet_name in sorted(api_sheets.keys()):
    try:
        api = extract_api_detailed(sheet_name)
        
        output.append(f"\n{'='*120}")
        output.append(f"API: {api.get('service', 'Unknown')}")
        output.append(f"{'='*120}")
        output.append(f"\nHTTP Method: {api.get('method', 'N/A')}")
        output.append(f"Endpoint URL: {api.get('url', 'N/A')}")
        output.append(f"Message Type: {api.get('messageType', 'N/A')}")
        
        # Request Headers
        if api.get('headers'):
            output.append(f"\nREQUEST HEADERS:")
            for h in api['headers']:
                output.append(f"  • {h[1]} ({h[2]}) - Mandatory: {h[4]}")
                if h[6]:
                    output.append(f"    Description: {h[6]}")
                if h[5]:
                    output.append(f"    Sample: {h[5]}")
        
        # Request Parameters
        if api.get('parameters'):
            output.append(f"\nREQUEST PARAMETERS:")
            for p in api['parameters']:
                output.append(f"  • {p[1]} ({p[2]}) - Mandatory: {p[4]}")
                if p[6]:
                    output.append(f"    Description: {p[6]}")
                if p[5]:
                    output.append(f"    Sample: {p[5]}")
        
        # Request Body
        if api.get('body'):
            output.append(f"\nREQUEST BODY:")
            for b in api['body']:
                output.append(f"  • {b[1]} ({b[2]})")
                if len(b) > 6 and b[6]:
                    output.append(f"    Description: {b[6]}")
        
        # Response Structure
        if api.get('response'):
            output.append(f"\nRESPONSE STRUCTURE:")
            for r in api['response']:
                if r[1]:
                    output.append(f"  • {r[1]} ({r[2] if len(r) > 2 else 'Object'})")
                    if len(r) > 6 and r[6]:
                        output.append(f"    Description: {r[6]}")
        
        # Business Rules
        if api.get('businessRules'):
            output.append(f"\nBUSINESS RULES:")
            for rule in api['businessRules']:
                if rule and rule not in ['None', '']:
                    output.append(f"  • {rule}")
        
        # Error Codes
        if api.get('errorCodes'):
            output.append(f"\nERROR CODES:")
            for code, desc in api['errorCodes']:
                if code and code not in ['None', '']:
                    output.append(f"  • {code}: {desc if desc else 'N/A'}")
        
        # Validations
        if api.get('validations'):
            output.append(f"\nVALIDATIONS:")
            for val in api['validations']:
                if val and val not in ['None', '']:
                    output.append(f"  • {val}")
        
        # Sample Request
        if api.get('requestSample'):
            output.append(f"\nREQUEST SAMPLE URL:")
            output.append(f"  {api['requestSample']}")
        
    except Exception as e:
        output.append(f"\nError processing {sheet_name}: {str(e)}")

# Write to file and print
output_text = "\n".join(output)
with open('api_specifications_extracted.txt', 'w', encoding='utf-8') as f:
    f.write(output_text)

print(output_text)

