import openpyxl
from openpyxl.utils import get_column_letter
import json

# Load the workbook
wb = openpyxl.load_workbook(r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx")

def extract_sheet_data(sheet):
    """Extract all data from a sheet as a structured format"""
    data = []
    for row in sheet.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            value = cell.value
            if value is None:
                row_data.append(None)
            else:
                row_data.append(str(value))
        data.append(row_data)
    return data

def print_sheet_content(sheet_name, max_rows=30):
    """Print formatted content of a sheet"""
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    data = extract_sheet_data(ws)
    
    for i, row in enumerate(data[:max_rows], 1):
        # Filter out None values for cleaner display
        row_display = [f"{j}:{v}" if v else f"{j}:None" for j, v in enumerate(row) if j < 10]
        if any(v for v in row[:10]):
            print(f"Row {i}: {row_display}")

# Print README sheet first to understand document structure
print_sheet_content("README", max_rows=50)

# Print API_Specs_Index to understand the API list
print_sheet_content("API_Specs_Index", max_rows=50)

# Print key API sheets
api_sheets = [s for s in wb.sheetnames if s.startswith('get') or s.startswith('set') or s.startswith('block') or s.startswith('reactivate')]
print(f"\n\nFound {len(api_sheets)} API sheets: {api_sheets}")

# Print one sample API sheet
if api_sheets:
    print_sheet_content(api_sheets[0], max_rows=50)

