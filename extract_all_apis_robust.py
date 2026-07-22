#!/usr/bin/env python3
"""Comprehensive API extraction handling all format variations"""

import openpyxl
import json
from typing import Dict, List

class RobustAPIExtractor:
    def __init__(self, excel_file):
        self.wb = openpyxl.load_workbook(excel_file, data_only=True)
        self.apis = []
        
    def extract_all_apis(self) -> List[Dict]:
        """Extract all API specifications"""
        api_sheets = [s for s in self.wb.sheetnames if s not in ['README', 'INDEX']]
        
        print(f"Processing {len(api_sheets)} API sheets...\n")
        
        for sheet_name in api_sheets:
            api_spec = self.extract_api_spec(sheet_name)
            if api_spec and api_spec.get('service'):
                self.apis.append(api_spec)
                print(f"  ✓ {api_spec['service']:40s} | {api_spec['method']:6s} | {api_spec['url']}")
        
        return self.apis
    
    def extract_api_spec(self, sheet_name: str) -> Dict:
        """Extract API specification from a single sheet"""
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        spec = {
            'sheet_name': sheet_name,
            'service': None,
            'description': None,
            'business_rule': None,
            'method': None,
            'url': None,
            'message_type': None,
            'headers': [],
            'body_params': [],
            'query_params': [],
            'response_fields': [],
        }
        
        # Extract basic info from first rows
        for i, row in enumerate(rows[:20]):
            if not any(row):
                continue
            
            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            # Handle various naming conventions
            if col_a in ['Service', 'API Name']:
                spec['service'] = col_b.strip()
            elif col_a == 'Description':
                spec['description'] = col_b
            elif col_a in ['Business Rule', 'IBO Rule']:
                spec['business_rule'] = col_b
            elif col_a == 'Method':
                spec['method'] = col_b
            elif col_a == 'URL':
                spec['url'] = col_b.replace('https: ', 'https://').replace('http: ', 'http://')
            elif col_a == 'Message Type':
                spec['message_type'] = col_b
        
        # Extract parameters
        for i, row in enumerate(rows):
            if not any(row):
                continue
            
            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            # HTTP Headers section
            if col_a == 'Request' and col_b == 'HTTP Header':
                self._extract_parameters(rows, i + 2, spec, 'headers')
            
            # HTTP Body section
            elif col_a is not None and 'HTTP Body' in str(col_a):
                self._extract_parameters(rows, i + 2, spec, 'body_params')
            
            # Request Parameters section
            elif col_a is not None and 'Request Parameter' in str(col_a):
                self._extract_parameters(rows, i + 2, spec, 'query_params')
            
            # Response section
            elif col_a == 'Response' or (col_a is not None and 'Response' in str(col_a) and col_b == 'Name'):
                self._extract_parameters(rows, i + 1, spec, 'response_fields')
        
        return spec
    
    def _extract_parameters(self, rows: list, start_idx: int, spec: dict, param_type: str):
        """Extract parameter list from a section"""
        for i in range(start_idx, min(start_idx + 50, len(rows))):
            row = rows[i]
            
            if not any(row):
                continue
            
            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            # Stop if we hit another section
            if col_a in ['HTTP Body', 'Request Parameter', 'Response', 'Request Sample', 'Response Sample']:
                break
            
            if col_a and 'Body' in col_a or col_a and 'Parameter' in col_a or col_a and 'Sample' in col_a:
                break
            
            # Check if this is a header row
            if col_b == 'Name' and len(row) >= 7:
                continue
            
            # Extract parameter if it has a name
            if col_b and col_b != 'Name' and col_a not in ['Request', 'Response', 'Name', '']:
                param = {
                    'name': col_b,
                    'type': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    'length': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                    'mandatory': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                    'sample': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                    'description': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                }
                spec[param_type].append(param)

# Main execution
if __name__ == '__main__':
    print("=" * 150)
    print("EXTRACTING ALL API SPECIFICATIONS FROM DDD FILE")
    print("=" * 150)
    
    extractor = RobustAPIExtractor(r'input\api\DEP_App_Dashboard_DDD_API_Design_v1.xlsx')
    apis = extractor.extract_all_apis()
    
    print(f"\n{'='*150}")
    print(f"✓ Successfully extracted {len(apis)} APIs")
    print(f"{'='*150}\n")
    
    # Save to JSON
    with open('all_apis_complete.json', 'w') as f:
        json.dump({
            'total_apis': len(apis),
            'apis': apis
        }, f, indent=2)
    
    print(f"✓ API metadata saved to all_apis_complete.json\n")
    
    # Statistics
    print("STATISTICS:")
    print(f"  Average headers per API: {sum(len(a['headers']) for a in apis) / len(apis):.1f}")
    print(f"  Average body params per API: {sum(len(a['body_params']) for a in apis) / len(apis):.1f}")
    print(f"  Average query params per API: {sum(len(a['query_params']) for a in apis) / len(apis):.1f}")
    print(f"  Average response fields per API: {sum(len(a['response_fields']) for a in apis) / len(apis):.1f}")
