#!/usr/bin/env python3
"""Format all test cases into Markdown and Excel outputs"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

# Load test cases
with open('all_test_cases_complete.json', 'r') as f:
    data = json.load(f)

test_cases = data['test_cases']
apis = list({tc['service']: tc for tc in test_cases}.keys())

print("=" * 150)
print("FORMATTING TEST CASES INTO MARKDOWN AND EXCEL")
print("=" * 150)

# Ensure output directories exist
markdown_dir = 'artifacts\\Test_Case\\Markdown'
excel_dir = 'artifacts\\Test_Case\\Excel'
os.makedirs(markdown_dir, exist_ok=True)
os.makedirs(excel_dir, exist_ok=True)

# ===== MARKDOWN OUTPUT =====
print("\nGenerating Markdown file...")

markdown_output = """# DEP App Dashboard API Validation Test Cases

**Reference Document:** DEP_App_Dashboard_DDD_API_Design_v1.xlsx  
**Generated Date:** """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """  
**Total APIs Covered:** """ + str(len(apis)) + """  
**Total Test Cases:** """ + str(len(test_cases)) + """  

---

## 1. TEST SUMMARY

### Overview Statistics
| Metric | Count |
|--------|-------|
| **Total APIs** | """ + str(len(apis)) + """ |
| **Total Test Cases** | """ + str(len(test_cases)) + """ |
| **Positive Test Cases** | """ + str(len([tc for tc in test_cases if tc['type'] == 'Positive'])) + """ |
| **Negative Test Cases** | """ + str(len([tc for tc in test_cases if tc['type'] == 'Negative'])) + """ |

### APIs Identified
| # | API Name | HTTP Method | Endpoint |
|---|---|---|---|
"""

apis_info = {}
for tc in test_cases:
    if tc['service'] not in apis_info:
        apis_info[tc['service']] = {'method': tc['method'], 'endpoint': tc['endpoint']}

for i, (api_name, info) in enumerate(apis_info.items(), 1):
    markdown_output += f"| {i} | {api_name} | {info['method']} | {info['endpoint']} |\n"

markdown_output += """
### Key Validation Areas Covered
- **Happy Path / Valid Request** - Valid requests with all mandatory fields
- **Authentication / Missing Header** - Missing Authorization header
- **Authentication / Invalid Token** - Invalid or expired authentication tokens
- **HTTP Protocol / Invalid Method** - Incorrect HTTP method usage
- **Format / Malformed Payload** - Malformed JSON in request body
- **Parameter Validation / Missing Field** - Missing mandatory query parameters
- **Data Validation / Type Mismatch** - Invalid data types in requests
- **Security / SQL Injection** - SQL injection payload handling
- **Security / XSS Attack** - XSS payload handling and prevention
- **Input Validation / Special Characters** - Special character handling
- **Boundary Testing / Range Validation** - Out-of-range and boundary violations

### Assumptions Made
1. **Authentication:** Tests assume Bearer token format for Authorization header
2. **Error Handling:** Standard HTTP status codes (400, 401, 405) with descriptive error messages
3. **JSON Format:** All API communications use application/json content type
4. **Input Sanitization:** API implements input sanitization for security testing
5. **Token Validity:** Expired tokens have clear TTL validation mechanisms

---

## 2. TEST CASES BY API

"""

# Group test cases by API
test_cases_by_api = {}
for tc in test_cases:
    api_name = tc['service']
    if api_name not in test_cases_by_api:
        test_cases_by_api[api_name] = []
    test_cases_by_api[api_name].append(tc)

# Generate markdown for each API
for api_name in sorted(test_cases_by_api.keys()):
    api_tcs = test_cases_by_api[api_name]
    method = api_tcs[0]['method']
    endpoint = api_tcs[0]['endpoint']
    
    markdown_output += f"""### {api_name}
**Endpoint:** `{method} {endpoint}`  
**Test Cases Count:** {len(api_tcs)}  

#### Test Case Summary
| Test ID | Title | Type | Category | Priority |
|---------|-------|------|----------|----------|
"""
    
    for tc in api_tcs:
        markdown_output += f"| {tc['id']} | {tc['title']} | {tc['type']} | {tc['category']} | {tc['priority']} |\n"
    
    markdown_output += "\n#### Detailed Test Cases\n\n"
    
    for tc in api_tcs:
        markdown_output += f"""**Test ID:** {tc['id']}  
**Title:** {tc['title']}  
**Type:** {tc['type']}  
**Category:** {tc['category']}  
**Priority:** {tc['priority']}  
**Endpoint:** `{tc['method']} {tc['endpoint']}`  

**Pre-conditions:**  
{tc['preconditions']}

**Test Steps:**  
{tc['test_steps']}

**Input Data:**  
```
{tc['input_data']}
```

**Expected Result:**  
{tc['expected_result']}

---

"""

# Save markdown
markdown_file = os.path.join(markdown_dir, 'DEP_App_Dashboard_DDD_API_Design_v1_All_Test_Cases.md')
with open(markdown_file, 'w', encoding='utf-8') as f:
    f.write(markdown_output)

print(f"✓ Markdown file: {markdown_file}")

# ===== EXCEL OUTPUT =====
print("Generating Excel file...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Summary sheet
ws_summary = wb.create_sheet("Summary", 0)
ws_summary['A1'] = "DEP APP DASHBOARD API VALIDATION TEST CASES"
ws_summary['A1'].font = Font(bold=True, size=14)

ws_summary['A3'] = "Generated Date:"
ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

ws_summary['A5'] = "Total APIs:"
ws_summary['B5'] = len(apis)
ws_summary['B5'].font = Font(bold=True)

ws_summary['A6'] = "Total Test Cases:"
ws_summary['B6'] = len(test_cases)
ws_summary['B6'].font = Font(bold=True)

ws_summary['A7'] = "Positive Test Cases:"
ws_summary['B7'] = len([tc for tc in test_cases if tc['type'] == 'Positive'])

ws_summary['A8'] = "Negative Test Cases:"
ws_summary['B8'] = len([tc for tc in test_cases if tc['type'] == 'Negative'])

ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 20

# Validation Areas sheet
ws_validation = wb.create_sheet("Validation Areas")
ws_validation['A1'] = "KEY VALIDATION AREAS COVERED"
ws_validation['A1'].font = Font(bold=True, size=12)

areas = [
    "Happy Path / Valid Request",
    "Authentication / Missing Header",
    "Authentication / Invalid Token",
    "HTTP Protocol / Invalid Method",
    "Format / Malformed Payload",
    "Format / Empty Payload",
    "Parameter Validation / Missing Field",
    "Data Validation / Type Mismatch",
    "Security / SQL Injection",
    "Security / XSS Attack",
    "Input Validation / Special Characters",
    "Boundary Testing / Range Validation"
]

for i, area in enumerate(areas, 2):
    ws_validation[f'A{i}'] = area
    ws_validation.row_dimensions[i].height = 20

ws_validation.column_dimensions['A'].width = 50

# Test cases sheets - one per API
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for api_name in sorted(test_cases_by_api.keys()):
    # Shorten sheet name to fit Excel's 31 character limit
    sheet_name = api_name[:31]
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ['Test ID', 'Title', 'Type', 'Category', 'Priority', 'Method', 'Endpoint', 'Pre-conditions', 'Test Steps', 'Input Data', 'Expected Result']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add test cases for this API
    api_tcs = test_cases_by_api[api_name]
    for row_idx, tc in enumerate(api_tcs, 2):
        ws.cell(row=row_idx, column=1).value = tc['id']
        ws.cell(row=row_idx, column=2).value = tc['title']
        ws.cell(row=row_idx, column=3).value = tc['type']
        ws.cell(row=row_idx, column=4).value = tc['category']
        ws.cell(row=row_idx, column=5).value = tc['priority']
        ws.cell(row=row_idx, column=6).value = tc['method']
        ws.cell(row=row_idx, column=7).value = tc['endpoint']
        ws.cell(row=row_idx, column=8).value = tc['preconditions']
        ws.cell(row=row_idx, column=9).value = tc['test_steps']
        ws.cell(row=row_idx, column=10).value = tc['input_data']
        ws.cell(row=row_idx, column=11).value = tc['expected_result']
        
        # Apply styling
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 40
    
    # Freeze header row
    ws.freeze_panes = 'A2'

# Save Excel file
excel_file = os.path.join(excel_dir, 'DEP_App_Dashboard_DDD_API_Design_v1_All_Test_Cases.xlsx')
wb.save(excel_file)

print(f"✓ Excel file: {excel_file}")

print("\n" + "=" * 150)
print("COMPLETE TEST CASE SUITE GENERATED")
print("=" * 150)
print(f"\nDeliverables:")
print(f"  ✓ Markdown: {markdown_file}")
print(f"  ✓ Excel: {excel_file}")
print(f"\nStatistics:")
print(f"  APIs: {len(apis)}")
print(f"  Total Test Cases: {len(test_cases)}")
print(f"  Positive Cases: {len([tc for tc in test_cases if tc['type'] == 'Positive'])}")
print(f"  Negative Cases: {len([tc for tc in test_cases if tc['type'] == 'Negative'])}")
