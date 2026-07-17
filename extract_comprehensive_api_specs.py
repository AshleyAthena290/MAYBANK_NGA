import openpyxl
import json
from datetime import datetime

# Load the workbook
wb = openpyxl.load_workbook(r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx")

def extract_api_specs(sheet_name):
    """Extract API specifications from a sheet"""
    ws = wb[sheet_name]
    
    api_info = {
        'name': None,
        'method': None,
        'url': None,
        'messageType': None,
        'requestHeaders': [],
        'requestBody': [],
        'requestParameters': [],
        'requestSample': None,
        'responseStructure': [],
        'responseSample': None,
        'businessRules': [],
        'errorCodes': [],
        'validations': []
    }
    
    rows = list(ws.iter_rows(values_only=False))
    current_section = None
    
    for i, row in enumerate(rows):
        values = [cell.value for cell in row]
        
        if not any(v for v in values):
            continue
            
        # Extract key values
        if values[0] == 'Service' and values[1]:
            api_info['name'] = values[1]
        elif values[0] == 'Method' and values[1]:
            api_info['method'] = values[1]
        elif values[0] == 'URL' and values[1]:
            api_info['url'] = str(values[1]).replace('https: //', 'https://')
        elif values[0] == 'Message Type' and values[1]:
            api_info['messageType'] = values[1]
        
        # Identify sections
        if 'HTTP Header' in str(values[1] if len(values) > 1 else ''):
            current_section = 'headers'
        elif 'HTTP Body' in str(values[1] if len(values) > 1 else ''):
            current_section = 'body'
        elif 'Request Parameter' in str(values[0] if len(values) > 0 else ''):
            current_section = 'parameters'
        elif 'Request Sample' in str(values[0] if len(values) > 0 else ''):
            current_section = 'request_sample'
        elif values[0] == 'Response' and values[1] == 'Name':
            current_section = 'response'
        elif 'Response Sample' in str(values[0] if len(values) > 0 else ''):
            current_section = 'response_sample'
        elif 'Business Rule' in str(values[0] if len(values) > 0 else ''):
            current_section = 'business_rules'
        elif 'Error Code' in str(values[0] if len(values) > 0 else ''):
            current_section = 'error_codes'
        elif 'Validation' in str(values[0] if len(values) > 0 else ''):
            current_section = 'validations'
        
        # Skip header rows
        if 'Name' in str(values[1]) and 'Type' in str(values[2]):
            continue
        
        # Extract parameter/header data
        if current_section in ['headers', 'body', 'parameters', 'response']:
            if values[1] and values[1] not in ['Name', 'Type', 'Length']:
                param = {
                    'name': values[1],
                    'type': values[2] if len(values) > 2 else None,
                    'length': values[3] if len(values) > 3 else None,
                    'mandatory': values[4] if len(values) > 4 else None,
                    'sample': values[5] if len(values) > 5 else None,
                    'description': values[6] if len(values) > 6 else None,
                }
                
                if current_section == 'headers':
                    api_info['requestHeaders'].append(param)
                elif current_section == 'body':
                    api_info['requestBody'].append(param)
                elif current_section == 'parameters':
                    api_info['requestParameters'].append(param)
                elif current_section == 'response':
                    api_info['responseStructure'].append(param)
        
        # Extract samples
        elif current_section == 'request_sample' and values[1]:
            api_info['requestSample'] = str(values[1])
        elif current_section == 'response_sample' and values[1]:
            api_info['responseSample'] = str(values[1])[:500]  # Limit for readability
        
        # Extract business rules, errors, validations
        elif current_section in ['business_rules', 'error_codes', 'validations']:
            if values[1]:
                if current_section == 'business_rules':
                    api_info['businessRules'].append(str(values[1]))
                elif current_section == 'error_codes':
                    api_info['errorCodes'].append({
                        'code': values[1],
                        'description': values[2] if len(values) > 2 else None
                    })
                elif current_section == 'validations':
                    api_info['validations'].append(str(values[1]))
    
    return api_info

# Extract all APIs from API_Specs_Index first
def get_api_summary():
    """Get summary from API_Specs_Index sheet"""
    ws = wb['API_Specs_Index']
    apis = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # API Name column
            apis.append({
                'api_name': row[2],
                'method': row[3],
                'endpoint': row[5],
                'microservice': row[4],
                'sheet_name': row[1]
            })
    
    return apis

# Main extraction
print("=" * 100)
print("API SPECIFICATIONS EXTRACTION REPORT")
print("=" * 100)
print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print()

# Get summary
api_summary = get_api_summary()
print(f"Found {len(set(a['api_name'] for a in api_summary if a['api_name']))} unique APIs\n")

# Extract detailed specs for known API sheets
api_sheets = {s: s for s in wb.sheetnames 
              if any(x in s for x in ['get', 'set', 'block', 'reactivate', 'link', 'Auto', 'Remove'])}

print("=" * 100)
print("DETAILED API SPECIFICATIONS")
print("=" * 100)

for sheet_name in sorted(api_sheets.keys()):
    print(f"\n{'='*100}")
    print(f"API SHEET: {sheet_name}")
    print(f"{'='*100}")
    
    try:
        api = extract_api_specs(sheet_name)
        
        print(f"\nService Name: {api['name']}")
        print(f"HTTP Method: {api['method']}")
        print(f"URL: {api['url']}")
        print(f"Message Type: {api['messageType']}")
        
        # Request Headers
        if api['requestHeaders']:
            print(f"\n--- REQUEST HEADERS ---")
            for h in api['requestHeaders']:
                mandatory = "✓ MANDATORY" if h['mandatory'] == 'Y' else "(optional)"
                print(f"  • {h['name']} ({h['type']}) {mandatory}")
                if h['description']:
                    print(f"    Description: {h['description']}")
                if h['sample']:
                    print(f"    Sample: {h['sample']}")
        
        # Request Parameters
        if api['requestParameters']:
            print(f"\n--- REQUEST PARAMETERS ---")
            for p in api['requestParameters']:
                mandatory = "✓ MANDATORY" if p['mandatory'] == 'Y' else "(optional)"
                print(f"  • {p['name']} ({p['type']}) {mandatory}")
                if p['description']:
                    print(f"    Description: {p['description']}")
                if p['sample']:
                    print(f"    Sample: {p['sample']}")
        
        # Request Body
        if api['requestBody']:
            print(f"\n--- REQUEST BODY ---")
            for b in api['requestBody']:
                print(f"  • {b['name']} ({b['type']})")
                if b['description']:
                    print(f"    Description: {b['description']}")
        
        # Response Structure
        if api['responseStructure']:
            print(f"\n--- RESPONSE STRUCTURE ---")
            for r in api['responseStructure']:
                print(f"  • {r['name']} ({r['type']})")
                if r['sample']:
                    print(f"    Sample: {r['sample']}")
                if r['description']:
                    print(f"    Description: {r['description']}")
        
        # Business Rules
        if api['businessRules']:
            print(f"\n--- BUSINESS RULES ---")
            for rule in api['businessRules']:
                print(f"  • {rule}")
        
        # Error Codes
        if api['errorCodes']:
            print(f"\n--- ERROR CODES ---")
            for err in api['errorCodes']:
                print(f"  • {err['code']}: {err['description']}")
        
        # Validations
        if api['validations']:
            print(f"\n--- VALIDATIONS ---")
            for val in api['validations']:
                print(f"  • {val}")
        
        # Sample URLs
        if api['requestSample']:
            print(f"\n--- REQUEST SAMPLE ---")
            print(f"  {api['requestSample']}")
            
    except Exception as e:
        print(f"Error extracting {sheet_name}: {str(e)}")

print("\n\n" + "="*100)
print("API SUMMARY TABLE")
print("="*100)

print(f"\n{'No.':<5} {'API Name':<40} {'Method':<10} {'Endpoint':<60}")
print("-" * 120)

for i, api in enumerate(api_summary, 1):
    if api['api_name']:
        endpoint = str(api['endpoint'])[:57] + "..." if len(str(api['endpoint'])) > 60 else str(api['endpoint'])
        print(f"{i:<5} {str(api['api_name']):<40} {str(api['method']):<10} {endpoint:<60}")

