#!/usr/bin/env python3
"""Create the Finance and Non-Financing Dashboard API analysis deliverables."""

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import re
from zipfile import BadZipFile

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "input" / "Journey_Test_cases" / "Finance and Non Financing Dashboard.xlsx"
OUTPUT_XLSX = ROOT / "artifacts" / "Finance_and_Non_Financing_Dashboard_API_Test_Analysis.xlsx"
OUTPUT_MD = ROOT / "artifacts" / "Finance_and_Non_Financing_Dashboard_API_Test_Analysis.md"

API_FILES = {
    "Credit Card": "ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx",
    "Debit Card": "ECLIPSE_Account Dashboard_Debit_Card_DDD_API_Design_v1_Workshop (1).xlsx",
    "Insurance": "ECLIPSE_Account Dashboard_Insurance_DDD_API_Design_v1_Workshop.xlsx",
    "Loan": "ECLIPSE_Account Dashboard_Loan_DDD_API_Design_v1_Workshop.xlsx",
    "CASA": "ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx",
    "Maintenance or shared services": "ECLIPSE_Maintenance_DDD_API_Design_v1.xlsx",
}

TRACE_HEADERS = [
    "Worksheet", "Source Row Number", "Test Case ID", "Journey or Feature", "Test Scenario",
    "Dashboard Domain", "API Design Document", "API Name", "HTTP Method", "Endpoint",
    "Request Inputs", "Expected Status Code", "Expected Response or Dashboard Behaviour",
    "Mapping Status", "Coverage Type", "Finding or Gap", "Recommendation", "Confidence",
    "Source Reference",
]


def text(value):
    return "" if value is None else str(value).strip()


def compact(value, limit=700):
    value = re.sub(r"\s+", " ", text(value))
    return value if len(value) <= limit else value[:limit - 3] + "..."


def domain_for(row):
    haystack = " ".join(text(v).lower() for v in row)
    if "credit card" in haystack or "credit-card" in haystack:
        return "Credit Card"
    if "debit card" in haystack or "debit-card" in haystack:
        return "Debit Card"
    if "insurance" in haystack or "takaful" in haystack:
        return "Insurance"
    if any(word in haystack for word in ("loan", "financing", "mortgage")):
        return "Loan"
    if any(word in haystack for word in ("casa", "savings", "current account", "deposit")):
        return "CASA"
    return "Maintenance or shared services"


def coverage_type(value):
    value = text(value).lower()
    if "security" in value or "auth" in value:
        return "Security"
    if "resilien" in value or any(x in value for x in ("timeout", "unavailable", "failure")):
        return "Resilience"
    if "boundary" in value or "limit" in value:
        return "Boundary"
    if "negative" in value or any(x in value for x in ("invalid", "missing", "reject", "error")):
        return "Negative"
    if "positive" in value or "open" in value or "validate" in value:
        return "Positive"
    return "Other"


def api_operations():
    operations = []
    unreadable = []
    for domain, filename in API_FILES.items():
        path = ROOT / "input" / "api" / filename
        if not path.exists():
            unreadable.append((domain, filename, "File not found"))
            continue
        try:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except (BadZipFile, OSError, ValueError) as error:
            unreadable.append((domain, filename, f"Could not open as XLSX: {error}"))
            continue
        if "API_Specs_Index" not in workbook.sheetnames:
            unreadable.append((domain, filename, "API_Specs_Index sheet not found"))
            continue
        ws = workbook["API_Specs_Index"]
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            values = [text(v) for v in row]
            method = values[3] if len(values) > 3 else ""
            endpoint = values[5] if len(values) > 5 else ""
            name = values[2] if len(values) > 2 else ""
            sheet = values[1] if len(values) > 1 else ""
            if not method or not endpoint:
                continue
            operations.append({
                "domain": domain, "document": filename, "name": name or "Not specified",
                "method": method.upper(), "endpoint": endpoint, "sheet": sheet or "Not specified",
                "index_row": row_number,
            })
    return operations, unreadable


def choose_operation(row, operations, domain):
    haystack = " ".join(text(v).lower() for v in row)
    candidates = [op for op in operations if op["domain"] == domain]
    scored = []
    for op in candidates:
        terms = set(re.findall(r"[a-z0-9]+", (op["name"] + " " + op["sheet"] + " " + op["endpoint"]).lower()))
        score = sum(1 for term in terms if len(term) > 3 and term in haystack)
        if "dashboard" in op["name"].lower() and any(x in haystack for x in ("dashboard", "tab", "landing")):
            score += 2
        scored.append((score, op))
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0] if scored and scored[0][0] >= 2 else None


def expected_status(row):
    found = re.findall(r"\b([1-5]\d{2})\b", " ".join(text(v) for v in row))
    return ", ".join(dict.fromkeys(found)) or "Not specified"


def source_cases():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    ws = workbook["Test Cases"]
    headers = [text(v) for v in next(ws.iter_rows(values_only=True))]
    cases = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        values = [text(v) for v in row]
        if not any(values):
            continue
        data = dict(zip(headers, values))
        data["_row"] = row_number
        cases.append(data)
    return cases


def build_traceability(cases, operations):
    rows = []
    for case in cases:
        domain = domain_for(case.values())
        selected = choose_operation(case.values(), operations, domain)
        operation, score = (selected[1], selected[0]) if selected else (None, 0)
        expected = case.get("Expected Result", "")
        source_ref = f"Test Cases!A{case['_row']}:R{case['_row']}"
        if operation:
            status = "Mapped" if score >= 4 else "Partially Mapped"
            finding = ("Documented API operation is a keyword-supported candidate, but request/response assertions are not fully specified in the source test case."
                       if status == "Mapped" else "A generic or weak keyword match suggests an API candidate; the source case does not uniquely evidence this operation.")
            confidence = "Medium" if status == "Mapped" else "Low"
            api_doc, api_name, method, endpoint = operation["document"], operation["name"], operation["method"], operation["endpoint"]
        else:
            status = "Partially Mapped" if domain in API_FILES else "Not Mapped"
            finding = "No uniquely evidenced API operation in the source test case; endpoint and API assertions are not specified." 
            confidence = "Low"
            api_doc = API_FILES.get(domain, "Not specified")
            api_name = method = endpoint = "Not specified"
        rows.append({
            "Worksheet": "Test Cases", "Source Row Number": case["_row"],
            "Test Case ID": case.get("Test Case ID", "Not specified"),
            "Journey or Feature": case.get("Journey", "Not specified"),
            "Test Scenario": case.get("Test Case Title", "Not specified"),
            "Dashboard Domain": domain, "API Design Document": api_doc, "API Name": api_name,
            "HTTP Method": method, "Endpoint": endpoint,
            "Request Inputs": compact("Test data: " + case.get("Test Data / Variants", "Not specified") + "; Steps: " + case.get("Test Steps", "Not specified")),
            "Expected Status Code": expected_status([case.get("Expected Result"), case.get("Test Steps")]),
            "Expected Response or Dashboard Behaviour": compact(expected or "Not specified"),
            "Mapping Status": status, "Coverage Type": coverage_type(case.get("Test Type")),
            "Finding or Gap": finding,
            "Recommendation": "Add API-level request, response, status, and failure assertions using the documented operation." if operation else "Confirm the owning API and add an API-level mapping before implementation.",
            "Confidence": confidence, "Source Reference": source_ref,
        })
    return rows


def build_api_coverage(operations, trace):
    by_op = defaultdict(list)
    for row in trace:
        key = (row["API Design Document"], row["API Name"], row["HTTP Method"], row["Endpoint"])
        if row["API Name"] != "Not specified":
            by_op[key].append(row)
    result = []
    for op in operations:
        key = (op["document"], op["name"], op["method"], op["endpoint"])
        linked = by_op.get(key, [])
        types = Counter(r["Coverage Type"] for r in linked)
        missing = [label for label, kind in (("positive", "Positive"), ("negative", "Negative"), ("boundary", "Boundary"), ("security", "Security"), ("resilience", "Resilience")) if not types[kind]]
        result.append({
            "API design document": op["document"], "API or operation name": op["name"],
            "HTTP method": op["method"], "Endpoint": op["endpoint"],
            "Covered by test case IDs": ", ".join(r["Test Case ID"] for r in linked) or "None",
            "Positive coverage": "Yes" if types["Positive"] else "No", "Negative coverage": "Yes" if types["Negative"] else "No",
            "Boundary coverage": "Yes" if types["Boundary"] else "No", "Security coverage": "Yes" if types["Security"] else "No",
            "Resilience coverage": "Yes" if types["Resilience"] else "No",
            "Coverage status": "Covered" if linked and not missing else ("Partially covered" if linked else "Not covered"),
            "Missing scenarios": ", ".join(missing) or "None identified from category labels",
        })
    return result


def build_gaps(trace, coverage, unreadable):
    gaps = []
    for domain, filename, reason in unreadable:
        gaps.append({"Domain": domain, "Severity": "High", "Gap": "Requested API design document could not be reviewed.", "Impact": "Operations and contract requirements for this domain remain unverified.", "Supporting Evidence": f"{filename}: {reason}", "Recommended Action": "Provide a valid readable source workbook and rerun the analysis."})
    for domain in API_FILES:
        domain_rows = [r for r in trace if r["Dashboard Domain"] == domain]
        if not domain_rows:
            gaps.append({"Domain": domain, "Severity": "High", "Gap": "No source test cases were identified for this domain.", "Impact": "Documented domain APIs cannot be validated from the supplied workbook.", "Supporting Evidence": "No Test Cases rows classified to this domain.", "Recommended Action": "Confirm scope and add domain journey coverage."})
            continue
        if not any(r["API Name"] != "Not specified" for r in domain_rows):
            gaps.append({"Domain": domain, "Severity": "High", "Gap": "Domain test cases do not identify a documented API operation.", "Impact": "API contract and integration behaviour cannot be traced.", "Supporting Evidence": f"{len(domain_rows)} source rows classified to {domain}; no unique API candidate.", "Recommended Action": "Confirm API ownership and map each journey to an indexed operation."})
    for row in coverage:
        if row["Coverage status"] != "Covered":
            gaps.append({"Domain": next((d for d, f in API_FILES.items() if f == row["API design document"]), "Cross-cutting concerns"), "Severity": "High" if row["Coverage status"] == "Not covered" else "Medium", "Gap": f"{row['API or operation name']} lacks complete category coverage.", "Impact": "One or more documented API failure or control conditions are not represented by source test cases.", "Supporting Evidence": f"Endpoint {row['Endpoint']}; missing: {row['Missing scenarios']}.", "Recommended Action": "Add focused API contract tests for the missing scenarios, using only documented fields and status codes."})
    cross = [r for r in trace if r["Coverage Type"] in ("Security", "Resilience")]
    if not any(r["Coverage Type"] == "Resilience" for r in trace):
        gaps.append({"Domain": "Cross-cutting concerns", "Severity": "High", "Gap": "No resilience test category is present in the source test cases.", "Impact": "Timeout, unavailable service, and dependency failure behaviour is not validated.", "Supporting Evidence": f"{len(trace)} source test cases reviewed; no Resilience-labelled case.", "Recommended Action": "Add timeout, 5xx, unavailable-service, and partial-response tests only where the API contract defines expected behaviour."})
    return gaps


def build_proposals(gaps, coverage):
    proposals = []
    for index, gap in enumerate(gaps, 1):
        matching = next((r for r in coverage if r["API design document"] == API_FILES.get(gap["Domain"])), None)
        proposals.append({
            "Proposed Test Case ID": f"PROP-FND-{index:03d}", "Dashboard Domain": gap["Domain"], "Journey or Feature": "API contract coverage gap",
            "Test Objective": gap["Gap"], "Preconditions": "Use a valid source-supported test profile and documented authentication prerequisites; otherwise Not specified.",
            "API and Endpoint": (matching["API or operation name"] + " | " + matching["Endpoint"]) if matching else "Not specified",
            "HTTP Method": matching["HTTP method"] if matching else "Not specified", "Request Headers": "Not specified",
            "Request Parameters or Body": "Not specified", "Test Steps": "Prepare the documented request; inject the gap condition; invoke the documented operation; verify the documented response and dashboard behaviour.",
            "Expected HTTP Status": "Not specified unless documented in the API design workbook", "Expected API Response": "Not specified unless documented in the API design workbook",
            "Expected Dashboard Behaviour": "The dashboard handles the documented condition without exposing incorrect or unverified data.", "Test Data Requirements": "Source-supported valid and invalid values for the affected operation; exact values Not specified.",
            "Priority": gap["Severity"], "Requirement or Source Reference": gap["Supporting Evidence"],
        })
    return proposals


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in ws.columns:
        letter = column[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(text(c.value)) for c in column) + 2, 12), 45)


def write_workbook(trace, coverage, gaps, proposals, cases, operations):
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Executive Summary"
    summary_rows = [
        ["Metric", "Value"], ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source test cases reviewed", len(cases)], ["Mapped", sum(r["Mapping Status"] == "Mapped" for r in trace)],
        ["Partially Mapped", sum(r["Mapping Status"] == "Partially Mapped" for r in trace)], ["Not Mapped", sum(r["Mapping Status"] == "Not Mapped" for r in trace)],
        ["Documented API operations reviewed", len(operations)], ["Coverage gaps recorded", len(gaps)],
        ["High-risk gaps (Critical/High)", sum(g["Severity"] in ("Critical", "High") for g in gaps)],
        ["Source workbook", str(SOURCE.relative_to(ROOT))], ["Method note", "Mappings are evidence-based keyword candidates; unverified API details are Not specified."],
    ]
    for row in summary_rows: summary.append(row)
    sheets = [("Traceability Matrix", TRACE_HEADERS, trace), ("API Coverage", list(coverage[0]) if coverage else [], coverage), ("Gap Analysis", list(gaps[0]) if gaps else [], gaps), ("Proposed Test Cases", list(proposals[0]) if proposals else [], proposals)]
    for name, headers, rows in sheets:
        ws = wb.create_sheet(name); ws.append(headers)
        for row in rows: ws.append([row.get(h, "") for h in headers])
    ws = wb.create_sheet("Assumptions and Questions")
    ws.append(["Type", "Item"])
    ws.append(["Assumption", "The Test Cases worksheet is the authoritative detailed source; other source sheets provide context and are not counted as additional test cases."])
    ws.append(["Assumption", "A keyword-supported API candidate is not proof of a functional mapping; confidence is retained in the traceability matrix."])
    ws.append(["Question", "Which API operation and contract should be used for source cases that describe only UI/FSD behaviour?"])
    ws.append(["Question", "What status codes and response schemas are required for timeout, unavailable-service, empty, null, and partial responses?"])
    ws.append(["Question", "Should the Common DDD workbook be treated as an additional source for shared reference data, or is Maintenance the intended source?"])
    for sheet in wb.worksheets: style_sheet(sheet)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def markdown(trace, coverage, gaps, proposals, cases, operations, unreadable):
    counts = Counter(r["Mapping Status"] for r in trace)
    lines = ["# Finance and Non-Financing Dashboard API Test Analysis", "", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "", "## Executive Summary", "", f"- Test cases reviewed: **{len(cases)}**", f"- Mapped: **{counts['Mapped']}**; partially mapped: **{counts['Partially Mapped']}**; not mapped: **{counts['Not Mapped']}**", f"- Documented API operations reviewed: **{len(operations)}**", f"- Gaps recorded: **{len(gaps)}**", "", "Mappings are evidence-based candidates. The source workbook frequently specifies dashboard behaviour without an API name, request contract, or status code; those values are recorded as `Not specified` rather than inferred.", "", "## Mapping Statistics by Domain", "", "| Domain | Cases | Mapped | Partial | Not mapped |", "|---|---:|---:|---:|---:|"]
    for domain in API_FILES:
        rows = [r for r in trace if r["Dashboard Domain"] == domain]
        lines.append(f"| {domain} | {len(rows)} | {sum(r['Mapping Status']=='Mapped' for r in rows)} | {sum(r['Mapping Status']=='Partially Mapped' for r in rows)} | {sum(r['Mapping Status']=='Not Mapped' for r in rows)} |")
    lines += ["", "## High-Risk Findings", ""]
    for gap in gaps:
        if gap["Severity"] in ("Critical", "High"):
            lines.append(f"- **{gap['Severity']} - {gap['Domain']}:** {gap['Gap']} {gap['Impact']}")
    lines += ["", "## Gap Analysis", "", "| Domain | Severity | Gap | Evidence | Recommendation |", "|---|---|---|---|---|"]
    for gap in gaps: lines.append(f"| {gap['Domain']} | {gap['Severity']} | {compact(gap['Gap'], 220)} | {compact(gap['Supporting Evidence'], 220)} | {compact(gap['Recommended Action'], 220)} |")
    lines += ["", "## Proposed Test Cases", ""]
    for proposal in proposals:
        lines += [f"### {proposal['Proposed Test Case ID']} - {proposal['Dashboard Domain']}", f"- Objective: {proposal['Test Objective']}", f"- API and endpoint: {proposal['API and Endpoint']}", f"- Method/status: {proposal['HTTP Method']} / {proposal['Expected HTTP Status']}", f"- Steps: {proposal['Test Steps']}", f"- Priority: {proposal['Priority']}", f"- Source: {proposal['Requirement or Source Reference']}", ""]
    lines += ["## Assumptions and Unresolved Questions", "", "### Assumptions", "", "- The `Test Cases` worksheet is the authoritative detailed source; other source sheets provide context and are not counted as additional test cases.", "- A keyword-supported API candidate is not proof of a functional mapping; confidence is retained in the traceability matrix.", "", "### Questions", "", "- Which API operation and contract should be used for source cases that describe only UI/FSD behaviour?", "- What status codes and response schemas are required for timeout, unavailable-service, empty, null, and partial responses?", "- Should the Common DDD workbook be treated as an additional source for shared reference data, or is Maintenance the intended source?", "", "## Source Review", "", f"- Source workbook reviewed: `{SOURCE.relative_to(ROOT)}`", f"- API workbooks requested: {len(API_FILES)}; parseable: {len(API_FILES) - len(unreadable)}; unreadable or unavailable: {len(unreadable)}", "- Original inputs were not modified."]
    OUTPUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    operations, unreadable = api_operations()
    cases = source_cases()
    trace = build_traceability(cases, operations)
    coverage = build_api_coverage(operations, trace)
    gaps = build_gaps(trace, coverage, unreadable)
    proposals = build_proposals(gaps, coverage)
    write_workbook(trace, coverage, gaps, proposals, cases, operations)
    markdown(trace, coverage, gaps, proposals, cases, operations, unreadable)
    print(f"Reviewed {len(cases)} cases; operations {len(operations)}; unreadable API docs {len(unreadable)}; mapped {sum(r['Mapping Status']=='Mapped' for r in trace)}; partial {sum(r['Mapping Status']=='Partially Mapped' for r in trace)}; not mapped {sum(r['Mapping Status']=='Not Mapped' for r in trace)}")
    print(OUTPUT_XLSX)
    print(OUTPUT_MD)


if __name__ == "__main__":
    main()