#!/usr/bin/env python3
"""
Extract ECLIPSE API Detailed Specifications
These are the detailed API specs sheets (getCreditLimits, creditCardBlock, etc.)
"""

import openpyxl
from pathlib import Path
from datetime import datetime

EXCEL_FILE = r"input\api\ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx"
OUTPUT_FILE = r"artifacts\CreditCard_ECLIPSE_API_Detailed_Specs.md"

class EclipseAPIExtractor:
    def __init__(self, excel_path: str):
        """Initialize the extractor"""
        self.wb = openpyxl.load_workbook(excel_path)
        
    @staticmethod
    def clean_value(val):
        """Clean cell value"""
        if val is None:
            return None
        return str(val).strip()
    
    def get_cell(self, ws, row: int, col: int):
        """Get cell value safely"""
        try:
            cell = ws.cell(row=row, column=col)
            return self.clean_value(cell.value) if cell.value else None
        except:
            return None
    
    def find_section_row(self, ws, marker: str, start_row: int = 1) -> int:
        """Find the row where a section starts"""
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = self.get_cell(ws, row_idx, 1)
            if cell_value and marker.lower() in cell_value.lower():
                return row_idx
        return -1
    
    def extract_table_section(self, ws, marker: str, max_rows: int = 20) -> list:
        """Extract a table section"""
        start_row = self.find_section_row(ws, marker)
        if start_row == -1:
            return []
        
        # Headers are in the row after marker
        header_row = start_row + 1
        headers = []
        for col_idx in range(2, 12):
            header = self.get_cell(ws, header_row, col_idx)
            if header:
                headers.append(header)
            else:
                break
        
        if not headers:
            return []
        
        # Data rows
        rows = []
        for row_idx in range(header_row + 1, min(header_row + max_rows, ws.max_row + 1)):
            cell_a = self.get_cell(ws, row_idx, 1)
            
            # Stop at next section or empty marker
            if cell_a and any(x.lower() in cell_a.lower() for x in 
                            ['request', 'response', 'error', 'validation', 'business', 'sample', 'auth', 'security']):
                if marker.lower() not in cell_a.lower():
                    break
            
            # Extract row data
            row_data = []
            has_data = False
            for col_idx in range(2, 2 + len(headers)):
                val = self.get_cell(ws, row_idx, col_idx)
                row_data.append(val)
                if val:
                    has_data = True
            
            if has_data:
                rows.append(dict(zip(headers, row_data)))
        
        return rows
    
    def extract_eclipse_api(self, sheet_name: str) -> dict:
        """Extract detailed API specification"""
        if sheet_name not in self.wb.sheetnames:
            return {}
        
        ws = self.wb[sheet_name]
        api = {
            'sheet_name': sheet_name,
            'service': self.get_cell(ws, 1, 2),
            'method': self.get_cell(ws, 2, 2),
            'url': self.get_cell(ws, 3, 2),
            'message_type': self.get_cell(ws, 4, 2),
            'request_headers': self.extract_table_section(ws, 'HTTP Header', 20),
            'path_variables': self.extract_table_section(ws, 'Path Variable', 10),
            'request_parameters': self.extract_table_section(ws, 'Request Parameter', 15),
            'request_body': self.extract_table_section(ws, 'HTTP Body', 20),
            'response_structure': self.extract_table_section(ws, 'Response', 20),
            'request_sample': None,
            'response_sample': None,
            'business_rules': [],
            'error_codes': [],
            'validations': [],
        }
        
        # Extract samples
        for row_idx in range(1, ws.max_row + 1):
            cell_a = self.get_cell(ws, row_idx, 1)
            cell_b = self.get_cell(ws, row_idx, 2)
            
            if cell_a == 'Request Sample':
                api['request_sample'] = cell_b
            elif cell_a == 'Response Sample':
                api['response_sample'] = cell_b
            elif cell_a and 'business rule' in cell_a.lower():
                # Extract business rules
                for check_row in range(row_idx + 1, min(row_idx + 15, ws.max_row + 1)):
                    rule = self.get_cell(ws, check_row, 2)
                    if rule:
                        api['business_rules'].append(rule)
                    else:
                        break
            elif cell_a and 'error' in cell_a.lower() and 'code' in cell_a.lower():
                # Extract error codes
                for check_row in range(row_idx + 1, min(row_idx + 15, ws.max_row + 1)):
                    code = self.get_cell(ws, check_row, 2)
                    desc = self.get_cell(ws, check_row, 3)
                    if code and desc:
                        api['error_codes'].append({'code': code, 'description': desc})
                    elif not code:
                        break
            elif cell_a and 'validation' in cell_a.lower():
                # Extract validations
                for check_row in range(row_idx + 1, min(row_idx + 15, ws.max_row + 1)):
                    val = self.get_cell(ws, check_row, 2)
                    if val:
                        api['validations'].append(val)
                    else:
                        break
        
        return api
    
    def get_eclipse_api_sheets(self) -> list:
        """Get list of ECLIPSE API sheets"""
        eclipse_keywords = ['get', 'post', 'put', 'patch', 'delete', 'credit']
        api_sheets = []
        
        for sheet in self.wb.sheetnames:
            sheet_lower = sheet.lower()
            if any(keyword in sheet_lower for keyword in eclipse_keywords) and sheet not in ['README', 'API_Specs_Index', 'ECLIPSE APIs >>']:
                # Skip the Dashboard/Transaction/Summary/Manage sheets
                if not any(x in sheet for x in ['2.1', '2.2', '2.3', '2.4', 'Dashboard', 'Transactions', 'Summary', 'Manage']):
                    api_sheets.append(sheet)
        
        return api_sheets
    
    def generate_report(self, output_file: str):
        """Generate comprehensive report"""
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        eclipse_sheets = self.get_eclipse_api_sheets()
        print(f"Found {len(eclipse_sheets)} ECLIPSE API sheets")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"# ECLIPSE Account Dashboard - Credit Card Detailed API Specifications\n")
            f.write(f"## Backend ECLIPSE API Detailed Specifications\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write(f"## Executive Summary\n---\n\n")
            f.write(f"- **Total ECLIPSE APIs:** {len(eclipse_sheets)}\n")
            f.write(f"- **These are backend ECLIPSE DDD APIs** that power the Credit Card Dashboard\n")
            f.write(f"- Each API has detailed specifications for request/response structures, validations, and error handling\n\n")
            
            # Summary table
            f.write(f"## API List\n---\n\n")
            f.write(f"| No. | API Name | Method | URL |\n")
            f.write(f"|-----|----------|--------|-----|\n")
            
            for idx, sheet in enumerate(eclipse_sheets, 1):
                api = self.extract_eclipse_api(sheet)
                if api.get('service'):
                    method = api.get('method', 'GET')
                    url = api.get('url', '')[:50]
                    if len(api.get('url', '')) > 50:
                        url += '...'
                    f.write(f"| {idx} | {api.get('service')} | {method} | {url} |\n")
            
            f.write(f"\n")
            
            # Detailed specifications
            f.write(f"## Detailed API Specifications\n---\n\n")
            
            for idx, sheet in enumerate(eclipse_sheets, 1):
                api = self.extract_eclipse_api(sheet)
                
                if not api.get('service'):
                    continue
                
                f.write(f"### {idx}. {api.get('service', 'Unknown API')}\n\n")
                
                f.write(f"#### API Overview\n")
                f.write(f"- **Sheet:** {sheet}\n")
                f.write(f"- **HTTP Method:** `{api.get('method', 'N/A')}`\n")
                f.write(f"- **URL:** `{api.get('url', 'N/A')}`\n")
                f.write(f"- **Message Type:** {api.get('message_type', 'JSON')}\n\n")
                
                # Request Headers
                if api.get('request_headers'):
                    f.write(f"#### Request Headers\n")
                    f.write(f"| Field | Type | Length | Mandatory | Sample Value | Description |\n")
                    f.write(f"|-------|------|--------|-----------|--------------|-------------|\n")
                    for header in api['request_headers']:
                        f.write(f"| {header.get('Name', '')} | {header.get('Type', '')} | {header.get('Length', '')} | {header.get('Mandatory', '')} | {header.get('Sample Value', '')} | {header.get('Description', '')} |\n")
                    f.write(f"\n")
                
                # Path Variables
                if api.get('path_variables'):
                    f.write(f"#### Path Variables\n")
                    f.write(f"| Variable | Type | Length | Mandatory | Description |\n")
                    f.write(f"|----------|------|--------|-----------|-------------|\n")
                    for var in api['path_variables']:
                        f.write(f"| {var.get('Name', '')} | {var.get('Type', '')} | {var.get('Length', '')} | {var.get('Mandatory', '')} | {var.get('Description', '')} |\n")
                    f.write(f"\n")
                
                # Request Parameters
                if api.get('request_parameters'):
                    f.write(f"#### Request Parameters\n")
                    f.write(f"| Parameter | Type | Mandatory | Description |\n")
                    f.write(f"|-----------|------|-----------|-------------|\n")
                    for param in api['request_parameters']:
                        f.write(f"| {param.get('Name', '')} | {param.get('Type', '')} | {param.get('Mandatory', '')} | {param.get('Description', '')} |\n")
                    f.write(f"\n")
                
                # Request Body
                if api.get('request_body'):
                    f.write(f"#### Request Body Structure\n")
                    f.write(f"| Field | Type | Length | Mandatory | Description |\n")
                    f.write(f"|-------|------|--------|-----------|-------------|\n")
                    for field in api['request_body']:
                        f.write(f"| {field.get('Name', '')} | {field.get('Type', '')} | {field.get('Length', '')} | {field.get('Mandatory', '')} | {field.get('Description', '')} |\n")
                    f.write(f"\n")
                
                # Response Structure
                if api.get('response_structure'):
                    f.write(f"#### Response Structure\n")
                    f.write(f"| Field | Type | Description |\n")
                    f.write(f"|-------|------|-------------|\n")
                    for field in api['response_structure']:
                        f.write(f"| {field.get('Name', '')} | {field.get('Type', '')} | {field.get('Description', '')} |\n")
                    f.write(f"\n")
                
                # Business Rules
                if api.get('business_rules'):
                    f.write(f"#### Business Rules\n")
                    for rule in api['business_rules']:
                        f.write(f"- {rule}\n")
                    f.write(f"\n")
                
                # Error Codes
                if api.get('error_codes'):
                    f.write(f"#### Error Codes\n")
                    f.write(f"| Code | Description |\n")
                    f.write(f"|------|-------------|\n")
                    for error in api['error_codes']:
                        f.write(f"| {error.get('code', '')} | {error.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Validations
                if api.get('validations'):
                    f.write(f"#### Field Validations\n")
                    for val in api['validations']:
                        f.write(f"- {val}\n")
                    f.write(f"\n")
                
                # Samples
                if api.get('request_sample'):
                    f.write(f"#### Request Sample\n")
                    f.write(f"```json\n{api['request_sample']}\n```\n\n")
                
                if api.get('response_sample'):
                    f.write(f"#### Response Sample\n")
                    f.write(f"```json\n{api['response_sample']}\n```\n\n")
                
                f.write(f"---\n\n")


def main():
    """Main execution"""
    try:
        print("=" * 80)
        print("ECLIPSE Account Dashboard - Credit Card Detailed API Extraction")
        print("=" * 80)
        print(f"Source: {EXCEL_FILE}")
        print(f"Output: {OUTPUT_FILE}")
        print("")
        
        extractor = EclipseAPIExtractor(EXCEL_FILE)
        extractor.generate_report(OUTPUT_FILE)
        
        print("")
        print("=" * 80)
        print("✓ ECLIPSE API Detailed Specifications extracted successfully")
        print("=" * 80)
        
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
