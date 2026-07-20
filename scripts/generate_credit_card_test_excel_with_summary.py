#!/usr/bin/env python3
"""
ECLIPSE Account Dashboard - Credit Card DDD API
Test Case Generation with Summary Sheet
Organized by API Subfolders from artifacts/Acc_Dashboard_CC
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# TEST CASE DATA - Organized by API Subfolder
# ============================================================================

# Define APIs from artifacts/Acc_Dashboard_CC subfolders
APIS = [
    {"id": "CC_001", "name": "getCreditLimits", "method": "GET", "count": 10},
    {"id": "CC_002", "name": "getCreditLimitConfigurations", "method": "GET", "count": 4},
    {"id": "CC_003", "name": "getInstalmentConversionSchedule", "method": "GET", "count": 5},
    {"id": "CC_004", "name": "creditCardActivation", "method": "POST", "count": 8},
    {"id": "CC_005", "name": "creditCardBlock", "method": "POST", "count": 7},
    {"id": "CC_006", "name": "creditCardUnblock", "method": "POST", "count": 5},
    {"id": "CC_007", "name": "creditCardResetPin", "method": "POST", "count": 6},
    {"id": "CC_008", "name": "creditCardPinValidation", "method": "POST", "count": 6},
    {"id": "CC_009", "name": "cvvInquiry", "method": "POST", "count": 6},
    {"id": "CC_010", "name": "replaceCreditCard", "method": "POST", "count": 6},
    {"id": "CC_011", "name": "increaseCreditLimit", "method": "POST", "count": 10},
    {"id": "CC_012", "name": "applyInstConvert", "method": "POST", "count": 9},
]

TEST_CASES_DATA = [
    # getCreditLimits (10 cases)
    {"id": "CC_001_001", "scenario": "Valid request with all parameters", "objective": "Retrieve credit limits successfully", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P0", "type": "Positive"},
    {"id": "CC_001_002", "scenario": "Missing authentication header", "objective": "Verify auth validation", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P0", "type": "Negative"},
    {"id": "CC_001_003", "scenario": "Invalid/expired Bearer token", "objective": "Test token validation", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P1", "type": "Negative"},
    {"id": "CC_001_004", "scenario": "Missing client-id header", "objective": "Verify client-id validation", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P1", "type": "Negative"},
    {"id": "CC_001_005", "scenario": "Missing env header", "objective": "Verify env header validation", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P1", "type": "Negative"},
    {"id": "CC_001_006", "scenario": "Boundary - limit amount validation", "objective": "Test min/max limits", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P1", "type": "Boundary"},
    {"id": "CC_001_007", "scenario": "Cross-customer data access", "objective": "Test authorization check", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P2", "type": "Security"},
    {"id": "CC_001_008", "scenario": "Service unavailable handling", "objective": "Verify 503 handling", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P2", "type": "Negative"},
    {"id": "CC_001_009", "scenario": "Concurrent requests", "objective": "Test concurrency", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P3", "type": "Integration"},
    {"id": "CC_001_010", "scenario": "Response time validation", "objective": "Performance baseline", "api": "getCreditLimits", "method": "GET", "endpoint": "/v1/external/cards/limits", "priority": "P3", "type": "Performance"},
    
    # getCreditLimitConfigurations (4 cases)
    {"id": "CC_002_001", "scenario": "Valid request for configurations", "objective": "Retrieve configurations", "api": "getCreditLimitConfigurations", "method": "GET", "endpoint": "/v1/external/cards/limit-configurations", "priority": "P0", "type": "Positive"},
    {"id": "CC_002_002", "scenario": "Missing authentication", "objective": "Auth validation", "api": "getCreditLimitConfigurations", "method": "GET", "endpoint": "/v1/external/cards/limit-configurations", "priority": "P1", "type": "Negative"},
    {"id": "CC_002_003", "scenario": "Invalid token", "objective": "Token validation", "api": "getCreditLimitConfigurations", "method": "GET", "endpoint": "/v1/external/cards/limit-configurations", "priority": "P1", "type": "Negative"},
    {"id": "CC_002_004", "scenario": "Configuration consistency", "objective": "Data consistency check", "api": "getCreditLimitConfigurations", "method": "GET", "endpoint": "/v1/external/cards/limit-configurations", "priority": "P1", "type": "Integration"},
    
    # getInstalmentConversionSchedule (5 cases)
    {"id": "CC_003_001", "scenario": "Valid request with transaction ref", "objective": "Retrieve schedule", "api": "getInstalmentConversionSchedule", "method": "GET", "endpoint": "/v1/external/cards/instalment-schedule", "priority": "P0", "type": "Positive"},
    {"id": "CC_003_002", "scenario": "Missing transaction reference", "objective": "Parameter validation", "api": "getInstalmentConversionSchedule", "method": "GET", "endpoint": "/v1/external/cards/instalment-schedule", "priority": "P1", "type": "Negative"},
    {"id": "CC_003_003", "scenario": "Invalid transaction reference", "objective": "Invalid data handling", "api": "getInstalmentConversionSchedule", "method": "GET", "endpoint": "/v1/external/cards/instalment-schedule", "priority": "P1", "type": "Negative"},
    {"id": "CC_003_004", "scenario": "Boundary - max duration", "objective": "Duration limits", "api": "getInstalmentConversionSchedule", "method": "GET", "endpoint": "/v1/external/cards/instalment-schedule", "priority": "P1", "type": "Boundary"},
    {"id": "CC_003_005", "scenario": "Auth validation", "objective": "Authentication check", "api": "getInstalmentConversionSchedule", "method": "GET", "endpoint": "/v1/external/cards/instalment-schedule", "priority": "P1", "type": "Security"},
    
    # creditCardActivation (8 cases)
    {"id": "CC_004_001", "scenario": "Activate with PIN", "objective": "Successful activation", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P0", "type": "Positive"},
    {"id": "CC_004_002", "scenario": "PIN mismatch", "objective": "PIN validation", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P1", "type": "Negative"},
    {"id": "CC_004_003", "scenario": "Invalid card number", "objective": "Card validation", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P1", "type": "Negative"},
    {"id": "CC_004_004", "scenario": "Invalid PIN length", "objective": "PIN format validation", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P1", "type": "Boundary"},
    {"id": "CC_004_005", "scenario": "Already activated card", "objective": "State validation", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P1", "type": "Negative"},
    {"id": "CC_004_006", "scenario": "Missing required fields", "objective": "Mandatory field check", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P1", "type": "Negative"},
    {"id": "CC_004_007", "scenario": "Invalid activation code", "objective": "Code validation", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P2", "type": "Security"},
    {"id": "CC_004_008", "scenario": "Cross-customer access", "objective": "Authorization check", "api": "creditCardActivation", "method": "POST", "endpoint": "/v1/external/cards/activate", "priority": "P2", "type": "Security"},
    
    # creditCardBlock (7 cases)
    {"id": "CC_005_001", "scenario": "Valid block request", "objective": "Block card successfully", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P0", "type": "Positive"},
    {"id": "CC_005_002", "scenario": "Block already blocked card", "objective": "State validation", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P1", "type": "Negative"},
    {"id": "CC_005_003", "scenario": "Invalid card number", "objective": "Card validation", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P1", "type": "Negative"},
    {"id": "CC_005_004", "scenario": "Non-existent card", "objective": "Card existence check", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P1", "type": "Negative"},
    {"id": "CC_005_005", "scenario": "Missing reason", "objective": "Mandatory field check", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P1", "type": "Negative"},
    {"id": "CC_005_006", "scenario": "Invalid reason value", "objective": "Enum validation", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P1", "type": "Negative"},
    {"id": "CC_005_007", "scenario": "Block inactive card", "objective": "Business rule validation", "api": "creditCardBlock", "method": "POST", "endpoint": "/v1/external/cards/block", "priority": "P2", "type": "Negative"},
    
    # creditCardUnblock (5 cases)
    {"id": "CC_006_001", "scenario": "Unblock blocked card", "objective": "Unblock successfully", "api": "creditCardUnblock", "method": "POST", "endpoint": "/v1/external/cards/unblock", "priority": "P0", "type": "Positive"},
    {"id": "CC_006_002", "scenario": "Unblock non-blocked card", "objective": "State validation", "api": "creditCardUnblock", "method": "POST", "endpoint": "/v1/external/cards/unblock", "priority": "P1", "type": "Negative"},
    {"id": "CC_006_003", "scenario": "Invalid card number", "objective": "Card validation", "api": "creditCardUnblock", "method": "POST", "endpoint": "/v1/external/cards/unblock", "priority": "P1", "type": "Negative"},
    {"id": "CC_006_004", "scenario": "Card not found", "objective": "Resource check", "api": "creditCardUnblock", "method": "POST", "endpoint": "/v1/external/cards/unblock", "priority": "P1", "type": "Negative"},
    {"id": "CC_006_005", "scenario": "Missing card number", "objective": "Mandatory field", "api": "creditCardUnblock", "method": "POST", "endpoint": "/v1/external/cards/unblock", "priority": "P1", "type": "Negative"},
    
    # creditCardResetPin (6 cases)
    {"id": "CC_007_001", "scenario": "Reset with valid PIN", "objective": "Reset PIN successfully", "api": "creditCardResetPin", "method": "POST", "endpoint": "/v1/external/cards/reset-pin", "priority": "P0", "type": "Positive"},
    {"id": "CC_007_002", "scenario": "PIN mismatch", "objective": "PIN validation", "api": "creditCardResetPin", "method": "POST", "endpoint": "/v1/external/cards/reset-pin", "priority": "P1", "type": "Negative"},
    {"id": "CC_007_003", "scenario": "Invalid PIN length", "objective": "PIN format", "api": "creditCardResetPin", "method": "POST", "endpoint": "/v1/external/cards/reset-pin", "priority": "P1", "type": "Boundary"},
    {"id": "CC_007_004", "scenario": "Card not found", "objective": "Resource check", "api": "creditCardResetPin", "method": "POST", "endpoint": "/v1/external/cards/reset-pin", "priority": "P1", "type": "Negative"},
    {"id": "CC_007_005", "scenario": "Same as old PIN", "objective": "Business rule", "api": "creditCardResetPin", "method": "POST", "endpoint": "/v1/external/cards/reset-pin", "priority": "P1", "type": "Negative"},
    {"id": "CC_007_006", "scenario": "Blocked card PIN reset", "objective": "State validation", "api": "creditCardResetPin", "method": "POST", "endpoint": "/v1/external/cards/reset-pin", "priority": "P2", "type": "Negative"},
    
    # creditCardPinValidation (6 cases)
    {"id": "CC_008_001", "scenario": "Validate correct PIN", "objective": "Success validation", "api": "creditCardPinValidation", "method": "POST", "endpoint": "/v1/external/cards/validate-pin", "priority": "P0", "type": "Positive"},
    {"id": "CC_008_002", "scenario": "Validate incorrect PIN", "objective": "Failure validation", "api": "creditCardPinValidation", "method": "POST", "endpoint": "/v1/external/cards/validate-pin", "priority": "P1", "type": "Negative"},
    {"id": "CC_008_003", "scenario": "Too many attempts", "objective": "Rate limiting", "api": "creditCardPinValidation", "method": "POST", "endpoint": "/v1/external/cards/validate-pin", "priority": "P2", "type": "Security"},
    {"id": "CC_008_004", "scenario": "Missing PIN", "objective": "Mandatory field", "api": "creditCardPinValidation", "method": "POST", "endpoint": "/v1/external/cards/validate-pin", "priority": "P1", "type": "Negative"},
    {"id": "CC_008_005", "scenario": "Blocked card validation", "objective": "State check", "api": "creditCardPinValidation", "method": "POST", "endpoint": "/v1/external/cards/validate-pin", "priority": "P1", "type": "Negative"},
    {"id": "CC_008_006", "scenario": "Response security check", "objective": "No PIN exposure", "api": "creditCardPinValidation", "method": "POST", "endpoint": "/v1/external/cards/validate-pin", "priority": "P2", "type": "Security"},
    
    # cvvInquiry (6 cases)
    {"id": "CC_009_001", "scenario": "Valid CVV inquiry", "objective": "Retrieve CVV", "api": "cvvInquiry", "method": "POST", "endpoint": "/v1/external/cards/cvv-inquiry", "priority": "P0", "type": "Positive"},
    {"id": "CC_009_002", "scenario": "Card not found", "objective": "Resource check", "api": "cvvInquiry", "method": "POST", "endpoint": "/v1/external/cards/cvv-inquiry", "priority": "P1", "type": "Negative"},
    {"id": "CC_009_003", "scenario": "Invalid card number", "objective": "Card validation", "api": "cvvInquiry", "method": "POST", "endpoint": "/v1/external/cards/cvv-inquiry", "priority": "P1", "type": "Negative"},
    {"id": "CC_009_004", "scenario": "CVV security validation", "objective": "Security check", "api": "cvvInquiry", "method": "POST", "endpoint": "/v1/external/cards/cvv-inquiry", "priority": "P2", "type": "Security"},
    {"id": "CC_009_005", "scenario": "Cross-customer access", "objective": "Authorization check", "api": "cvvInquiry", "method": "POST", "endpoint": "/v1/external/cards/cvv-inquiry", "priority": "P2", "type": "Security"},
    {"id": "CC_009_006", "scenario": "Response masking verification", "objective": "Verify no full CVV", "api": "cvvInquiry", "method": "POST", "endpoint": "/v1/external/cards/cvv-inquiry", "priority": "P2", "type": "Security"},
    
    # replaceCreditCard (6 cases)
    {"id": "CC_010_001", "scenario": "Valid replacement request", "objective": "Replace successfully", "api": "replaceCreditCard", "method": "POST", "endpoint": "/v1/external/cards/replace", "priority": "P0", "type": "Positive"},
    {"id": "CC_010_002", "scenario": "Invalid replacement reason", "objective": "Enum validation", "api": "replaceCreditCard", "method": "POST", "endpoint": "/v1/external/cards/replace", "priority": "P1", "type": "Negative"},
    {"id": "CC_010_003", "scenario": "Missing delivery address", "objective": "Mandatory field", "api": "replaceCreditCard", "method": "POST", "endpoint": "/v1/external/cards/replace", "priority": "P1", "type": "Negative"},
    {"id": "CC_010_004", "scenario": "Card not found", "objective": "Resource check", "api": "replaceCreditCard", "method": "POST", "endpoint": "/v1/external/cards/replace", "priority": "P1", "type": "Negative"},
    {"id": "CC_010_005", "scenario": "Duplicate replacement request", "objective": "Business rule", "api": "replaceCreditCard", "method": "POST", "endpoint": "/v1/external/cards/replace", "priority": "P1", "type": "Negative"},
    {"id": "CC_010_006", "scenario": "Inactive card replacement", "objective": "State validation", "api": "replaceCreditCard", "method": "POST", "endpoint": "/v1/external/cards/replace", "priority": "P2", "type": "Negative"},
    
    # increaseCreditLimit (10 cases)
    {"id": "CC_011_001", "scenario": "Increase permanent limit", "objective": "Increase limit", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P0", "type": "Positive"},
    {"id": "CC_011_002", "scenario": "Increase to same limit", "objective": "No-op validation", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Negative"},
    {"id": "CC_011_003", "scenario": "Decrease limit attempt", "objective": "Direction validation", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Negative"},
    {"id": "CC_011_004", "scenario": "Exceed maximum limit", "objective": "Boundary check", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Boundary"},
    {"id": "CC_011_005", "scenario": "Invalid currency", "objective": "Enum validation", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Negative"},
    {"id": "CC_011_006", "scenario": "Invalid limit type", "objective": "Enum validation", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Negative"},
    {"id": "CC_011_007", "scenario": "Negative amount", "objective": "Range validation", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Boundary"},
    {"id": "CC_011_008", "scenario": "Customer not eligible", "objective": "Business rule", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P2", "type": "Negative"},
    {"id": "CC_011_009", "scenario": "Boundary - max limit value", "objective": "Edge case", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Boundary"},
    {"id": "CC_011_010", "scenario": "Boundary - min limit value", "objective": "Edge case", "api": "increaseCreditLimit", "method": "POST", "endpoint": "/v1/external/cards/increase-limit", "priority": "P1", "type": "Boundary"},
    
    # applyInstConvert (9 cases)
    {"id": "CC_012_001", "scenario": "Apply instalment conversion", "objective": "Convert successfully", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P0", "type": "Positive"},
    {"id": "CC_012_002", "scenario": "Invalid transaction reference", "objective": "Reference validation", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Negative"},
    {"id": "CC_012_003", "scenario": "Invalid instalment period", "objective": "Period validation", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Negative"},
    {"id": "CC_012_004", "scenario": "Insufficient transaction amount", "objective": "Amount validation", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Negative"},
    {"id": "CC_012_005", "scenario": "Already converted transaction", "objective": "State validation", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Negative"},
    {"id": "CC_012_006", "scenario": "Missing conversation ID", "objective": "Mandatory field", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Negative"},
    {"id": "CC_012_007", "scenario": "Duplicate conversion request", "objective": "Idempotency", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P2", "type": "Negative"},
    {"id": "CC_012_008", "scenario": "Boundary - maximum period (24 months)", "objective": "Period limit", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Boundary"},
    {"id": "CC_012_009", "scenario": "Boundary - minimum period (3 months)", "objective": "Period limit", "api": "applyInstConvert", "method": "POST", "endpoint": "/v1/external/cards/apply-instalment", "priority": "P1", "type": "Boundary"},
]

# Test data mapping
TEST_DATA = [
    {"id": "CC_001_001", "scenario": "Valid getCreditLimits", "status": "200", "response": "success", "source": "production_data"},
    {"id": "CC_001_002", "scenario": "Missing auth", "status": "401", "response": "unauthorized", "source": "test_data"},
    {"id": "CC_004_001", "scenario": "Activate card", "status": "202", "response": "accepted", "source": "test_data"},
    {"id": "CC_005_001", "scenario": "Block card", "status": "202", "response": "accepted", "source": "test_data"},
    {"id": "CC_011_001", "scenario": "Increase limit", "status": "202", "response": "accepted", "source": "test_data"},
]

# ============================================================================
# EXCEL GENERATION WITH SUMMARY
# ============================================================================

def apply_header_style(cell, header_fill=None, header_font=None):
    """Apply header styling"""
    if header_fill is None:
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    if header_font is None:
        header_font = Font(bold=True, color="FFFFFF", size=11)
    
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    cell.border = thin_border

def apply_data_style(cell, border_only=False):
    """Apply data cell styling"""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    cell.border = thin_border
    if not border_only:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def create_summary_sheet(wb):
    """Create summary sheet with metrics"""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "ECLIPSE Account Dashboard - Credit Card DDD API"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Test Case Generation Summary"
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')
    
    # Generation date
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(italic=True, size=10)
    
    # Overall metrics
    ws['A5'] = "OVERALL METRICS"
    ws['A5'].font = Font(bold=True, size=11, color="FFFFFF")
    ws['A5'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A5:B5')
    
    row = 6
    metrics = [
        ("Total APIs Tested", len(APIS)),
        ("Total Test Cases", len(TEST_CASES_DATA)),
        ("Test Data Records", len(TEST_DATA)),
        ("Positive Test Cases", len([t for t in TEST_CASES_DATA if t["type"] == "Positive"])),
        ("Negative Test Cases", len([t for t in TEST_CASES_DATA if t["type"] == "Negative"])),
        ("Boundary Test Cases", len([t for t in TEST_CASES_DATA if t["type"] == "Boundary"])),
        ("Security Test Cases", len([t for t in TEST_CASES_DATA if t["type"] == "Security"])),
        ("Integration Test Cases", len([t for t in TEST_CASES_DATA if t["type"] == "Integration"])),
        ("Performance Test Cases", len([t for t in TEST_CASES_DATA if t["type"] == "Performance"])),
    ]
    
    for metric, value in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'B{row}'].alignment = Alignment(horizontal="center")
        row += 1
    
    # API breakdown
    row += 1
    ws[f'A{row}'] = "API BREAKDOWN (By Subfolder)"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    # Headers
    headers = ["API ID", "API Name", "HTTP Method", "Test Cases"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_header_style(cell)
    
    row += 1
    # Data
    for api in APIS:
        ws[f'A{row}'] = api["id"]
        ws[f'B{row}'] = api["name"]
        ws[f'C{row}'] = api["method"]
        ws[f'D{row}'] = api["count"]
        for col in range(1, 5):
            apply_data_style(ws.cell(row=row, column=col))
        row += 1
    
    # Priority breakdown
    row += 2
    ws[f'A{row}'] = "TEST PRIORITY BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    headers = ["Priority", "Count", "Execution Time"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_header_style(cell)
    
    row += 1
    priorities = [
        ("P0 (Critical)", len([t for t in TEST_CASES_DATA if t["priority"] == "P0"]), "20 min"),
        ("P1 (High)", len([t for t in TEST_CASES_DATA if t["priority"] == "P1"]), "1-2 hours"),
        ("P2 (Medium)", len([t for t in TEST_CASES_DATA if t["priority"] == "P2"]), "1-2 hours"),
        ("P3 (Low)", len([t for t in TEST_CASES_DATA if t["priority"] == "P3"]), "30-60 min"),
    ]
    
    for priority, count, time in priorities:
        ws[f'A{row}'] = priority
        ws[f'B{row}'] = count
        ws[f'C{row}'] = time
        for col in range(1, 4):
            apply_data_style(ws.cell(row=row, column=col))
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

def create_test_cases_sheet(wb):
    """Create test cases sheet organized by API"""
    ws = wb.create_sheet("Test Cases")
    
    # Headers
    headers = [
        "Test Case ID", "Test Scenario", "Test Objective", "API Name",
        "HTTP Method", "Endpoint", "Priority", "Test Type", "Expected Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        apply_header_style(cell)
    
    # Data - organized by API
    row = 2
    for test in TEST_CASES_DATA:
        ws[f'A{row}'] = test["id"]
        ws[f'B{row}'] = test["scenario"]
        ws[f'C{row}'] = test["objective"]
        ws[f'D{row}'] = test["api"]
        ws[f'E{row}'] = test["method"]
        ws[f'F{row}'] = test["endpoint"]
        ws[f'G{row}'] = test["priority"]
        ws[f'H{row}'] = test["type"]
        ws[f'I{row}'] = "200/201/202/204" if test["type"] == "Positive" else "400/401/403/404/422"
        
        for col in range(1, 10):
            apply_data_style(ws.cell(row=row, column=col))
        row += 1
    
    # Freeze panes
    ws.freeze_panes = "A2"
    
    # Set column widths
    widths = [15, 25, 25, 30, 12, 30, 8, 15, 18]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def create_test_data_sheet(wb):
    """Create test data sheet"""
    ws = wb.create_sheet("Test Data")
    
    # Headers
    headers = [
        "Test Case ID", "Test Scenario", "Expected Status", "Expected Response", "Data Source", "Execution Date", "Actual Result"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        apply_header_style(cell)
    
    # Data
    row = 2
    for test in TEST_DATA:
        ws[f'A{row}'] = test["id"]
        ws[f'B{row}'] = test["scenario"]
        ws[f'C{row}'] = test["status"]
        ws[f'D{row}'] = test["response"]
        ws[f'E{row}'] = test["source"]
        ws[f'F{row}'] = ""  # For execution date
        ws[f'G{row}'] = ""  # For actual result
        
        for col in range(1, 8):
            apply_data_style(ws.cell(row=row, column=col))
        row += 1
    
    # Freeze panes
    ws.freeze_panes = "A2"
    
    # Set column widths
    widths = [15, 25, 15, 20, 20, 18, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def create_api_coverage_sheet(wb):
    """Create API coverage sheet by subfolder"""
    ws = wb.create_sheet("API Coverage")
    
    # Title
    ws['A1'] = "API Test Coverage by Subfolder (artifacts/Acc_Dashboard_CC)"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Headers
    row = 3
    headers = ["API Subfolder", "API Method", "Endpoint", "Total Cases", "Positive", "Negative"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_header_style(cell)
    
    row += 1
    for api in APIS:
        api_cases = [t for t in TEST_CASES_DATA if t["api"] == api["name"]]
        positive = len([t for t in api_cases if t["type"] == "Positive"])
        negative = len([t for t in api_cases if t["type"] != "Positive"])
        
        ws[f'A{row}'] = api["name"]
        ws[f'B{row}'] = api["method"]
        ws[f'C{row}'] = f"/v1/external/cards/{api['name'].lower()}"
        ws[f'D{row}'] = api["count"]
        ws[f'E{row}'] = positive
        ws[f'F{row}'] = negative
        
        for col in range(1, 7):
            apply_data_style(ws.cell(row=row, column=col))
        row += 1
    
    # Set column widths
    widths = [30, 12, 35, 12, 10, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def main():
    """Main execution"""
    print("=" * 70)
    print("ECLIPSE Account Dashboard - Credit Card DDD API")
    print("Test Case & Test Data Generation with Summary")
    print("=" * 70)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    create_summary_sheet(wb)
    create_test_cases_sheet(wb)
    create_test_data_sheet(wb)
    create_api_coverage_sheet(wb)
    
    # Save workbook
    output_path = "artifacts/Test_Case/Excel/Account_Dashboard_Credit_Card_DDD_Organized.xlsx"
    wb.save(output_path)
    
    print(f"\n✓ Test Cases organized by subfolder: {len(TEST_CASES_DATA)} cases")
    print(f"✓ APIs Covered: {len(APIS)}")
    print(f"✓ Sheets Generated: 4 (Summary, Test Cases, Test Data, API Coverage)")
    print(f"\n✓ File saved: {output_path}")
    
    print("\n" + "=" * 70)
    print("SUMMARY:")
    print(f"  Total Test Cases: {len(TEST_CASES_DATA)}")
    print(f"  Total APIs: {len(APIS)}")
    print(f"  Positive Cases: {len([t for t in TEST_CASES_DATA if t['type'] == 'Positive'])}")
    print(f"  Negative Cases: {len([t for t in TEST_CASES_DATA if t['type'] == 'Negative'])}")
    print(f"  Boundary Cases: {len([t for t in TEST_CASES_DATA if t['type'] == 'Boundary'])}")
    print(f"  Security Cases: {len([t for t in TEST_CASES_DATA if t['type'] == 'Security'])}")
    print("=" * 70)

if __name__ == "__main__":
    main()
