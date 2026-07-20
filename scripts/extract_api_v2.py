import openpyxl
import json
from typing import Dict, List, Any

file_path = r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx"

def extract_table_data(ws, start_row: int, column_mapping: Dict[str, int]) -> List[Dict[str, Any]]:
    """Extract table data starting from start_row + 1 (skip header row)"""
    data = []
    current_row = start_row + 1
    
    while current_row <= ws.max_row:
        row_data = {}
        has_data = False
        
        for col_name, col_idx in column_mapping.items():
            cell_value = ws.cell(row=current_row, column=col_idx).value
            if cell_value and str(cell_value).strip() not in ['-', '']:
                has_data = True
            row_data[col_name] = str(cell_value).strip() if cell_value and str(cell_value).strip() != '-' else '-'
        
        # Stop if no data found in this row (empty row or section marker)
        if not has_data or row_data.get('Name', '') == '-':
            break
        
        # Skip header rows (marked with "Name" in first column on its own)
        if row_data.get('Name', '').lower() == 'name' and len([v for v in row_data.values() if v != '-']) <= 1:
            current_row += 1
            continue
        
        data.append(row_data)
        current_row += 1
    
    return data

def extract_api_spec(sheet_name: str, ws) -> Dict[str, Any]:
    """Extract complete API specification from worksheet"""
    spec = {}
    
    # Row 1: Service
    spec['Service'] = str(ws['B1'].value).strip() if ws['B1'].value else 'Unknown'
    
    # Row 2: Method
    spec['Method'] = str(ws['B2'].value).strip() if ws['B2'].value else 'Unknown'
    
    # Row 3: URL
    spec['URL'] = str(ws['B3'].value).strip() if ws['B3'].value else 'Unknown'
    
    # Row 4: Message Type
    spec['MessageType'] = str(ws['B4'].value).strip() if ws['B4'].value else 'JSON'
    
    spec['HTTPHeaders'] = []
    spec['HTTPBody'] = []
    spec['RequestParameters'] = []
    spec['Response'] = []
    spec['ErrorResponses'] = []
    
    # Scan for sections
    for row in range(5, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        cell_b = ws.cell(row=row, column=2).value
        
        if not cell_b:
            continue
        
        cell_b_str = str(cell_b).strip() if cell_b else ''
        
        # HTTP Headers section
        if 'HTTP Header' in cell_b_str and 'Response' not in cell_b_str:
            column_mapping = {
                'Name': 2, 'Type': 3, 'Length': 4,
                'Mandatory': 5, 'Sample Value': 6, 'Description': 7
            }
            spec['HTTPHeaders'] = extract_table_data(ws, row, column_mapping)
        
        # HTTP Body section
        elif 'HTTP Body' in cell_b_str:
            column_mapping = {
                'Name': 2, 'Type': 3, 'Length': 4,
                'Mandatory': 5, 'Sample Value': 6, 'Description': 7
            }
            spec['HTTPBody'] = extract_table_data(ws, row, column_mapping)
        
        # Request Parameter section
        elif 'Request Parameter' in cell_b_str:
            column_mapping = {
                'Name': 2, 'Type': 3, 'Length': 4,
                'Mandatory': 5, 'Sample Value': 6, 'Description': 7
            }
            spec['RequestParameters'] = extract_table_data(ws, row, column_mapping)
        
        # Response section (but not Error Response)
        elif cell_b_str == 'Response':
            column_mapping = {
                'Name': 2, 'Type': 3, 'Length': 4,
                'Mandatory': 5, 'Sample Value': 6, 'Description': 7
            }
            spec['Response'] = extract_table_data(ws, row, column_mapping)
        
        # Error Response section
        elif 'Error Response' in cell_b_str:
            column_mapping = {
                'Error Code': 2, 'Error Message': 3, 'Description': 4
            }
            spec['ErrorResponses'] = extract_table_data(ws, row, column_mapping)
        
        # Path Variable section
        elif 'Path Variable' in cell_b_str:
            column_mapping = {
                'Name': 2, 'Type': 3, 'Length': 4,
                'Mandatory': 5, 'Sample Value': 6, 'Description': 7
            }
            path_vars = extract_table_data(ws, row, column_mapping)
            spec['PathVariables'] = path_vars
    
    return spec

# Load workbook
wb = openpyxl.load_workbook(file_path)
sheet_names = wb.sheetnames

# Filter out metadata sheets
skip_sheets = ['README', 'API_Specs_Index', 'Sheet1']
api_sheets = [s for s in sheet_names if s not in skip_sheets]

# Extract all specs
all_specs = {}
for sheet_name in api_sheets:
    try:
        ws = wb[sheet_name]
        spec = extract_api_spec(sheet_name, ws)
        all_specs[sheet_name] = spec
    except Exception as e:
        print(f"Error processing {sheet_name}: {e}", file=__import__('sys').stderr)

# Save to JSON
output_file = r'artifacts\api_endpoints_summary.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(all_specs, f, indent=2, ensure_ascii=False)

print(f"Extraction complete: {len(all_specs)} endpoints")
print(f"Output: {output_file}")

# Display summary
for sheet_name, spec in all_specs.items():
    print(f"\n{spec['Service']}: {spec['Method']} - Headers: {len(spec['HTTPHeaders'])}, Body: {len(spec['HTTPBody'])}, Errors: {len(spec['ErrorResponses'])}")

