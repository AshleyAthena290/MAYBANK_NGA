#!/usr/bin/env python3
"""
Comprehensive API Analysis Extraction Script
Extracts all API specifications from Excel for test case generation
Target: ECLIPSE Account Dashboard Credit Card DDD API Design v1
"""

import openpyxl
import json
import sys
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

# File path
EXCEL_FILE = r"input\api\ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx"
OUTPUT_FILE = r"artifacts\CreditCard_API_Analysis.md"

class APIAnalyzer:
    def __init__(self, excel_path: str):
        """Initialize the analyzer with Excel workbook"""
        self.wb = openpyxl.load_workbook(excel_path)
        self.apis = []
        self.api_index = {}
        
    @staticmethod
    def clean_value(val) -> Optional[str]:
        """Clean cell value"""
        if val is None:
            return None
        return str(val).strip()
    
    def get_cell_value(self, ws, row: int, col: int) -> Optional[str]:
        """Get cell value safely"""
        try:
            cell = ws.cell(row=row, column=col)
            return self.clean_value(cell.value) if cell.value else None
        except:
            return None
    
    def extract_api_index(self) -> List[Dict[str, Any]]:
        """Extract API index from the index sheet"""
        api_list = []
        
        # Try to find the index sheet
        index_sheets = [s for s in self.wb.sheetnames if 'index' in s.lower()]
        if not index_sheets:
            return api_list
        
        ws = self.wb[index_sheets[0]]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=200, values_only=False)):
            row_data = [self.clean_value(cell.value) if cell else None for cell in row]
            
            if not row_data or not any(row_data):
                continue
            
            if len(row_data) >= 6:
                api_info = {
                    'number': row_data[0],
                    'sheet_name': row_data[1],
                    'api_name': row_data[2],
                    'http_method': row_data[3],
                    'microservice': row_data[4] if len(row_data) > 4 else None,
                    'endpoint': row_data[5] if len(row_data) > 5 else None,
                    'status': row_data[6] if len(row_data) > 6 else None,
                    'remarks': row_data[7] if len(row_data) > 7 else None,
                }
                
                if api_info['api_name'] and api_info['sheet_name']:
                    api_list.append(api_info)
                    self.api_index[api_info['sheet_name']] = api_info
        
        return api_list
    
    def extract_section_from_sheet(self, ws, section_marker: str, 
                                   start_row: int = 1, 
                                   end_row: int = 500) -> List[List[str]]:
        """Extract a section from worksheet"""
        rows = []
        found_section = False
        in_section = False
        
        for row_idx in range(start_row, end_row):
            row_data = []
            for col_idx in range(1, 12):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(self.clean_value(cell.value) if cell else None)
            
            # Check if we found the section marker
            first_col = row_data[0] if row_data else None
            if first_col and section_marker.lower() in first_col.lower():
                found_section = True
                in_section = True
                # Skip the header row
                if row_idx + 1 < end_row:
                    # Read next row which should have column headers
                    header_row = []
                    for col_idx in range(1, 12):
                        cell = ws.cell(row=row_idx + 1, column=col_idx)
                        header_row.append(self.clean_value(cell.value) if cell else None)
                    rows.append(header_row)
                continue
            
            # If in section, check if we hit another section
            if in_section:
                if first_col and any(marker in first_col.lower() for marker in 
                                     ['request', 'response', 'business', 'error', 
                                      'validation', 'sample', 'auth', 'security']):
                    if section_marker.lower() not in first_col.lower():
                        break
                
                # Add data rows
                if any(row_data[1:]):  # Has data beyond first column
                    rows.append(row_data)
        
        return rows
    
    def extract_api_details(self, sheet_name: str) -> Dict[str, Any]:
        """Extract all details for a single API from its sheet"""
        if sheet_name not in self.wb.sheetnames:
            return {}
        
        ws = self.wb[sheet_name]
        api = {
            'sheet_name': sheet_name,
            'basic_info': {},
            'endpoint_info': {},
            'request_headers': [],
            'request_parameters': [],
            'request_body': [],
            'response_codes': {},
            'response_structure': [],
            'business_rules': [],
            'error_scenarios': [],
            'validations': [],
            'authentication': [],
            'authorization': [],
            'samples': {
                'request': None,
                'response': None
            }
        }
        
        # Extract basic info (first 15 rows)
        for row_idx in range(1, 16):
            col_a = self.get_cell_value(ws, row_idx, 1)
            col_b = self.get_cell_value(ws, row_idx, 2)
            
            if col_a:
                key = col_a.lower().replace(' ', '_')
                if col_a in ['Service', 'Method', 'URL', 'Message Type', 'Endpoint', 'Path']:
                    api['basic_info'][key] = col_b
                    if col_a == 'Method':
                        api['endpoint_info']['http_method'] = col_b
                    elif col_a == 'URL':
                        api['endpoint_info']['url'] = col_b
                        # Parse URL to extract path
                        if col_b:
                            api['endpoint_info']['path'] = col_b.split(' ')[-1] if ' ' in col_b else col_b
        
        # Extract sections
        sections = {
            'request_headers': ('HTTP Header', 'Request Header'),
            'request_parameters': ('Request Parameter', 'Path Parameter', 'Query Parameter'),
            'request_body': ('HTTP Body', 'Request Body'),
            'response_structure': ('Response', 'Response Structure'),
        }
        
        for field, markers in sections.items():
            for marker in markers:
                rows = self.extract_section_from_sheet(ws, marker)
                if rows and len(rows) > 1:  # Has header + data
                    api[field] = self._parse_section_rows(rows, marker)
                    break
        
        # Extract business rules
        br_rows = self.extract_section_from_sheet(ws, 'Business Rule')
        if br_rows and len(br_rows) > 1:
            api['business_rules'] = [row[1] for row in br_rows[1:] if row[1]]
        
        # Extract error scenarios
        error_rows = self.extract_section_from_sheet(ws, 'Error')
        if error_rows and len(error_rows) > 1:
            for row in error_rows[1:]:
                if row[1] and len(row) > 2:
                    api['error_scenarios'].append({
                        'error_code': row[1],
                        'description': row[2] if len(row) > 2 else None,
                        'http_status': row[3] if len(row) > 3 else None,
                    })
        
        # Extract validations
        val_rows = self.extract_section_from_sheet(ws, 'Validation')
        if val_rows and len(val_rows) > 1:
            for row in val_rows[1:]:
                if row[1]:
                    api['validations'].append({
                        'rule': row[1],
                        'description': row[2] if len(row) > 2 else None,
                        'field': row[3] if len(row) > 3 else None,
                    })
        
        # Extract authentication
        auth_rows = self.extract_section_from_sheet(ws, 'Authentication')
        if auth_rows and len(auth_rows) > 1:
            api['authentication'] = [row[1] for row in auth_rows[1:] if row[1]]
        
        # Extract authorization
        authz_rows = self.extract_section_from_sheet(ws, 'Authorization')
        if authz_rows and len(authz_rows) > 1:
            api['authorization'] = [row[1] for row in authz_rows[1:] if row[1]]
        
        # Extract samples
        for row_idx in range(1, ws.max_row + 1):
            col_a = self.get_cell_value(ws, row_idx, 1)
            col_b = self.get_cell_value(ws, row_idx, 2)
            
            if col_a == 'Request Sample' and col_b:
                api['samples']['request'] = col_b
            elif col_a == 'Response Sample' and col_b:
                api['samples']['response'] = col_b
        
        # Extract HTTP status codes from response structure
        for row in api['response_structure']:
            if isinstance(row, dict) and 'http_status' in row:
                status = row['http_status']
                if status:
                    api['response_codes'][status] = row.get('description', '')
        
        return api
    
    def _parse_section_rows(self, rows: List[List[str]], section_type: str) -> List[Dict[str, Any]]:
        """Parse section rows into structured data"""
        if not rows or len(rows) < 2:
            return []
        
        headers = rows[0]
        parsed = []
        
        for row in rows[1:]:
            if not any(row):
                continue
            
            item = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    value = row[i]
                    if header and value:
                        # Normalize header names
                        key = header.lower().replace(' ', '_').strip()
                        item[key] = value
            
            if item:
                parsed.append(item)
        
        return parsed
    
    def analyze_all_apis(self) -> List[Dict[str, Any]]:
        """Analyze all APIs in the workbook"""
        # Get API index
        api_index = self.extract_api_index()
        
        # Extract details for each API
        for api_info in api_index:
            sheet_name = api_info.get('sheet_name')
            if sheet_name:
                details = self.extract_api_details(sheet_name)
                if details and details.get('basic_info'):
                    api_info['details'] = details
                    self.apis.append(api_info)
        
        return self.apis
    
    def generate_markdown_report(self, output_file: str):
        """Generate comprehensive markdown report"""
        lines = []
        
        # Header
        lines.append("# ECLIPSE Account Dashboard - Credit Card API Specifications")
        lines.append("## Comprehensive Analysis for Test Case Generation")
        lines.append("")
        lines.append(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"**Excel Source:** ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx")
        lines.append("")
        
        # Executive Summary
        lines.append("## Executive Summary")
        lines.append("---")
        lines.append(f"- **Total APIs Analyzed:** {len(self.apis)}")
        
        # Count methods
        methods = {}
        for api in self.apis:
            method = api.get('http_method', 'UNKNOWN')
            methods[method] = methods.get(method, 0) + 1
        
        lines.append(f"- **API Distribution by Method:**")
        for method, count in sorted(methods.items()):
            lines.append(f"  - {method}: {count}")
        
        lines.append("")
        
        # API List Summary
        lines.append("## 1. API Endpoints Summary")
        lines.append("---")
        lines.append("")
        lines.append("| No. | API Name | HTTP Method | Endpoint | Status |")
        lines.append("|-----|----------|-------------|----------|--------|")
        
        for idx, api in enumerate(self.apis, 1):
            name = api.get('api_name', 'N/A')
            method = api.get('http_method', 'N/A')
            endpoint = api.get('endpoint', 'N/A')
            status = api.get('status', 'N/A')
            
            # Truncate if too long
            if len(endpoint) > 50:
                endpoint = endpoint[:47] + "..."
            
            lines.append(f"| {idx} | {name} | {method} | {endpoint} | {status} |")
        
        lines.append("")
        
        # Detailed API Specifications
        lines.append("## 2. Detailed API Specifications")
        lines.append("---")
        lines.append("")
        
        for idx, api in enumerate(self.apis, 1):
            details = api.get('details', {})
            basic_info = details.get('basic_info', {})
            endpoint_info = details.get('endpoint_info', {})
            
            # API Header
            lines.append(f"### 2.{idx} {api.get('api_name', 'Unknown API')}")
            lines.append("")
            
            # Basic Info
            lines.append("#### Basic Information")
            lines.append(f"- **HTTP Method:** `{api.get('http_method', 'N/A')}`")
            lines.append(f"- **Endpoint URL:** `{endpoint_info.get('url', 'N/A')}`")
            lines.append(f"- **Endpoint Path:** `{endpoint_info.get('path', 'N/A')}`")
            lines.append(f"- **Microservice:** {api.get('microservice', 'N/A')}")
            lines.append(f"- **Status:** {api.get('status', 'N/A')}")
            if api.get('remarks'):
                lines.append(f"- **Remarks:** {api.get('remarks')}")
            lines.append("")
            
            # Request Headers
            if details.get('request_headers'):
                lines.append("#### Request Headers")
                lines.append("| Field | Type | Mandatory | Sample Value | Description |")
                lines.append("|-------|------|-----------|--------------|-------------|")
                for header in details['request_headers']:
                    field = header.get('field_name', header.get('name', 'N/A'))
                    typ = header.get('type', 'string')
                    mandatory = header.get('mandatory', 'No')
                    sample = header.get('sample_value', '-')
                    desc = header.get('description', '-')
                    lines.append(f"| {field} | {typ} | {mandatory} | {sample} | {desc} |")
                lines.append("")
            
            # Request Parameters
            if details.get('request_parameters'):
                lines.append("#### Request Parameters")
                lines.append("| Parameter | Type | Mandatory | Sample Value | Description |")
                lines.append("|-----------|------|-----------|--------------|-------------|")
                for param in details['request_parameters']:
                    name = param.get('field_name', param.get('name', param.get('parameter', 'N/A')))
                    typ = param.get('type', 'string')
                    mandatory = param.get('mandatory', 'No')
                    sample = param.get('sample_value', '-')
                    desc = param.get('description', '-')
                    lines.append(f"| {name} | {typ} | {mandatory} | {sample} | {desc} |")
                lines.append("")
            
            # Request Body
            if details.get('request_body'):
                lines.append("#### Request Body Structure")
                lines.append("| Field | Type | Mandatory | Sample Value | Description |")
                lines.append("|-------|------|-----------|--------------|-------------|")
                for body_field in details['request_body']:
                    name = body_field.get('field_name', body_field.get('name', 'N/A'))
                    typ = body_field.get('type', 'string')
                    mandatory = body_field.get('mandatory', 'No')
                    sample = body_field.get('sample_value', '-')
                    desc = body_field.get('description', '-')
                    lines.append(f"| {name} | {typ} | {mandatory} | {sample} | {desc} |")
                lines.append("")
            
            # Response Codes
            if details.get('response_codes'):
                lines.append("#### HTTP Response Codes")
                lines.append("| Status Code | Description |")
                lines.append("|-------------|-------------|")
                for code, desc in sorted(details['response_codes'].items()):
                    lines.append(f"| {code} | {desc} |")
                lines.append("")
            
            # Response Structure
            if details.get('response_structure'):
                lines.append("#### Response Structure")
                lines.append("| Field | Type | Description |")
                lines.append("|-------|------|-------------|")
                for resp_field in details['response_structure']:
                    name = resp_field.get('field_name', resp_field.get('name', 'N/A'))
                    typ = resp_field.get('type', 'string')
                    desc = resp_field.get('description', '-')
                    lines.append(f"| {name} | {typ} | {desc} |")
                lines.append("")
            
            # Business Rules
            if details.get('business_rules'):
                lines.append("#### Business Rules")
                for rule in details['business_rules']:
                    lines.append(f"- {rule}")
                lines.append("")
            
            # Validations
            if details.get('validations'):
                lines.append("#### Field Validations")
                for validation in details['validations']:
                    if isinstance(validation, dict):
                        lines.append(f"- **{validation.get('field', 'N/A')}:** {validation.get('rule', '-')} ({validation.get('description', '')})")
                    else:
                        lines.append(f"- {validation}")
                lines.append("")
            
            # Error Scenarios
            if details.get('error_scenarios'):
                lines.append("#### Error Scenarios")
                lines.append("| Error Code | HTTP Status | Description |")
                lines.append("|------------|-------------|-------------|")
                for error in details['error_scenarios']:
                    code = error.get('error_code', 'N/A')
                    status = error.get('http_status', 'N/A')
                    desc = error.get('description', '-')
                    lines.append(f"| {code} | {status} | {desc} |")
                lines.append("")
            
            # Authentication
            if details.get('authentication'):
                lines.append("#### Authentication Requirements")
                for auth in details['authentication']:
                    lines.append(f"- {auth}")
                lines.append("")
            
            # Authorization
            if details.get('authorization'):
                lines.append("#### Authorization Requirements")
                for authz in details['authorization']:
                    lines.append(f"- {authz}")
                lines.append("")
            
            # Samples
            if details.get('samples', {}).get('request'):
                lines.append("#### Request Sample")
                lines.append("```json")
                lines.append(details['samples']['request'])
                lines.append("```")
                lines.append("")
            
            if details.get('samples', {}).get('response'):
                lines.append("#### Response Sample")
                lines.append("```json")
                lines.append(details['samples']['response'])
                lines.append("```")
                lines.append("")
            
            lines.append("---")
            lines.append("")
        
        # Test Case Generation Guide
        lines.append("## 3. Test Case Generation Guide")
        lines.append("---")
        lines.append("")
        lines.append("### 3.1 Positive Test Scenarios (Happy Path)")
        lines.append("- Valid requests with all required parameters")
        lines.append("- Successful API responses with HTTP 200/201/202/204")
        lines.append("- Response data matches expected structure")
        lines.append("- Business logic validation (account status, permissions, etc.)")
        lines.append("")
        
        lines.append("### 3.2 Negative Test Scenarios")
        lines.append("- Missing mandatory parameters → HTTP 400")
        lines.append("- Invalid parameter values → HTTP 400/422")
        lines.append("- Unauthorized requests (missing auth headers) → HTTP 401")
        lines.append("- Forbidden requests (insufficient permissions) → HTTP 403")
        lines.append("- Resource not found → HTTP 404")
        lines.append("- Invalid resource state → HTTP 409")
        lines.append("")
        
        lines.append("### 3.3 Boundary Value Testing")
        lines.append("- Minimum and maximum field length values")
        lines.append("- Empty strings and null values")
        lines.append("- Special characters in string fields")
        lines.append("- Maximum numeric values")
        lines.append("- Date and timestamp boundaries")
        lines.append("")
        
        lines.append("### 3.4 Field Validation Testing")
        lines.append("For each parameter/field:")
        lines.append("- Test with correct data type")
        lines.append("- Test with incorrect data type")
        lines.append("- Test with out-of-range values")
        lines.append("- Test with invalid format (for formatted fields)")
        lines.append("- Test case sensitivity (if applicable)")
        lines.append("")
        
        lines.append("### 3.5 Authentication Testing")
        lines.append("- Valid Bearer token in Authorization header")
        lines.append("- Missing Authorization header → HTTP 401")
        lines.append("- Expired Bearer token → HTTP 401")
        lines.append("- Invalid Bearer token → HTTP 401")
        lines.append("- Malformed Authorization header → HTTP 401")
        lines.append("")
        
        lines.append("### 3.6 Authorization Testing")
        lines.append("- Authorized user accessing allowed resources")
        lines.append("- User attempting cross-customer access → HTTP 403")
        lines.append("- User with insufficient permissions → HTTP 403")
        lines.append("- Admin user accessing restricted endpoints")
        lines.append("")
        
        lines.append("### 3.7 Error Handling Testing")
        lines.append("For each documented error code:")
        lines.append("- Trigger the error condition")
        lines.append("- Verify correct error code is returned")
        lines.append("- Verify correct HTTP status code")
        lines.append("- Verify error message is descriptive")
        lines.append("- Verify response structure includes error details")
        lines.append("")
        
        lines.append("### 3.8 Business Logic Validation")
        lines.append("For each business rule:")
        lines.append("- Test scenarios that satisfy the rule")
        lines.append("- Test scenarios that violate the rule")
        lines.append("- Test edge cases around rule conditions")
        lines.append("- Test interactions between multiple rules")
        lines.append("")
        
        lines.append("### 3.9 Integration Testing")
        lines.append("- Data consistency across multiple API calls")
        lines.append("- State propagation after API operations")
        lines.append("- Cascading effects of operations on related entities")
        lines.append("")
        
        lines.append("### 3.10 Performance Testing")
        lines.append("- Response time within acceptable limits")
        lines.append("- Batch operation performance")
        lines.append("- Large payload handling")
        lines.append("")
        
        # Reference Data
        lines.append("## 4. Reference Data for Test Cases")
        lines.append("---")
        lines.append("")
        
        lines.append("### 4.1 HTTP Status Codes Reference")
        lines.append("- **2xx Success**")
        lines.append("  - 200 OK: Successful GET/PATCH request")
        lines.append("  - 201 Created: Resource created via POST")
        lines.append("  - 202 Accepted: Request accepted for processing")
        lines.append("  - 204 No Content: Successful request with no response body")
        lines.append("")
        lines.append("- **4xx Client Errors**")
        lines.append("  - 400 Bad Request: Invalid parameters/malformed request")
        lines.append("  - 401 Unauthorized: Missing/invalid authentication")
        lines.append("  - 403 Forbidden: Authenticated but no permission")
        lines.append("  - 404 Not Found: Resource does not exist")
        lines.append("  - 409 Conflict: Request conflicts with current state")
        lines.append("  - 410 Gone: Resource no longer available")
        lines.append("  - 422 Unprocessable Entity: Request format valid but semantic error")
        lines.append("")
        lines.append("- **5xx Server Errors**")
        lines.append("  - 500 Internal Server Error: Unexpected server error")
        lines.append("  - 503 Service Unavailable: Server temporarily unavailable")
        lines.append("")
        
        lines.append("### 4.2 Common Test Data Sets")
        lines.append("- **Valid Customer IDs:** Numeric format, non-zero")
        lines.append("- **Valid Account Numbers:** Follow banking account format")
        lines.append("- **Valid Amounts:** Positive decimal numbers")
        lines.append("- **Valid Dates:** ISO 8601 format (YYYY-MM-DD)")
        lines.append("- **Valid Currencies:** IDR, USD, SGD, MYR, EUR")
        lines.append("- **Valid Statuses:** Based on business domain (ACTIVE, INACTIVE, BLOCKED, etc.)")
        lines.append("")
        
        lines.append("### 4.3 Negative Test Data Sets")
        lines.append("- **Invalid Customer IDs:** Zero, negative, non-numeric")
        lines.append("- **Invalid Account Numbers:** Special characters, incorrect format")
        lines.append("- **Invalid Amounts:** Negative, zero, oversized numbers")
        lines.append("- **Invalid Dates:** Wrong format, future dates (where applicable), past dates")
        lines.append("- **SQL Injection Payloads:** `'; DROP TABLE--`, `1 OR 1=1`")
        lines.append("- **XSS Payloads:** `<script>alert('XSS')</script>`, `<img src=x onerror=alert('XSS')>`")
        lines.append("")
        
        # Write to file
        output_dir = Path(output_file).parent
        output_dir.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"✓ Report generated: {output_file}")
        return output_file


def main():
    """Main execution"""
    try:
        print("=" * 80)
        print("ECLIPSE Account Dashboard - Credit Card API Analysis")
        print("=" * 80)
        print(f"Source: {EXCEL_FILE}")
        print(f"Output: {OUTPUT_FILE}")
        print("")
        
        # Initialize analyzer
        analyzer = APIAnalyzer(EXCEL_FILE)
        
        # Analyze all APIs
        print("Extracting API specifications...")
        apis = analyzer.analyze_all_apis()
        print(f"✓ Found {len(apis)} APIs")
        
        # Generate markdown report
        print("Generating comprehensive report...")
        report_path = analyzer.generate_markdown_report(OUTPUT_FILE)
        
        print("")
        print("=" * 80)
        print("ANALYSIS COMPLETE")
        print("=" * 80)
        print(f"Total APIs Analyzed: {len(apis)}")
        print(f"Report Location: {report_path}")
        
    except Exception as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
