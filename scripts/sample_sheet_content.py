#!/usr/bin/env python3
"""
Sample API sheet content analysis
"""

import openpyxl

EXCEL_FILE = r"input\api\ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx"

wb = openpyxl.load_workbook(EXCEL_FILE)

# Check the index sheet
print("=" * 100)
print("API_Specs_Index CONTENT")
print("=" * 100)

ws = wb['API_Specs_Index']
print(f"\nSheet: API_Specs_Index ({ws.max_row} rows x {ws.max_column} cols)\n")

for row_idx in range(1, min(15, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value if cell.value else ""
        row_data.append(str(value)[:25])
    print(f"Row {row_idx:2d}: | {' | '.join(row_data)} |")

# Check a sample API sheet
print("\n" + "=" * 100)
print("SAMPLE API SHEET: getCreditLimits")
print("=" * 100)

ws = wb['getCreditLimits']
print(f"\nSheet: getCreditLimits ({ws.max_row} rows x {ws.max_column} cols)\n")

for row_idx in range(1, min(20, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(8, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value if cell.value else ""
        row_data.append(str(value)[:30])
    print(f"Row {row_idx:2d}: | {' | '.join(row_data)} |")

# Check another API sheet with detail
print("\n" + "=" * 100)
print("SAMPLE API SHEET: creditCardBlock")
print("=" * 100)

ws = wb['creditCardBlock']
print(f"\nSheet: creditCardBlock ({ws.max_row} rows x {ws.max_column} cols)\n")

for row_idx in range(1, min(20, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(8, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value if cell.value else ""
        row_data.append(str(value)[:30])
    print(f"Row {row_idx:2d}: | {' | '.join(row_data)} |")
