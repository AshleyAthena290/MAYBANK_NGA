#!/usr/bin/env python3
"""
Generate Excel files for Credit Card API Test Cases and Test Data
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import json

def create_test_cases_excel():
    """Generate test cases Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define columns
    headers = [
        "Test Case ID",
        "Test Scenario",
        "Test Objective",
        "API Name",
        "HTTP Method",
        "Endpoint",
        "Preconditions",
        "Priority",
        "Test Type",
        "Expected Response",
        "Expected HTTP Status",
        "Business Validation"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 25
    
    # Test case data
    test_cases = [
        ["CC_DASHBOARD_001", "Retrieve Credit Card Dashboard", "Get all active credit cards", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "Service reachable, valid token", "P1", "Positive", "200 OK with cards array", "200", "All cards active status"],
        ["CC_DASHBOARD_002", "No Cards Response", "User with no credit cards", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "User has no cards", "P2", "Negative", "200 with empty array", "200", "Empty array returned"],
        ["CC_DASHBOARD_003", "Missing Auth Token", "Request without Bearer token", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "No bearer token in headers", "P1", "Security", "401 Unauthorized", "401", "No card data exposed"],
        ["CC_DASHBOARD_004", "Expired Token", "Request with expired token", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "Token is expired", "P1", "Security", "401 Token Expired", "401", "Access denied"],
        ["CC_DASHBOARD_005", "Invalid Client-ID", "Request with invalid client-id", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "Invalid client-id provided", "P1", "Security", "403 Forbidden", "403", "Client not authorized"],
        ["CC_DASHBOARD_006", "Missing Client-ID Header", "Request without client-id", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "Missing client-id header", "P1", "Security", "400 Bad Request", "400", "Required header validation"],
        ["CC_DASHBOARD_007", "Invalid Environment", "Request with invalid env value", "Credit Card Dashboard", "GET", "/dep/dashboard/v1/external/credit-cards", "env header = INVALID", "P1", "Negative", "400 Bad Request", "400", "Environment validation"],
        ["CC_DETAIL_001", "Get Card Details", "Retrieve complete card details", "Credit Card Detail", "GET", "/dep/card/v1/external/credit-cards/{cardId}/detail", "Valid cardId, authorized user", "P1", "Positive", "200 OK with all details", "200", "All card fields present"],
        ["CC_DETAIL_002", "Invalid Card ID Format", "Request with invalid cardId", "Credit Card Detail", "GET", "/dep/card/v1/external/credit-cards/{cardId}/detail", "cardId format invalid", "P1", "Negative", "400 Bad Request", "400", "Format validation"],
        ["CC_DETAIL_003", "Non-Existent Card", "Card ID doesn't exist", "Credit Card Detail", "GET", "/dep/card/v1/external/credit-cards/{cardId}/detail", "cardId not in system", "P1", "Negative", "404 Not Found", "404", "Proper error handling"],
        ["CC_DETAIL_004", "Unauthorized Access", "User accessing other user's card", "Credit Card Detail", "GET", "/dep/card/v1/external/credit-cards/{cardId}/detail", "Card belongs to different user", "P1", "Security", "403 Forbidden", "403", "Cross-customer access prevented"],
        ["CC_TRANSACTION_HISTORY_001", "Get Transaction History", "Retrieve transactions within date range", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "Valid date range, pagination", "P1", "Positive", "200 OK with transactions", "200", "Correct pagination, dates"],
        ["CC_TRANSACTION_HISTORY_002", "Missing Start Date", "Request without startDate", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "startDate parameter missing", "P1", "Negative", "400 Bad Request", "400", "Required parameter check"],
        ["CC_TRANSACTION_HISTORY_003", "Invalid Date Format", "Date in wrong format", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "Date format is DD/MM/YYYY", "P1", "Negative", "400 Bad Request", "400", "Date format validation"],
        ["CC_TRANSACTION_HISTORY_004", "Start After End Date", "startDate > endDate", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "startDate > endDate", "P1", "Negative", "400 Bad Request", "400", "Date range validation"],
        ["CC_TRANSACTION_HISTORY_005", "Date Range > 12 Months", "Query exceeds 12 months", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "Date range > 12 months", "P1", "Negative", "400 Bad Request", "400", "Date range limit check"],
        ["CC_TRANSACTION_HISTORY_006", "Invalid Page Number", "Page < 1", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "page=0", "P2", "Negative", "400 Bad Request", "400", "Page validation"],
        ["CC_TRANSACTION_HISTORY_007", "Limit > Maximum", "Limit exceeds max (100)", "Credit Card Transactions", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "limit=1000", "P2", "Negative", "400 Bad Request", "400", "Limit boundary check"],
        ["CC_TRANSACTION_SEARCH_001", "Search Transactions", "Search by keyword", "Credit Card Search", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions/search", "Valid keyword provided", "P1", "Positive", "200 OK with matching transactions", "200", "Keyword search functionality"],
        ["CC_TRANSACTION_SEARCH_002", "No Search Results", "Keyword has no matches", "Credit Card Search", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions/search", "Keyword returns empty", "P2", "Negative", "200 OK with empty array", "200", "Empty result handling"],
        ["CC_TRANSACTION_DETAIL_001", "Get Transaction Detail", "Retrieve specific transaction", "Transaction Detail", "GET", "/info/customer/v1/external/transaction/detail/{transactionId}", "Valid transactionId", "P1", "Positive", "200 OK with details", "200", "Complete transaction info"],
        ["CC_TRANSACTION_DETAIL_002", "Invalid Transaction ID", "Malformed transaction ID", "Transaction Detail", "GET", "/info/customer/v1/external/transaction/detail/{transactionId}", "transactionId format invalid", "P1", "Negative", "400 Bad Request", "400", "Format validation"],
        ["CC_TRANSACTION_DETAIL_003", "Non-Existent Transaction", "Transaction doesn't exist", "Transaction Detail", "GET", "/info/customer/v1/external/transaction/detail/{transactionId}", "transactionId not found", "P1", "Negative", "404 Not Found", "404", "Proper 404 handling"],
        ["CC_STATEMENTS_001", "Get Statements", "Retrieve account statements", "Statements", "GET", "/info/card/v1/external/credit-cards/{cardId}/statements", "Valid date range", "P1", "Positive", "200 OK with statement", "200", "Statement retrieval"],
        ["CC_STATEMENTS_002", "Missing Date Parameter", "Request without required date", "Statements", "GET", "/info/card/v1/external/credit-cards/{cardId}/statements", "Date parameter missing", "P1", "Negative", "400 Bad Request", "400", "Required field validation"],
        ["CC_REWARDS_001", "Get Rewards Points", "Retrieve rewards balance", "Rewards Program", "GET", "/info/card/v1/external/credit-cards/{cardId}/rewards", "Card has rewards program", "P1", "Positive", "200 OK with points", "200", "Rewards balance valid"],
        ["CC_REWARDS_002", "Zero Rewards", "Card with no rewards", "Rewards Program", "GET", "/info/card/v1/external/credit-cards/{cardId}/rewards", "Card has no points", "P2", "Negative", "200 OK with zero points", "200", "Zero amount handling"],
        ["CC_OFFERS_001", "Get Available Offers", "Retrieve active offers", "Card Offers", "GET", "/info/card/v1/external/credit-cards/{cardId}/offers", "Card eligible for offers", "P1", "Positive", "200 OK with offers array", "200", "Offers retrieval"],
        ["CC_BILLING_001", "Get Billing Summary", "Retrieve current billing info", "Billing Information", "GET", "/info/card/v1/external/credit-cards/{cardId}/billing", "Card has active billing", "P1", "Positive", "200 OK with billing details", "200", "Billing data validation"],
        ["CC_LIMIT_001", "Get Credit Limit", "Retrieve limit details", "Credit Limit", "GET", "/info/card/v1/external/credit-cards/{cardId}/limit", "Valid cardId", "P1", "Positive", "200 OK with limit info", "200", "Limit details present"],
        ["CC_LINKED_ACCOUNTS_001", "Get Linked Accounts", "Retrieve linked accounts", "Linked Accounts", "GET", "/info/card/v1/external/credit-cards/{cardId}/linked-accounts", "Card has linked accounts", "P1", "Positive", "200 OK with accounts", "200", "Linked accounts list"],
        ["CC_PIN_001", "Get PIN Status", "Check PIN status", "PIN Management", "GET", "/dep/card/v1/external/credit-cards/{cardId}/pin-status", "Valid cardId", "P2", "Positive", "200 OK with status", "200", "PIN status valid"],
        ["CC_BLOCK_001", "Block Card", "Block active credit card", "Block Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/block", "Card is ACTIVE, S2U capable", "P1", "Positive", "202 Accepted with S2U", "202", "S2U transaction initiated"],
        ["CC_BLOCK_002", "Block Non-Existent Card", "Card doesn't exist", "Block Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/block", "Card ID doesn't exist", "P1", "Negative", "404 Not Found", "404", "Card not found error"],
        ["CC_BLOCK_003", "Block Already Blocked", "Attempt to block blocked card", "Block Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/block", "Card status is BLOCKED", "P2", "Negative", "409 Conflict", "409", "Conflict error"],
        ["CC_BLOCK_004", "Missing Block Reason", "Request without reason field", "Block Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/block", "Reason field missing", "P1", "Negative", "400 Bad Request", "400", "Required field validation"],
        ["CC_BLOCK_005", "Invalid Block Reason", "Invalid reason value", "Block Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/block", "Reason is INVALID_REASON", "P1", "Negative", "400 Bad Request", "400", "Enum validation"],
        ["CC_UNBLOCK_001", "Unblock Card", "Unblock previously blocked card", "Unblock Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/unblock", "Card status is BLOCKED", "P1", "Positive", "202 Accepted with S2U", "202", "S2U confirmation required"],
        ["CC_UNBLOCK_002", "Unblock Active Card", "Card is already active", "Unblock Card", "POST", "/dep/card/v1/external/credit-cards/{cardId}/unblock", "Card status is ACTIVE", "P2", "Negative", "409 Conflict", "409", "Not blocked error"],
        ["CC_LIMIT_REQUEST_001", "Request Limit Increase", "Request higher credit limit", "Limit Increase Request", "POST", "/dep/card/v1/external/credit-cards/{cardId}/limit-increase", "Valid increase amount", "P1", "Positive", "202 Accepted", "202", "Request submitted"],
        ["CC_LIMIT_REQUEST_002", "Limit Below Current", "Requested < current limit", "Limit Increase Request", "POST", "/dep/card/v1/external/credit-cards/{cardId}/limit-increase", "Requested 50000 < current 100000", "P1", "Negative", "400 Bad Request", "400", "Validation error"],
        ["CC_LIMIT_REQUEST_003", "Limit Exceeds Max", "Request exceeds system max", "Limit Increase Request", "POST", "/dep/card/v1/external/credit-cards/{cardId}/limit-increase", "Requested 10000000", "P1", "Negative", "400 Bad Request", "400", "Max limit check"],
        ["CC_PIN_RESET_001", "Reset PIN", "Change card PIN", "PIN Reset", "POST", "/dep/card/v1/external/credit-cards/{cardId}/pin/reset", "Valid new PIN provided", "P1", "Positive", "200 OK", "200", "PIN reset confirmation"],
        ["CC_PIN_RESET_002", "PIN Mismatch", "New and confirm don't match", "PIN Reset", "POST", "/dep/card/v1/external/credit-cards/{cardId}/pin/reset", "newPin != confirmPin", "P1", "Negative", "400 Bad Request", "400", "PIN match validation"],
        ["CC_PIN_RESET_003", "Weak PIN", "PIN doesn't meet requirements", "PIN Reset", "POST", "/dep/card/v1/external/credit-cards/{cardId}/pin/reset", "PIN = 1111", "P1", "Negative", "400 Bad Request", "400", "PIN strength check"],
        ["CC_ACTIVATE_001", "Activate Card", "Activate new credit card", "Card Activation", "POST", "/dep/card/v1/external/credit-cards/{cardId}/activate", "Valid OTP provided", "P1", "Positive", "200 OK", "200", "Card activated"],
        ["CC_ACTIVATE_002", "Invalid OTP", "Incorrect OTP provided", "Card Activation", "POST", "/dep/card/v1/external/credit-cards/{cardId}/activate", "OTP = 999999", "P1", "Negative", "400 Bad Request", "400", "OTP validation"],
        ["CC_ACTIVATE_003", "Expired OTP", "OTP has expired", "Card Activation", "POST", "/dep/card/v1/external/credit-cards/{cardId}/activate", "OTP timestamp expired", "P1", "Negative", "400 Bad Request", "400", "OTP expiry check"],
        ["CC_LINK_ACCOUNT_001", "Link Account", "Link account for auto-payment", "Link Account", "POST", "/dep/card/v1/external/credit-cards/{cardId}/link-account", "Valid account and auto-pay type", "P1", "Positive", "200 OK", "200", "Account linked"],
        ["CC_LINK_ACCOUNT_002", "Account Already Linked", "Account is already linked", "Link Account", "POST", "/dep/card/v1/external/credit-cards/{cardId}/link-account", "Account already linked", "P2", "Negative", "409 Conflict", "409", "Duplicate link error"],
        ["CC_LINK_ACCOUNT_003", "Invalid Payment Type", "Invalid auto-payment type", "Link Account", "POST", "/dep/card/v1/external/credit-cards/{cardId}/link-account", "paymentType = INVALID", "P1", "Negative", "400 Bad Request", "400", "Enum validation"],
        ["CC_UNLINK_ACCOUNT_001", "Unlink Account", "Remove linked bank account", "Unlink Account", "POST", "/dep/card/v1/external/credit-cards/{cardId}/unlink-account", "Account is linked", "P1", "Positive", "200 OK", "200", "Account unlinked"],
        ["CC_UNLINK_ACCOUNT_002", "No Linked Account", "No account to unlink", "Unlink Account", "POST", "/dep/card/v1/external/credit-cards/{cardId}/unlink-account", "No linked account", "P2", "Negative", "404 Not Found", "404", "Not found error"],
        ["CC_PAYMENT_001", "Make Payment", "Pay credit card balance", "Card Payment", "POST", "/dep/card/v1/external/credit-cards/{cardId}/payment", "Valid amount and method", "P1", "Positive", "202 Accepted", "202", "Payment initiated"],
        ["CC_PAYMENT_002", "Invalid Amount", "Negative or invalid amount", "Card Payment", "POST", "/dep/card/v1/external/credit-cards/{cardId}/payment", "amount = -1000", "P1", "Negative", "400 Bad Request", "400", "Amount validation"],
        ["CC_PAYMENT_003", "Insufficient Funds", "Amount > account balance", "Card Payment", "POST", "/dep/card/v1/external/credit-cards/{cardId}/payment", "amount > balance", "P1", "Negative", "422 Unprocessable", "422", "Insufficient funds error"],
        ["CC_PAYMENT_004", "Zero Amount", "Payment of zero", "Card Payment", "POST", "/dep/card/v1/external/credit-cards/{cardId}/payment", "amount = 0", "P2", "Negative", "400 Bad Request", "400", "Amount > 0 check"],
        ["CC_NICKNAME_001", "Update Nickname", "Change card nickname", "Card Nickname", "PATCH", "/dep/card/v1/external/credit-cards/{cardId}/nickname", "Valid nickname provided", "P1", "Positive", "200 OK", "200", "Nickname updated"],
        ["CC_NICKNAME_002", "Empty Nickname", "Nickname is empty string", "Card Nickname", "PATCH", "/dep/card/v1/external/credit-cards/{cardId}/nickname", "nickname = ''", "P1", "Negative", "400 Bad Request", "400", "Empty field check"],
        ["CC_NICKNAME_003", "Nickname Too Long", "Nickname > 50 chars", "Card Nickname", "PATCH", "/dep/card/v1/external/credit-cards/{cardId}/nickname", "50 character limit", "P2", "Negative", "400 Bad Request", "400", "Length validation"],
        ["CC_NICKNAME_DELETE_001", "Delete Nickname", "Remove custom nickname", "Card Nickname", "DELETE", "/dep/card/v1/external/credit-cards/{cardId}/nickname", "Nickname exists", "P2", "Positive", "204 No Content", "204", "Nickname deleted"],
        ["CC_NICKNAME_DELETE_002", "No Nickname To Delete", "No nickname exists", "Card Nickname", "DELETE", "/dep/card/v1/external/credit-cards/{cardId}/nickname", "No nickname set", "P2", "Negative", "404 Not Found", "404", "Not found error"],
        ["CC_SECURITY_001", "SQL Injection", "Inject SQL in search keyword", "Security", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions/search", "keyword = '; DROP TABLE'", "P1", "Security", "400 Bad Request", "400", "Injection prevention"],
        ["CC_SECURITY_002", "XSS Attack", "XSS attempt in nickname", "Security", "PATCH", "/dep/card/v1/external/credit-cards/{cardId}/nickname", "nickname = <script>", "P1", "Security", "400 Bad Request", "400", "XSS prevention"],
        ["CC_SECURITY_003", "Path Traversal", "Path traversal in cardId", "Security", "GET", "/dep/card/v1/external/credit-cards/{cardId}/detail", "cardId = ../../../etc/passwd", "P1", "Security", "400 Bad Request", "400", "Path traversal prevention"],
        ["CC_PERFORMANCE_001", "Dashboard Response Time", "Response time < 500ms", "Performance", "GET", "/dep/dashboard/v1/external/credit-cards", "Normal load", "P2", "Performance", "Response < 500ms", "200", "Performance SLA"],
        ["CC_PERFORMANCE_002", "Large Transaction Query", "High volume transaction query", "Performance", "GET", "/info/card/v1/external/credit-cards/{cardId}/transactions", "Large date range", "P2", "Performance", "Response < 2000ms", "200", "Performance SLA"],
        ["CC_INTEGRATION_001", "Block Card Propagation", "Blocked card removed from dashboard", "Integration", "GET/POST", "Multiple endpoints", "Block then GET", "P1", "Integration", "Card not in dashboard", "200/202", "Data consistency"],
        ["CC_INTEGRATION_002", "Payment Reflects Balance", "Payment updates balance correctly", "Integration", "GET/POST", "Multiple endpoints", "Pay then GET", "P1", "Integration", "Balance reflects payment", "200/202", "Data consistency"],
    ]
    
    # Write test case data
    for row_num, data in enumerate(test_cases, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Alternate row coloring
            if row_num % 2 == 0:
                cell.fill = alt_row_fill
    
    # Set row height
    for row in range(1, len(test_cases) + 2):
        ws.row_dimensions[row].height = 30
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save("c:\\Users\\aahmadkamar\\Maybank\\FSD Parser\\artifacts\\Test_Case\\Excel\\Account_Dashboard_Credit_Card_DDD_Test_Cases.xlsx")
    print(f"✓ Test Cases Excel file generated: {len(test_cases)} test cases")

def create_test_data_excel():
    """Generate test data Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"
    
    # Define styles
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_row_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define columns
    headers = [
        "Test Data ID",
        "Related Test Case ID",
        "Description",
        "Input Parameters",
        "Expected Result",
        "Positive/Negative",
        "Remarks"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30
    
    # Test data entries
    test_data = [
        ["TD_CC_001", "CC_DASHBOARD_001", "Valid dashboard request", json.dumps({"endpoint": "/dep/dashboard/v1/external/credit-cards", "headers": "Bearer token, client-id"}), "Returns 200 with cards array", "Positive", "Happy path"],
        ["TD_CC_002", "CC_DASHBOARD_002", "No active cards", json.dumps({"endpoint": "/dep/dashboard/v1/external/credit-cards", "user_cards": 0}), "Returns 200 with empty cards array", "Negative", "User has no cards"],
        ["TD_CC_003", "CC_DASHBOARD_003", "Missing auth token", json.dumps({"headers": "client-id, env (no Bearer)"}), "Returns 401 Unauthorized", "Negative", "Security: Auth required"],
        ["TD_CC_004", "CC_DASHBOARD_004", "Expired token", json.dumps({"auth": "Bearer <EXPIRED_TOKEN>"}), "Returns 401 Token Expired", "Negative", "Security: Token validation"],
        ["TD_CC_005", "CC_DASHBOARD_005", "Invalid client-id", json.dumps({"client-id": "invalid-client"}), "Returns 403 Forbidden", "Negative", "Security: Client auth"],
        ["TD_CC_006", "CC_DASHBOARD_006", "Missing client-id", json.dumps({"headers": "Bearer token, env"}), "Returns 400 Bad Request", "Negative", "Validation: Required header"],
        ["TD_CC_007", "CC_DASHBOARD_007", "Invalid environment", json.dumps({"env": "INVALID"}), "Returns 400 Bad Request", "Negative", "Validation: Enum check"],
        ["TD_CC_008", "CC_DETAIL_001", "Get card details", json.dumps({"cardId": "CC_001", "operation": "GET detail"}), "Returns 200 with all fields", "Positive", "Happy path"],
        ["TD_CC_009", "CC_DETAIL_002", "Invalid card ID", json.dumps({"cardId": "INVALID123"}), "Returns 400 Bad Request", "Negative", "Validation: Format"],
        ["TD_CC_010", "CC_DETAIL_003", "Non-existent card", json.dumps({"cardId": "CC_999999"}), "Returns 404 Not Found", "Negative", "Error: Not found"],
        ["TD_CC_011", "CC_DETAIL_004", "Cross-user access", json.dumps({"cardId": "OTHER_USER_CARD"}), "Returns 403 Forbidden", "Negative", "Security: Access control"],
        ["TD_CC_012", "CC_TRANSACTION_HISTORY_001", "Get transactions", json.dumps({"cardId": "CC_001", "startDate": "2026-06-17", "endDate": "2026-07-17", "page": 1, "limit": 10}), "Returns 200 with transactions", "Positive", "Pagination: Valid range"],
        ["TD_CC_013", "CC_TRANSACTION_HISTORY_002", "Missing start date", json.dumps({"endDate": "2026-07-17"}), "Returns 400 Bad Request", "Negative", "Validation: Required param"],
        ["TD_CC_014", "CC_TRANSACTION_HISTORY_003", "Invalid date format", json.dumps({"startDate": "17/06/2026", "endDate": "2026-07-17"}), "Returns 400 Bad Request", "Negative", "Validation: Date format"],
        ["TD_CC_015", "CC_TRANSACTION_HISTORY_004", "Start after end", json.dumps({"startDate": "2026-07-17", "endDate": "2026-06-17"}), "Returns 400 Bad Request", "Negative", "Validation: Date logic"],
        ["TD_CC_016", "CC_TRANSACTION_HISTORY_005", "Date range > 12mo", json.dumps({"startDate": "2024-01-01", "endDate": "2026-07-17"}), "Returns 400 Bad Request", "Negative", "Validation: Date range"],
        ["TD_CC_017", "CC_TRANSACTION_HISTORY_006", "Invalid page", json.dumps({"page": 0}), "Returns 400 Bad Request", "Negative", "Validation: Page > 0"],
        ["TD_CC_018", "CC_TRANSACTION_HISTORY_007", "Limit > max", json.dumps({"limit": 1000}), "Returns 400 or caps to 100", "Negative", "Boundary: Max limit"],
        ["TD_CC_019", "CC_TRANSACTION_SEARCH_001", "Search with keyword", json.dumps({"keyword": "pizza", "startDate": "2026-06-17", "endDate": "2026-07-17"}), "Returns 200 with matching txns", "Positive", "Search functionality"],
        ["TD_CC_020", "CC_TRANSACTION_SEARCH_002", "No search results", json.dumps({"keyword": "nonexistent"}), "Returns 200 empty array", "Negative", "Empty result handling"],
        ["TD_CC_021", "CC_TRANSACTION_DETAIL_001", "Get transaction detail", json.dumps({"transactionId": "TXN_CC_001"}), "Returns 200 with details", "Positive", "Happy path"],
        ["TD_CC_022", "CC_TRANSACTION_DETAIL_002", "Invalid txn ID", json.dumps({"transactionId": "INVALID"}), "Returns 400 Bad Request", "Negative", "Validation: Format"],
        ["TD_CC_023", "CC_TRANSACTION_DETAIL_003", "Txn not found", json.dumps({"transactionId": "TXN_FAKE"}), "Returns 404 Not Found", "Negative", "Error: Not found"],
        ["TD_CC_024", "CC_STATEMENTS_001", "Get statements", json.dumps({"cardId": "CC_001", "fromDate": "2026-01-01", "toDate": "2026-07-17"}), "Returns 200 with statement", "Positive", "Document retrieval"],
        ["TD_CC_025", "CC_STATEMENTS_002", "Missing date param", json.dumps({"cardId": "CC_001"}), "Returns 400 Bad Request", "Negative", "Validation: Required"],
        ["TD_CC_026", "CC_REWARDS_001", "Get rewards", json.dumps({"cardId": "CC_001"}), "Returns 200 with points balance", "Positive", "Happy path"],
        ["TD_CC_027", "CC_REWARDS_002", "Zero rewards", json.dumps({"cardId": "CC_001", "points": 0}), "Returns 200 with 0 points", "Negative", "Empty/zero handling"],
        ["TD_CC_028", "CC_OFFERS_001", "Get offers", json.dumps({"cardId": "CC_001"}), "Returns 200 with offers array", "Positive", "Data retrieval"],
        ["TD_CC_029", "CC_BILLING_001", "Get billing", json.dumps({"cardId": "CC_001"}), "Returns 200 with billing info", "Positive", "Happy path"],
        ["TD_CC_030", "CC_LIMIT_001", "Get credit limit", json.dumps({"cardId": "CC_001"}), "Returns 200 with limit details", "Positive", "Happy path"],
        ["TD_CC_031", "CC_LINKED_ACCOUNTS_001", "Get linked accounts", json.dumps({"cardId": "CC_001"}), "Returns 200 with accounts", "Positive", "Happy path"],
        ["TD_CC_032", "CC_PIN_001", "Get PIN status", json.dumps({"cardId": "CC_001"}), "Returns 200 with status", "Positive", "Happy path"],
        ["TD_CC_033", "CC_BLOCK_001", "Block card", json.dumps({"cardId": "CC_001", "reason": "LOST_CARD", "operation": "POST block"}), "Returns 202 with S2U", "Positive", "S2U required"],
        ["TD_CC_034", "CC_BLOCK_002", "Card not found", json.dumps({"cardId": "CC_999999"}), "Returns 404 Not Found", "Negative", "Error: Not found"],
        ["TD_CC_035", "CC_BLOCK_003", "Already blocked", json.dumps({"cardId": "CC_001", "status": "BLOCKED"}), "Returns 409 Conflict", "Negative", "State error"],
        ["TD_CC_036", "CC_BLOCK_004", "Missing reason", json.dumps({"cardId": "CC_001"}), "Returns 400 Bad Request", "Negative", "Validation: Required"],
        ["TD_CC_037", "CC_BLOCK_005", "Invalid reason", json.dumps({"reason": "INVALID_REASON"}), "Returns 400 Bad Request", "Negative", "Validation: Enum"],
        ["TD_CC_038", "CC_UNBLOCK_001", "Unblock card", json.dumps({"cardId": "CC_001", "status": "BLOCKED"}), "Returns 202 with S2U", "Positive", "S2U required"],
        ["TD_CC_039", "CC_UNBLOCK_002", "Card not blocked", json.dumps({"cardId": "CC_001", "status": "ACTIVE"}), "Returns 409 Conflict", "Negative", "State error"],
        ["TD_CC_040", "CC_LIMIT_REQUEST_001", "Request limit increase", json.dumps({"cardId": "CC_001", "requestedLimit": 150000, "current": 100000}), "Returns 202 Accepted", "Positive", "Request submitted"],
        ["TD_CC_041", "CC_LIMIT_REQUEST_002", "Limit below current", json.dumps({"requestedLimit": 50000, "current": 100000}), "Returns 400 Bad Request", "Negative", "Validation: Logic"],
        ["TD_CC_042", "CC_LIMIT_REQUEST_003", "Limit exceeds max", json.dumps({"requestedLimit": 10000000}), "Returns 400 Bad Request", "Negative", "Boundary: Max limit"],
        ["TD_CC_043", "CC_PIN_RESET_001", "Reset PIN", json.dumps({"cardId": "CC_001", "newPin": "****", "confirmPin": "****"}), "Returns 200 OK", "Positive", "Happy path"],
        ["TD_CC_044", "CC_PIN_RESET_002", "PIN mismatch", json.dumps({"newPin": "1234", "confirmPin": "5678"}), "Returns 400 Bad Request", "Negative", "Validation: Match"],
        ["TD_CC_045", "CC_PIN_RESET_003", "Weak PIN", json.dumps({"newPin": "1111", "confirmPin": "1111"}), "Returns 400 Bad Request", "Negative", "Validation: Strength"],
        ["TD_CC_046", "CC_ACTIVATE_001", "Activate card", json.dumps({"cardId": "CC_001", "otp": "123456"}), "Returns 200 OK", "Positive", "Happy path"],
        ["TD_CC_047", "CC_ACTIVATE_002", "Invalid OTP", json.dumps({"otp": "999999"}), "Returns 400 Bad Request", "Negative", "Validation: OTP"],
        ["TD_CC_048", "CC_ACTIVATE_003", "Expired OTP", json.dumps({"otp": "123456", "expired": True}), "Returns 400 Bad Request", "Negative", "Validation: Expiry"],
        ["TD_CC_049", "CC_LINK_ACCOUNT_001", "Link account", json.dumps({"cardId": "CC_001", "accountId": "123456789012", "paymentType": "FULL_BALANCE"}), "Returns 200 OK", "Positive", "Happy path"],
        ["TD_CC_050", "CC_LINK_ACCOUNT_002", "Already linked", json.dumps({"cardId": "CC_001", "accountId": "123456789012", "linked": True}), "Returns 409 Conflict", "Negative", "State error"],
        ["TD_CC_051", "CC_LINK_ACCOUNT_003", "Invalid payment type", json.dumps({"paymentType": "INVALID"}), "Returns 400 Bad Request", "Negative", "Validation: Enum"],
        ["TD_CC_052", "CC_UNLINK_ACCOUNT_001", "Unlink account", json.dumps({"cardId": "CC_001", "linked": True}), "Returns 200 OK", "Positive", "Happy path"],
        ["TD_CC_053", "CC_UNLINK_ACCOUNT_002", "No linked account", json.dumps({"cardId": "CC_001", "linked": False}), "Returns 404 Not Found", "Negative", "Error: Not found"],
        ["TD_CC_054", "CC_PAYMENT_001", "Make payment", json.dumps({"cardId": "CC_001", "amount": 5000, "method": "DEBIT_TRANSFER"}), "Returns 202 Accepted", "Positive", "S2U may be required"],
        ["TD_CC_055", "CC_PAYMENT_002", "Invalid amount", json.dumps({"amount": -1000}), "Returns 400 Bad Request", "Negative", "Validation: Amount"],
        ["TD_CC_056", "CC_PAYMENT_003", "Insufficient funds", json.dumps({"amount": 50000, "balance": 1000}), "Returns 422 Unprocessable", "Negative", "Business rule"],
        ["TD_CC_057", "CC_PAYMENT_004", "Zero amount", json.dumps({"amount": 0}), "Returns 400 Bad Request", "Negative", "Validation: > 0"],
        ["TD_CC_058", "CC_NICKNAME_001", "Update nickname", json.dumps({"cardId": "CC_001", "nickname": "Shopping Card"}), "Returns 200 OK", "Positive", "Happy path"],
        ["TD_CC_059", "CC_NICKNAME_002", "Empty nickname", json.dumps({"nickname": ""}), "Returns 400 Bad Request", "Negative", "Validation: Non-empty"],
        ["TD_CC_060", "CC_NICKNAME_003", "Nickname too long", json.dumps({"nickname": "x" * 51}), "Returns 400 Bad Request", "Negative", "Validation: Length"],
        ["TD_CC_061", "CC_NICKNAME_DELETE_001", "Delete nickname", json.dumps({"cardId": "CC_001", "has_nickname": True}), "Returns 204 No Content", "Positive", "Happy path"],
        ["TD_CC_062", "CC_NICKNAME_DELETE_002", "No nickname", json.dumps({"cardId": "CC_001", "has_nickname": False}), "Returns 404 Not Found", "Negative", "Error: Not found"],
        ["TD_CC_063", "CC_SECURITY_001", "SQL injection", json.dumps({"keyword": "pizza'; DROP TABLE;"}), "Returns 400 Bad Request", "Negative", "Security: Injection"],
        ["TD_CC_064", "CC_SECURITY_002", "XSS attack", json.dumps({"nickname": "<script>alert('XSS')</script>"}), "Returns 400 Bad Request", "Negative", "Security: XSS"],
        ["TD_CC_065", "CC_SECURITY_003", "Path traversal", json.dumps({"cardId": "../../../etc/passwd"}), "Returns 400 Bad Request", "Negative", "Security: Path traversal"],
    ]
    
    # Write test data
    for row_num, data in enumerate(test_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Alternate row coloring
            if row_num % 2 == 0:
                cell.fill = alt_row_fill
    
    # Set row height
    for row in range(1, len(test_data) + 2):
        ws.row_dimensions[row].height = 30
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save("c:\\Users\\aahmadkamar\\Maybank\\FSD Parser\\artifacts\\Test_Case\\Excel\\Account_Dashboard_Credit_Card_DDD_Test_Data.xlsx")
    print(f"✓ Test Data Excel file generated: {len(test_data)} test data entries")

if __name__ == "__main__":
    create_test_cases_excel()
    create_test_data_excel()
    print("\n✓ All Excel files generated successfully!")
    print(f"✓ Location: artifacts/Test_Case/Excel/")
