#!/usr/bin/env python3
"""
API Validation Test Case Generator
Extracts all APIs from a DDD file and generates comprehensive test cases
"""

import openpyxl
import json
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
import re

class TestCaseGenerator:
    def __init__(self, ddd_file_path):
        self.ddd_file_path = ddd_file_path
        self.workbook = openpyxl.load_workbook(ddd_file_path)
        self.apis = {}
        self.test_cases = []
        self.api_count = 0
        self.positive_count = 0
        self.negative_count = 0

    def extract_apis(self):
        """Extract all APIs from the DDD file"""
        print(f"Loading DDD from: {self.ddd_file_path}")
        print(f"Available sheets: {self.workbook.sheetnames}\n")
        
        skip_sheets = ['Summary', 'Overview', 'Index', 'Guidelines', 'Template']
        
        for sheet_name in self.workbook.sheetnames:
            if any(skip in sheet_name for skip in skip_sheets):
                continue
            
            ws = self.workbook[sheet_name]
            api_data = self._extract_sheet_data(sheet_name, ws)
            
            if api_data:
                self.apis[sheet_name] = api_data
                self.api_count += 1
                print(f"✓ Extracted: {sheet_name}")
        
        print(f"\nTotal APIs extracted: {self.api_count}\n")
        return self.apis

    def _extract_sheet_data(self, sheet_name, ws):
        """Extract API data from a single sheet"""
        data = {
            'sheet_name': sheet_name,
            'endpoint': f'/{sheet_name}',
            'method': 'POST',
            'description': '',
            'request_fields': [],
            'response_fields': [],
        }
        
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            
            row_str = str(row[0]).lower()
            
            if 'endpoint' in row_str or 'url' in row_str:
                data['endpoint'] = str(row[1] if len(row) > 1 else '').strip() or data['endpoint']
            
            if 'method' in row_str and 'http' in row_str:
                data['method'] = str(row[1] if len(row) > 1 else 'POST').strip().upper()
            
            if 'description' in row_str:
                data['description'] = str(row[1] if len(row) > 1 else '').strip()
        
        data['request_fields'] = self._extract_parameters(ws, 'request')
        data['response_fields'] = self._extract_parameters(ws, 'response')
        
        return data if data['endpoint'] else None

    def _extract_parameters(self, ws, param_type):
        """Extract parameters from sheet"""
        parameters = []
        found = False
        
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            
            row_text = str(row[0]).lower() if row[0] else ''
            
            if param_type in row_text and 'parameter' in row_text:
                found = True
                continue
            
            if found and row[0]:
                if any(keyword in row_text for keyword in ['response', 'request', 'status']):
                    break
                
                if len(row) > 1:
                    param = {
                        'name': str(row[0]).strip(),
                        'type': str(row[1]).strip() if len(row) > 1 else 'string',
                        'required': str(row[2]).strip().lower() == 'yes' if len(row) > 2 else False,
                        'description': str(row[3]).strip() if len(row) > 3 else '',
                    }
                    parameters.append(param)
        
        return parameters

    def generate_test_cases(self):
        """Generate test cases for all APIs"""
        self.test_cases = []
        
        for api_name, api_data in self.apis.items():
            api_tests = self._generate_api_tests(api_name, api_data)
            self.test_cases.extend(api_tests)
        
        print(f"Test cases generated:")
        print(f"  Positive: {self.positive_count}")
        print(f"  Negative: {self.negative_count}")
        print(f"  Total: {len(self.test_cases)}\n")
        
        return self.test_cases

    def _generate_api_tests(self, api_name, api_data):
        """Generate test cases for a single API"""
        tests = []
        api_id = re.sub(r'[^a-zA-Z0-9_]', '', api_name.replace(' ', '_'))[:10].upper()
        counter = 0
        
        endpoint = api_data.get('endpoint', f'/{api_name}')
        method = api_data.get('method', 'POST')
        req_fields = api_data.get('request_fields', [])
        mandatory = [f for f in req_fields if f.get('required')]
        
        # POSITIVE TEST CASES
        # TC1: All mandatory fields
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Valid request with all mandatory fields',
            'type': 'Positive',
            'category': 'Field Validation',
            'pre_conditions': 'Valid authentication token; User has required permissions',
            'steps': f'Set Authorization header → Send {method} {endpoint} with all mandatory fields → Verify 200/201 response',
            'input': 'Valid JSON with all mandatory fields populated',
            'expected': 'HTTP 200/201; Valid response schema; All response fields present',
            'priority': 'High'
        })
        self.positive_count += 1
        
        # TC2: Optional fields
        if len(req_fields) > len(mandatory):
            counter += 1
            tests.append({
                'tc_id': f'TC_{api_id}_{counter:03d}',
                'title': 'Valid request with optional fields',
                'type': 'Positive',
                'category': 'Field Validation',
                'pre_conditions': 'Valid authentication token',
                'steps': f'Send {method} {endpoint} with all fields (mandatory + optional)',
                'input': 'Valid JSON with all mandatory and optional fields',
                'expected': 'HTTP 200/201; Optional fields processed correctly',
                'priority': 'Medium'
            })
            self.positive_count += 1
        
        # TC3: Boundary values
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Valid boundary values (min/max)',
            'type': 'Positive',
            'category': 'Boundary Testing',
            'pre_conditions': 'Knowledge of min/max from DDD',
            'steps': f'Send {method} {endpoint} with min/max values for numeric/string fields',
            'input': 'Numeric fields at min/max boundaries; String fields at min/max length',
            'expected': 'HTTP 200/201; Boundary values accepted',
            'priority': 'High'
        })
        self.positive_count += 1
        
        # NEGATIVE TEST CASES
        
        # TC-N1 to N3: Missing mandatory fields
        for i, field in enumerate(mandatory[:3]):
            counter += 1
            tests.append({
                'tc_id': f'TC_{api_id}_{counter:03d}',
                'title': f'Missing mandatory field: {field["name"]}',
                'type': 'Negative',
                'category': 'Field Validation',
                'pre_conditions': 'Valid token available',
                'steps': f'Send {method} {endpoint} WITHOUT "{field["name"]}" field',
                'input': f'Omit "{field["name"]}" from payload',
                'expected': f'HTTP 400; Error: "Missing required field: {field["name"]}"',
                'priority': 'High'
            })
            self.negative_count += 1
        
        # TC-N4 to N5: Invalid data types
        for i, field in enumerate(req_fields[:2]):
            counter += 1
            invalid_type = 'string' if field['type'] != 'string' else 'integer'
            tests.append({
                'tc_id': f'TC_{api_id}_{counter:03d}',
                'title': f'Invalid data type for {field["name"]}: {invalid_type}',
                'type': 'Negative',
                'category': 'Data Type Validation',
                'pre_conditions': 'Valid token',
                'steps': f'Send {method} {endpoint} with {invalid_type} value for {field["name"]}',
                'input': f'"{field["name"]}": "invalid" (instead of {field["type"]})',
                'expected': f'HTTP 400; Error: "Invalid data type for {field["name"]}"',
                'priority': 'High'
            })
            self.negative_count += 1
        
        # TC-N6: Invalid format
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Invalid date/email format',
            'type': 'Negative',
            'category': 'Format Validation',
            'pre_conditions': 'Valid token',
            'steps': f'Send {method} {endpoint} with malformed date/email',
            'input': '{"date": "2024-13-45", "email": "invalid-email"}',
            'expected': 'HTTP 400; Error: "Invalid format"',
            'priority': 'Medium'
        })
        self.negative_count += 1
        
        # TC-N7: Out of range
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Out of range values',
            'type': 'Negative',
            'category': 'Boundary Testing',
            'pre_conditions': 'Knowledge of ranges',
            'steps': f'Send {method} {endpoint} with values below min or above max',
            'input': 'Numeric value outside [min-max] range',
            'expected': 'HTTP 400; Error: "Value out of range"',
            'priority': 'Medium'
        })
        self.negative_count += 1
        
        # TC-N8: Missing auth token
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Missing authentication token',
            'type': 'Negative',
            'category': 'Authentication',
            'pre_conditions': 'API requires authentication',
            'steps': f'Send {method} {endpoint} WITHOUT Authorization header',
            'input': 'Valid payload; NO Authorization header',
            'expected': 'HTTP 401; Error: "Missing authentication token"',
            'priority': 'High'
        })
        self.negative_count += 1
        
        # TC-N9: Invalid token
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Invalid/expired token',
            'type': 'Negative',
            'category': 'Authentication',
            'pre_conditions': 'Invalid token available',
            'steps': f'Send {method} {endpoint} with invalid/expired token',
            'input': 'Authorization: Bearer invalid_token_12345',
            'expected': 'HTTP 401; Error: "Invalid or expired token"',
            'priority': 'High'
        })
        self.negative_count += 1
        
        # TC-N10: Malformed JSON
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Malformed JSON payload',
            'type': 'Negative',
            'category': 'Payload Validation',
            'pre_conditions': 'Valid token',
            'steps': f'Send {method} {endpoint} with invalid JSON',
            'input': '{{"field": "value", }}  // Trailing comma',
            'expected': 'HTTP 400; Error: "Invalid JSON format"',
            'priority': 'Medium'
        })
        self.negative_count += 1
        
        # TC-N11: SQL Injection
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'SQL Injection payload',
            'type': 'Negative',
            'category': 'Security',
            'pre_conditions': 'Valid token',
            'steps': f'Send {method} {endpoint} with SQL injection in text field',
            'input': '{"name": "\' OR \'1\'=\'1\'; DROP TABLE users; --"}',
            'expected': 'HTTP 400/422; Payload rejected/sanitized',
            'priority': 'High'
        })
        self.negative_count += 1
        
        # TC-N12: XSS payload
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'XSS payload in input',
            'type': 'Negative',
            'category': 'Security',
            'pre_conditions': 'Valid token',
            'steps': f'Send {method} {endpoint} with XSS payload',
            'input': '{"message": "<script>alert(\'XSS\')</script>"}',
            'expected': 'HTTP 400/422; Payload rejected/sanitized',
            'priority': 'High'
        })
        self.negative_count += 1
        
        # TC-N13: Empty payload
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Empty JSON payload',
            'type': 'Negative',
            'category': 'Payload Validation',
            'pre_conditions': 'Valid token',
            'steps': f'Send {method} {endpoint} with empty JSON',
            'input': '{}',
            'expected': 'HTTP 400; Error: "Missing required fields"',
            'priority': 'Medium'
        })
        self.negative_count += 1
        
        # TC-N14: Invalid HTTP method
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': f'Invalid HTTP method (GET instead of {method})',
            'type': 'Negative',
            'category': 'HTTP Method',
            'pre_conditions': f'API expects {method}',
            'steps': f'Send GET {endpoint} (should be {method})',
            'input': f'Valid payload sent via GET',
            'expected': 'HTTP 405; Error: "Method Not Allowed"',
            'priority': 'Medium'
        })
        self.negative_count += 1
        
        # TC-N15: Rate limiting
        counter += 1
        tests.append({
            'tc_id': f'TC_{api_id}_{counter:03d}',
            'title': 'Rate limiting behavior',
            'type': 'Negative',
            'category': 'Rate Limiting',
            'pre_conditions': 'Rate limit defined in DDD',
            'steps': f'Send 100+ requests rapidly to {endpoint}',
            'input': 'Rapid succession of valid requests',
            'expected': 'HTTP 429; Error: "Too Many Requests"; Retry-After header',
            'priority': 'Medium'
        })
        self.negative_count += 1
        
        return tests

    def generate_markdown(self, output_path):
        """Generate markdown file"""
        os.makedirs(output_path, exist_ok=True)
        # Use DDD filename to create output filename
        ddd_name = os.path.splitext(os.path.basename(self.ddd_file_path))[0]
        md_file = os.path.join(output_path, f'{ddd_name}_Test_Cases.md')
        
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write('# API Validation Test Cases\n\n')
            f.write(f'**Generated**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')
            f.write('**Reference**: PT_Scan_and_Pay_DDD_API_Spec_v2_working.xlsx\n\n')
            
            # SUMMARY
            f.write('## TEST SUMMARY\n\n')
            f.write(f'### Statistics\n')
            f.write(f'- **Total APIs**: {self.api_count}\n')
            f.write(f'- **Total Test Cases**: {len(self.test_cases)}\n')
            f.write(f'  - Positive: {self.positive_count}\n')
            f.write(f'  - Negative: {self.negative_count}\n\n')
            
            # APIs List
            f.write('### APIs Identified\n\n')
            for i, (name, data) in enumerate(self.apis.items(), 1):
                f.write(f'{i}. **{name}**\n')
                f.write(f'   - Method: `{data["method"]}`\n')
                f.write(f'   - Endpoint: `{data["endpoint"]}`\n')
                f.write(f'   - Request Fields: {len(data["request_fields"])}\n')
                f.write(f'   - Response Fields: {len(data["response_fields"])}\n\n')
            
            # Coverage Areas
            f.write('### Coverage Areas\n\n')
            f.write('- ✓ Field Validation (mandatory, optional, types)\n')
            f.write('- ✓ Boundary Testing (min/max values)\n')
            f.write('- ✓ Authentication & Authorization\n')
            f.write('- ✓ Error Handling & Status Codes\n')
            f.write('- ✓ Security (SQL Injection, XSS)\n')
            f.write('- ✓ Rate Limiting\n')
            f.write('- ✓ Payload Validation\n')
            f.write('- ✓ HTTP Method Validation\n\n')
            
            # Assumptions
            f.write('### Assumptions\n\n')
            f.write('1. Bearer Token authentication\n')
            f.write('2. Content-Type: application/json\n')
            f.write('3. ISO 8601 timestamps\n')
            f.write('4. Standard HTTP status codes\n')
            f.write('5. All monetary values in decimal format\n\n')
            
            # TEST CASES
            f.write('---\n\n')
            f.write('## DETAILED TEST CASES\n\n')
            
            for api_name, api_data in self.apis.items():
                f.write(f'### API: {api_name}\n\n')
                f.write(f'**Method**: {api_data["method"]} | ')
                f.write(f'**Endpoint**: `{api_data["endpoint"]}`\n\n')
                
                if api_data.get('description'):
                    f.write(f'**Description**: {api_data["description"]}\n\n')
                
                # Request Fields Table
                if api_data.get('request_fields'):
                    f.write('#### Request Fields\n\n')
                    f.write('| Field | Type | Required | Description |\n')
                    f.write('|-------|------|----------|-------------|\n')
                    for field in api_data['request_fields']:
                        req = '✓' if field.get('required') else '✗'
                        f.write(f"| {field['name']} | {field['type']} | {req} | {field.get('description', '')} |\n")
                    f.write('\n')
                
                # Response Fields Table
                if api_data.get('response_fields'):
                    f.write('#### Response Fields\n\n')
                    f.write('| Field | Type | Description |\n')
                    f.write('|-------|------|-------------|\n')
                    for field in api_data['response_fields']:
                        f.write(f"| {field['name']} | {field['type']} | {field.get('description', '')} |\n")
                    f.write('\n')
                
                # Test Cases Table
                api_id = re.sub(r'[^a-zA-Z0-9_]', '', api_name.replace(' ', '_'))[:10].upper()
                api_tests = [tc for tc in self.test_cases if tc['tc_id'].startswith(f'TC_{api_id}')]
                f.write('#### Test Cases\n\n')
                f.write('| TC ID | Title | Type | Category | Pre-Conditions | Steps | Input | Expected | Priority |\n')
                f.write('|-------|-------|------|----------|---|---|---|---|---|\n')
                
                for tc in api_tests:
                    pre_cond = tc['pre_conditions'][:40].replace('|', '∣')
                    steps = tc['steps'][:50].replace('|', '∣')
                    input_data = tc['input'][:40].replace('|', '∣')
                    expected = tc['expected'][:40].replace('|', '∣')
                    f.write(f"| {tc['tc_id']} | {tc['title']} | {tc['type']} | {tc['category']} | {pre_cond}... | {steps}... | {input_data}... | {expected}... | {tc['priority']} |\n")
                
                f.write('\n---\n\n')
        
        print(f"✓ Markdown: {md_file}")
        return md_file

    def generate_excel(self, output_path):
        """Generate Excel file"""
        os.makedirs(output_path, exist_ok=True)
        # Use DDD filename to create output filename
        ddd_name = os.path.splitext(os.path.basename(self.ddd_file_path))[0]
        excel_file = os.path.join(output_path, f'{ddd_name}_Test_Cases.xlsx')
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Summary Sheet
        self._create_summary_sheet(wb)
        
        # Per-API sheets
        for api_name, api_data in self.apis.items():
            self._create_api_sheet(wb, api_name, api_data)
        
        # All test cases sheet
        self._create_all_tests_sheet(wb)
        
        wb.save(excel_file)
        print(f"✓ Excel: {excel_file}")
        return excel_file

    def _create_summary_sheet(self, wb):
        """Create summary sheet"""
        ws = wb.create_sheet('Summary', 0)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        ws['A1'] = 'API Validation Test Cases - Summary'
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        
        ws['A5'] = 'Statistics'
        ws['A5'].font = Font(bold=True, size=12)
        
        for col in range(1, 3):
            ws.cell(row=6, column=col).fill = header_fill
            ws.cell(row=6, column=col).font = header_font
        
        ws['A6'] = 'Metric'
        ws['B6'] = 'Count'
        ws['A7'] = 'Total APIs'
        ws['B7'] = self.api_count
        ws['A8'] = 'Positive Tests'
        ws['B8'] = self.positive_count
        ws['A9'] = 'Negative Tests'
        ws['B9'] = self.negative_count
        ws['A10'] = 'Total Tests'
        ws['B10'] = len(self.test_cases)
        ws['B10'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_api_sheet(self, wb, api_name, api_data):
        """Create sheet for each API"""
        sheet_name = api_name[:31]
        ws = wb.create_sheet(sheet_name)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        ws['A1'] = f'API: {api_name}'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f'Method: {api_data["method"]} | Endpoint: {api_data["endpoint"]}'
        
        row = 4
        headers = ['TC ID', 'Title', 'Type', 'Category', 'Pre-Conditions', 'Steps', 'Input', 'Expected', 'Priority']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        api_id = re.sub(r'[^a-zA-Z0-9_]', '', api_name.replace(' ', '_'))[:10].upper()
        api_tests = [tc for tc in self.test_cases if tc['tc_id'].startswith(f'TC_{api_id}')]
        row = 5
        
        for tc in api_tests:
            ws.cell(row=row, column=1).value = tc['tc_id']
            ws.cell(row=row, column=2).value = tc['title']
            ws.cell(row=row, column=3).value = tc['type']
            ws.cell(row=row, column=4).value = tc['category']
            ws.cell(row=row, column=5).value = tc['pre_conditions']
            ws.cell(row=row, column=6).value = tc['steps']
            ws.cell(row=row, column=7).value = tc['input']
            ws.cell(row=row, column=8).value = tc['expected']
            ws.cell(row=row, column=9).value = tc['priority']
            
            ws.row_dimensions[row].height = 45
            for col in range(1, 10):
                ws.cell(row=row, column=col).alignment = wrap
            
            row += 1
        
        for i, width in enumerate([12, 25, 10, 15, 20, 30, 25, 30, 10], 1):
            ws.column_dimensions[chr(64 + i)].width = width

    def _create_all_tests_sheet(self, wb):
        """Create combined test cases sheet"""
        ws = wb.create_sheet('All Test Cases')
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        headers = ['TC ID', 'API', 'Title', 'Type', 'Category', 'Pre-Conditions', 'Steps', 'Input', 'Expected', 'Priority']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for tc in self.test_cases:
            # Extract API name by matching TC ID prefix
            api_name = 'Unknown'
            for name in self.apis.keys():
                api_id = re.sub(r'[^a-zA-Z0-9_]', '', name.replace(' ', '_'))[:10].upper()
                if tc['tc_id'].startswith(f'TC_{api_id}'):
                    api_name = name
                    break
            
            ws.cell(row=row, column=1).value = tc['tc_id']
            ws.cell(row=row, column=2).value = api_name
            ws.cell(row=row, column=3).value = tc['title']
            ws.cell(row=row, column=4).value = tc['type']
            ws.cell(row=row, column=5).value = tc['category']
            ws.cell(row=row, column=6).value = tc['pre_conditions']
            ws.cell(row=row, column=7).value = tc['steps']
            ws.cell(row=row, column=8).value = tc['input']
            ws.cell(row=row, column=9).value = tc['expected']
            ws.cell(row=row, column=10).value = tc['priority']
            
            ws.row_dimensions[row].height = 45
            for col in range(1, 11):
                ws.cell(row=row, column=col).alignment = wrap
            
            row += 1
        
        for i, width in enumerate([12, 15, 25, 10, 15, 20, 30, 25, 30, 10], 1):
            ws.column_dimensions[chr(64 + i)].width = width

def main():
    """Main execution"""
    ddd = r'c:\Users\aahmadkamar\Maybank\FSD Parser\input\api\PT_Scan_and_Pay_DDD_API_Spec_v2_working.xlsx'
    md_out = r'c:\Users\aahmadkamar\Maybank\FSD Parser\artifacts\Test_Case\Markdown'
    xl_out = r'c:\Users\aahmadkamar\Maybank\FSD Parser\artifacts\Test_Case\Excel'
    
    print('=' * 70)
    print('API VALIDATION TEST CASE GENERATOR')
    print('=' * 70 + '\n')
    
    gen = TestCaseGenerator(ddd)
    
    print('Step 1: Extracting APIs...')
    gen.extract_apis()
    
    print('Step 2: Generating test cases...')
    gen.generate_test_cases()
    
    print('Step 3: Creating markdown...')
    gen.generate_markdown(md_out)
    
    print('Step 4: Creating Excel...')
    gen.generate_excel(xl_out)
    
    print('\n' + '=' * 70)
    print('GENERATION COMPLETE!')
    print('=' * 70)

if __name__ == '__main__':
    main()
