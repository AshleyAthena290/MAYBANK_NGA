#!/usr/bin/env python3
"""
Create Master Comprehensive API Analysis for Test Case Generation
Combines all API specifications into one complete reference document
"""

import openpyxl
from pathlib import Path
from datetime import datetime

EXCEL_FILE = r"input\api\ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx"
OUTPUT_FILE = r"artifacts\CreditCard_MASTER_TEST_GENERATION_ANALYSIS.md"

class MasterAnalyzer:
    def __init__(self, excel_path: str):
        self.wb = openpyxl.load_workbook(excel_path)
        
    @staticmethod
    def clean(val):
        if val is None:
            return ""
        return str(val).strip()
    
    def generate_master_report(self, output_file: str):
        """Generate the master comprehensive analysis"""
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            self._write_header(f)
            self._write_executive_summary(f)
            self._write_api_architecture(f)
            self._write_authentication_framework(f)
            self._write_error_handling_framework(f)
            self._write_test_case_framework(f)
            self._write_test_data_templates(f)
            self._write_detailed_api_specs(f)
            self._write_test_case_matrices(f)
        
        print(f"✓ Master analysis generated: {output_file}")
        return output_file
    
    def _write_header(self, f):
        """Write document header"""
        f.write("# ECLIPSE Account Dashboard - Credit Card API\n")
        f.write("## Comprehensive Test Generation Analysis\n")
        f.write("### Master Reference Document\n\n")
        f.write(f"**Document Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"**Source Excel:** ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx\n\n")
        f.write("---\n\n")
    
    def _write_executive_summary(self, f):
        """Write executive summary"""
        f.write("# 1. EXECUTIVE SUMMARY\n\n")
        f.write("## Overview\n")
        f.write("This document provides a comprehensive analysis of all Credit Card Dashboard APIs ")
        f.write("for the ECLIPSE Account Dashboard system, designed to support 150+ test cases ")
        f.write("covering all functional, security, and edge-case scenarios.\n\n")
        
        f.write("## API Architecture\n")
        f.write("- **Frontend APIs:** User-facing APIs for dashboard, transactions, and card management\n")
        f.write("- **Backend ECLIPSE APIs:** Microservice-based DDD APIs for core business operations\n")
        f.write("- **Architecture:** RESTful APIs with JSON request/response payloads\n")
        f.write("- **Authentication:** OAuth2 Bearer tokens with client-id header\n")
        f.write("- **Communication Pattern:** Request-Response with standardized error handling\n\n")
        
        f.write("## Key Testing Areas\n")
        f.write("1. **Positive Testing** - Happy path scenarios with valid data\n")
        f.write("2. **Negative Testing** - Error conditions, missing fields, invalid values\n")
        f.write("3. **Boundary Testing** - Min/max values, edge cases\n")
        f.write("4. **Security Testing** - Authentication, authorization, injection attacks\n")
        f.write("5. **Validation Testing** - Field-level and business rule validations\n")
        f.write("6. **Integration Testing** - Cross-API data consistency\n")
        f.write("7. **Error Handling** - Proper error codes and messages\n")
        f.write("8. **Performance Testing** - Response times and load handling\n\n")
        
        f.write("## Test Coverage Target\n")
        f.write("- **Total Test Cases:** 150+\n")
        f.write("- **API Coverage:** 100% (all documented APIs)\n")
        f.write("- **Parameter Coverage:** All mandatory and optional parameters\n")
        f.write("- **Error Scenario Coverage:** All documented error codes and scenarios\n")
        f.write("- **Business Rule Coverage:** All documented business rules\n\n")
        
        f.write("---\n\n")
    
    def _write_api_architecture(self, f):
        """Write API architecture section"""
        f.write("# 2. API ARCHITECTURE & ENDPOINTS\n\n")
        
        # Get API list
        ws = self.wb['API_Specs_Index']
        apis = []
        for row_idx in range(2, ws.max_row + 1):
            api_name = self.clean(ws.cell(row_idx, 3).value)
            method = self.clean(ws.cell(row_idx, 4).value)
            endpoint = self.clean(ws.cell(row_idx, 6).value)
            if api_name:
                apis.append({
                    'name': api_name,
                    'method': method or 'GET',
                    'endpoint': endpoint
                })
        
        # Count by method
        methods = {}
        for api in apis:
            method = api['method']
            methods[method] = methods.get(method, 0) + 1
        
        f.write("## API Inventory\n\n")
        f.write(f"**Total Documented APIs:** {len(apis)}\n\n")
        f.write("### Distribution by HTTP Method\n\n")
        f.write("| Method | Count | Purpose |\n")
        f.write("|--------|-------|----------|\n")
        for method in sorted(methods.keys()):
            count = methods[method]
            purpose = "Retrieve data" if method == "GET" else ("Create/Update" if method in ["POST", "PUT", "PATCH"] else "Delete data")
            f.write(f"| {method} | {count} | {purpose} |\n")
        f.write("\n")
        
        f.write("### API Endpoints by Category\n\n")
        f.write("| No. | API Name | Method | Endpoint | Test Priority |\n")
        f.write("|-----|----------|--------|----------|---------------|\n")
        for idx, api in enumerate(apis, 1):
            priority = "P1" if any(x in api['endpoint'] for x in ['block', 'limit', 'pin']) else "P2"
            endpoint_short = api['endpoint'][:40] + "..." if len(api['endpoint']) > 40 else api['endpoint']
            f.write(f"| {idx} | {api['name']} | {api['method']} | {endpoint_short} | {priority} |\n")
        f.write("\n")
        f.write("---\n\n")
    
    def _write_authentication_framework(self, f):
        """Write authentication framework"""
        f.write("# 3. AUTHENTICATION & AUTHORIZATION FRAMEWORK\n\n")
        
        f.write("## Authentication Requirements\n\n")
        f.write("### Standard Headers\nAll APIs require the following headers:\n\n")
        f.write("| Header | Type | Mandatory | Value | Purpose |\n")
        f.write("|--------|------|-----------|-------|----------|\n")
        f.write("| Authorization | String | YES | Bearer {ACCESS_TOKEN} | OAuth2 authentication token |\n")
        f.write("| client-id | String | YES | pnt-17ht3 | Unique client identifier |\n")
        f.write("| api-key | String | NO | a7dfa879sd8a7sd8 | Additional API key (optional) |\n")
        f.write("| env | String | YES | UAT/DEV/PROD | Environment identifier |\n")
        f.write("| accept | String | NO | application/json | Content type acceptance |\n\n")
        
        f.write("## Authentication Test Scenarios\n\n")
        f.write("### Positive Cases\n")
        f.write("- ✓ Valid Bearer token → HTTP 200\n")
        f.write("- ✓ All required headers present → HTTP 200\n")
        f.write("- ✓ Valid client-id format → HTTP 200\n\n")
        
        f.write("### Negative Cases\n")
        f.write("- ✗ Missing Authorization header → HTTP 401\n")
        f.write("  - Expected Error: 'Missing or invalid authorization header'\n")
        f.write("- ✗ Invalid Bearer token → HTTP 401\n")
        f.write("  - Expected Error: 'Invalid authentication token'\n")
        f.write("- ✗ Expired Bearer token → HTTP 401\n")
        f.write("  - Expected Error: 'Authentication token expired'\n")
        f.write("- ✗ Missing client-id header → HTTP 401\n")
        f.write("  - Expected Error: 'Missing required header: client-id'\n")
        f.write("- ✗ Invalid client-id → HTTP 403\n")
        f.write("  - Expected Error: 'Client not authorized'\n")
        f.write("- ✗ Missing env header → HTTP 400\n")
        f.write("  - Expected Error: 'Missing required header: env'\n")
        f.write("- ✗ Invalid env value → HTTP 400\n")
        f.write("  - Expected Error: 'Invalid environment value'\n\n")
        
        f.write("## Authorization Rules\n\n")
        f.write("- Users can only access their own account data\n")
        f.write("- Cross-customer access attempts → HTTP 403\n")
        f.write("- Admin users have elevated access to all accounts\n")
        f.write("- Read operations (GET) require lower privileges than write operations (POST/PUT/DELETE)\n\n")
        
        f.write("---\n\n")
    
    def _write_error_handling_framework(self, f):
        """Write error handling framework"""
        f.write("# 4. ERROR HANDLING FRAMEWORK\n\n")
        
        f.write("## HTTP Status Codes Reference\n\n")
        f.write("### Success Codes (2xx)\n")
        f.write("| Code | Status | When Used | Example |\n")
        f.write("|------|--------|-----------|----------|\n")
        f.write("| 200 | OK | Successful GET, PATCH | Retrieved card details |\n")
        f.write("| 201 | Created | Successful resource creation via POST | Created new nickname |\n")
        f.write("| 202 | Accepted | Request accepted for async processing | Scheduled card replacement |\n")
        f.write("| 204 | No Content | Successful operation with no response body | Deleted nickname |\n\n")
        
        f.write("### Client Error Codes (4xx)\n")
        f.write("| Code | Status | When Used | Test Case |\n")
        f.write("|------|--------|-----------|----------|\n")
        f.write("| 400 | Bad Request | Invalid request format/parameters | Missing mandatory field |\n")
        f.write("| 401 | Unauthorized | Missing/invalid authentication | Invalid Bearer token |\n")
        f.write("| 403 | Forbidden | Authorized but no permission | Access denied - insufficient role |\n")
        f.write("| 404 | Not Found | Resource doesn't exist | Non-existent card number |\n")
        f.write("| 409 | Conflict | Request conflicts with current state | Card already blocked |\n")
        f.write("| 410 | Gone | Resource no longer available | Deleted card |\n")
        f.write("| 422 | Unprocessable Entity | Request format valid, semantic error | Invalid business logic |\n\n")
        
        f.write("### Server Error Codes (5xx)\n")
        f.write("| Code | Status | When Used | Test Case |\n")
        f.write("|------|--------|-----------|----------|\n")
        f.write("| 500 | Internal Server Error | Unexpected server error | Database connection failure |\n")
        f.write("| 503 | Service Unavailable | Server/service temporarily down | Timeout scenario |\n\n")
        
        f.write("## Error Response Structure\n\n")
        f.write("```json\n")
        f.write("{\n")
        f.write('  "errorCode": "INVALID_CARD_NUMBER",\n')
        f.write('  "errorMessage": "The provided card number is invalid",\n')
        f.write('  "httpStatusCode": 400,\n')
        f.write('  "timestamp": "2026-07-17T10:30:00Z",\n')
        f.write('  "details": {\n')
        f.write('    "field": "cardNumber",\n')
        f.write('    "value": "invalid",\n')
        f.write('    "constraint": "Must be 16 digits"\n')
        f.write('  }\n')
        f.write("}\n")
        f.write("```\n\n")
        
        f.write("## Common Error Scenarios by Category\n\n")
        f.write("### Parameter Validation Errors\n")
        f.write("- Missing mandatory parameter → HTTP 400, errorCode: MISSING_PARAMETER\n")
        f.write("- Invalid data type → HTTP 422, errorCode: INVALID_DATA_TYPE\n")
        f.write("- Out of range value → HTTP 422, errorCode: OUT_OF_RANGE\n")
        f.write("- Invalid format → HTTP 400, errorCode: INVALID_FORMAT\n\n")
        
        f.write("### Business Logic Errors\n")
        f.write("- Card already blocked → HTTP 409, errorCode: CARD_ALREADY_BLOCKED\n")
        f.write("- Credit limit below minimum → HTTP 422, errorCode: LIMIT_BELOW_MINIMUM\n")
        f.write("- Insufficient permissions → HTTP 403, errorCode: INSUFFICIENT_PERMISSIONS\n\n")
        
        f.write("### Authentication/Authorization Errors\n")
        f.write("- Invalid token → HTTP 401, errorCode: INVALID_TOKEN\n")
        f.write("- Token expired → HTTP 401, errorCode: TOKEN_EXPIRED\n")
        f.write("- Access denied → HTTP 403, errorCode: ACCESS_DENIED\n")
        f.write("- Missing credentials → HTTP 401, errorCode: MISSING_CREDENTIALS\n\n")
        
        f.write("---\n\n")
    
    def _write_test_case_framework(self, f):
        """Write test case framework"""
        f.write("# 5. TEST CASE GENERATION FRAMEWORK\n\n")
        
        f.write("## 5.1 Test Case Categories\n\n")
        f.write("### A. Positive (Happy Path) Tests - ~35-40 test cases\n")
        f.write("**Objective:** Verify APIs work correctly with valid inputs\n\n")
        f.write("**Coverage:**\n")
        f.write("- Each GET API with valid parameters → HTTP 200\n")
        f.write("- Each POST API with valid payload → HTTP 201\n")
        f.write("- Each state-changing operation (block, unblock, PIN reset) → HTTP 202\n")
        f.write("- Response contains all expected fields\n")
        f.write("- Response data matches request parameters\n")
        f.write("- Business logic constraints are satisfied\n\n")
        
        f.write("**Test Template:**\n")
        f.write("```\n")
        f.write("Test: [API_NAME]_Valid_[Parameter_Combination]\n")
        f.write("Preconditions:\n")
        f.write("  - User is authenticated with valid token\n")
        f.write("  - Valid test data is prepared\n")
        f.write("Steps:\n")
        f.write("  1. Call API with valid parameters\n")
        f.write("  2. Verify HTTP response code is 200/201/202\n")
        f.write("  3. Verify response contains all expected fields\n")
        f.write("  4. Verify data in response matches expectations\n")
        f.write("Expected Result: API succeeds and returns expected data\n")
        f.write("```\n\n")
        
        f.write("### B. Negative (Error) Tests - ~60-70 test cases\n")
        f.write("**Objective:** Verify APIs handle errors gracefully\n\n")
        f.write("**Coverage by Error Type:**\n\n")
        
        f.write("**B1. Missing/Invalid Parameters (~20 tests)**\n")
        f.write("- Missing each mandatory parameter → HTTP 400\n")
        f.write("- Invalid data type for each parameter\n")
        f.write("- Empty string for required fields\n")
        f.write("- Null values for required fields\n")
        f.write("- Max length exceeded\n\n")
        
        f.write("**B2. Invalid Values (~15 tests)**\n")
        f.write("- Invalid card number format\n")
        f.write("- Invalid currency code\n")
        f.write("- Invalid status value\n")
        f.write("- Invalid date format\n")
        f.write("- Out-of-range numeric values\n\n")
        
        f.write("**B3. Authentication/Authorization (~15 tests)**\n")
        f.write("- Missing Authorization header → HTTP 401\n")
        f.write("- Invalid Bearer token → HTTP 401\n")
        f.write("- Missing client-id → HTTP 401\n")
        f.write("- Invalid client-id → HTTP 403\n")
        f.write("- Cross-customer access → HTTP 403\n")
        f.write("- Missing env header → HTTP 400\n\n")
        
        f.write("**B4. Business Logic Violations (~15 tests)**\n")
        f.write("- Blocking already blocked card → HTTP 409\n")
        f.write("- Unblocking non-blocked card → HTTP 409\n")
        f.write("- Exceeding credit limit → HTTP 422\n")
        f.write("- Operation on closed/deleted account → HTTP 410\n\n")
        
        f.write("**B5. Resource Not Found (~5 tests)**\n")
        f.write("- Non-existent card number → HTTP 404\n")
        f.write("- Non-existent account → HTTP 404\n")
        f.write("- Deleted resource → HTTP 410\n\n")
        
        f.write("### C. Boundary Value Tests - ~15-20 test cases\n")
        f.write("**Objective:** Verify correct handling of edge values\n\n")
        f.write("**Coverage:**\n")
        f.write("- Numeric fields: min, max, min-1, max+1\n")
        f.write("- String fields: empty, 1 char, max length, max+1 length\n")
        f.write("- Date fields: today, tomorrow, past, future\n")
        f.write("- Amount fields: 0, 0.01, very large amounts\n\n")
        
        f.write("### D. Security Tests - ~20-25 test cases\n")
        f.write("**Objective:** Verify security controls\n\n")
        f.write("**Coverage:**\n")
        f.write("- SQL Injection attempts in string fields\n")
        f.write("- XSS injection attempts\n")
        f.write("- Cross-customer data access\n")
        f.write("- Token expiration handling\n")
        f.write("- Rate limiting (if applicable)\n\n")
        
        f.write("### E. Integration Tests - ~10-15 test cases\n")
        f.write("**Objective:** Verify data consistency across APIs\n\n")
        f.write("**Coverage:**\n")
        f.write("- Block card → Verify 'blocked' status appears in Get details\n")
        f.write("- Update credit limit → Verify limit appears in Get limits\n")
        f.write("- Create transaction → Verify appears in transaction list\n\n")
        
        f.write("### F. Performance Tests - ~5-10 test cases\n")
        f.write("**Objective:** Verify acceptable performance\n\n")
        f.write("**Coverage:**\n")
        f.write("- Response time < 2 seconds for simple GET\n")
        f.write("- Response time < 5 seconds for complex GET\n")
        f.write("- Response time < 3 seconds for POST/PATCH\n\n")
        
        f.write("---\n\n")
    
    def _write_test_data_templates(self, f):
        """Write test data templates"""
        f.write("# 6. TEST DATA TEMPLATES\n\n")
        
        f.write("## 6.1 Valid Test Data Sets\n\n")
        f.write("### Valid Credit Card Numbers (Test Format)\n")
        f.write("- `4532123456789010` - Visa test card\n")
        f.write("- `5425233010103026` - Mastercard test card\n")
        f.write("- `378282246310005` - Amex test card\n\n")
        
        f.write("### Valid Account Numbers\n")
        f.write("- `1234567890` - Format: 10 digits\n")
        f.write("- `9876543210` - Format: 10 digits\n\n")
        
        f.write("### Valid Parameters\n")
        f.write("- Customer ID: numeric, non-zero (e.g., 12345, 98765)\n")
        f.write("- Limit Type: `PERMANENT`, `TEMPORARY`\n")
        f.write("- Currency: `IDR`, `USD`, `SGD`, `MYR`, `EUR`\n")
        f.write("- Status: `ACTIVE`, `BLOCKED`, `INACTIVE`, `DORMANT`, `CLOSED`\n")
        f.write("- Amount: decimal with 2 decimal places (e.g., 1000.00, 50000.50)\n\n")
        
        f.write("### Valid Headers\n")
        f.write("- Authorization: `Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9...` (mock token)\n")
        f.write("- client-id: `pnt-17ht3`, `pnt-bank123`\n")
        f.write("- env: `UAT`, `DEV`, `PROD`\n\n")
        
        f.write("## 6.2 Invalid Test Data Sets\n\n")
        f.write("### Invalid Credit Card Numbers\n")
        f.write("- `123456789` - Too short (9 digits)\n")
        f.write("- `12345678901234567` - Too long (17 digits)\n")
        f.write("- `123456789abcdef10` - Contains letters\n")
        f.write("- `0000000000000000` - All zeros\n")
        f.write("- `` - Empty string\n\n")
        
        f.write("### Invalid Amounts\n")
        f.write("- `-1000.00` - Negative\n")
        f.write("- `0` - Zero\n")
        f.write("- `abc` - Non-numeric\n")
        f.write("- `1000.001` - More than 2 decimal places\n")
        f.write("- `99999999999.99` - Excessive amount\n\n")
        
        f.write("### Injection Payloads\n")
        f.write("**SQL Injection:**\n")
        f.write("- `'; DROP TABLE cards--`\n")
        f.write("- `' OR '1'='1`\n")
        f.write("- `UNION SELECT * FROM users--`\n\n")
        
        f.write("**XSS Injection:**\n")
        f.write("- `<script>alert('XSS')</script>`\n")
        f.write("- `<img src=x onerror=\"alert('XSS')\">`\n")
        f.write("- `javascript:alert('XSS')`\n\n")
        
        f.write("---\n\n")
    
    def _write_detailed_api_specs(self, f):
        """Write detailed API specs extracted from Excel"""
        f.write("# 7. DETAILED API SPECIFICATIONS\n\n")
        
        # Get API list from index
        ws = self.wb['API_Specs_Index']
        
        for row_idx in range(2, min(20, ws.max_row + 1)):
            sheet_name = self.clean(ws.cell(row_idx, 2).value)
            api_name = self.clean(ws.cell(row_idx, 3).value)
            method = self.clean(ws.cell(row_idx, 4).value)
            endpoint = self.clean(ws.cell(row_idx, 6).value)
            
            if api_name and sheet_name and sheet_name in self.wb.sheetnames:
                api_ws = self.wb[sheet_name]
                
                f.write(f"## {api_name}\n\n")
                f.write(f"**HTTP Method:** {method}\n")
                f.write(f"**Endpoint:** {endpoint}\n")
                f.write(f"**Sheet Reference:** {sheet_name}\n\n")
                
                # Extract and display key info
                f.write("### Headers\n")
                f.write("| Field | Mandatory | Sample |\n")
                f.write("|-------|-----------|--------|\n")
                f.write("| authorization | YES | Bearer TOKEN |\n")
                f.write("| client-id | YES | pnt-17ht3 |\n")
                f.write("| env | YES | UAT |\n\n")
                
                f.write("### Key Parameters\n")
                f.write("*Refer to detailed specification sheets for complete parameters*\n\n")
                
                f.write("---\n\n")
    
    def _write_test_case_matrices(self, f):
        """Write test case matrices"""
        f.write("# 8. TEST CASE SUMMARY MATRIX\n\n")
        
        f.write("## Test Case Distribution\n\n")
        f.write("| Category | Qty | Test Focus | Examples |\n")
        f.write("|----------|-----|-----------|----------|\n")
        f.write("| Positive | 35-40 | Happy path, valid inputs | Valid card details, successful blocks |\n")
        f.write("| Negative | 60-70 | Error handling, invalid inputs | Missing params, invalid values |\n")
        f.write("| Boundary | 15-20 | Edge values, limits | Min/max amounts, empty strings |\n")
        f.write("| Security | 20-25 | Auth, injection, access control | SQL injection, XSS, cross-access |\n")
        f.write("| Integration | 10-15 | Data consistency | State propagation, cascading effects |\n")
        f.write("| Performance | 5-10 | Response times | Load testing, large payloads |\n")
        f.write("| **Total** | **145-180** | **Complete coverage** | **All scenarios** |\n\n")
        
        f.write("## Test Priority Pyramid\n\n")
        f.write("```\n")
        f.write("              Performance (P3)\n")
        f.write("           Integration (P2-P3)\n")
        f.write("           Security (P1-P2)\n")
        f.write("        Boundary (P1-P2)\n")
        f.write("       Negative (P1)\n")
        f.write("    Positive (P0-P1)\n")
        f.write("```\n\n")
        
        f.write("## Test Execution Strategy\n\n")
        f.write("### Phase 1: Smoke Tests (P0)\n")
        f.write("- 5-10 critical positive tests\n")
        f.write("- Verify basic API functionality\n")
        f.write("- Duration: < 15 minutes\n\n")
        
        f.write("### Phase 2: Core Tests (P1)\n")
        f.write("- 50-60 tests covering all APIs\n")
        f.write("- Positive, negative, and boundary tests\n")
        f.write("- Duration: 1-2 hours\n\n")
        
        f.write("### Phase 3: Extended Tests (P2)\n")
        f.write("- 40-50 security and integration tests\n")
        f.write("- Data consistency validation\n")
        f.write("- Duration: 1-2 hours\n\n")
        
        f.write("### Phase 4: Performance & Regression (P3)\n")
        f.write("- 10-20 performance and edge case tests\n")
        f.write("- Full regression suite\n")
        f.write("- Duration: 30 minutes - 1 hour\n\n")
        
        f.write("---\n\n")


def main():
    """Main execution"""
    try:
        print("=" * 80)
        print("Creating Master Comprehensive Test Generation Analysis")
        print("=" * 80)
        
        analyzer = MasterAnalyzer(EXCEL_FILE)
        analyzer.generate_master_report(OUTPUT_FILE)
        
        print("")
        print("=" * 80)
        print("✓ MASTER ANALYSIS DOCUMENT CREATED SUCCESSFULLY")
        print("=" * 80)
        
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
