import openpyxl
import json
from typing import Dict, List, Any, Optional
from pathlib import Path

file_path = r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx"

def extract_section(ws, start_row: int, section_label: str) -> Dict[str, Any]:
    """Extract a section from the worksheet"""
    result = {}
    current_row = start_row
    
    while current_row <= ws.max_row:
        cell_a = ws[f'A{current_row}'].value
        cell_b = ws[f'B{current_row}'].value
        
        # Check if this is the section header
        if cell_a and str(cell_a).strip() == section_label:
            current_row += 1
            # Now extract the table
            while current_row <= ws.max_row:
                cell_a = ws[f'A{current_row}'].value
                cell_b = ws[f'B{current_row}'].value
                
                # Check for end of section (empty row or new section)
                if not cell_a or not cell_b:
                    if current_row > start_row + 1:
                        break
                
                # This is a data row
                if cell_a and cell_b and str(cell_a).strip() != "-":
                    result[str(cell_b).strip()] = {
                        'Name': str(cell_b).strip(),
                        'Type': ws[f'C{current_row}'].value or '-',
                        'Length': ws[f'D{current_row}'].value or '-',
                        'Mandatory': ws[f'E{current_row}'].value or '-',
                        'Sample Value': ws[f'F{current_row}'].value or '-',
                        'Description': ws[f'G{current_row}'].value or '-'
                    }
                
                current_row += 1
            break
        
        current_row += 1
    
    return result

def extract_api_endpoint(ws) -> Dict[str, Any]:
    """Extract complete API endpoint specification from a sheet"""
    api_spec = {}
    
    # Extract Service/API Name
    service_cell = ws['B1'].value
    api_spec['Service'] = str(service_cell).strip() if service_cell else 'Unknown'
    
    # Extract Method
    method_cell = ws['B2'].value
    api_spec['Method'] = str(method_cell).strip() if method_cell else 'Unknown'
    
    # Extract URL
    url_cell = ws['B3'].value
    api_spec['URL'] = str(url_cell).strip() if url_cell else 'Unknown'
    
    # Extract Message Type
    msg_type_cell = ws['B4'].value
    api_spec['MessageType'] = str(msg_type_cell).strip() if msg_type_cell else 'JSON'
    
    # Extract HTTP Headers
    api_spec['HTTPHeaders'] = []
    for row in range(7, ws.max_row + 1):
        cell_a = ws[f'A{row}'].value
        cell_b = ws[f'B{row}'].value
        
        # Check for HTTP Header section
        if cell_a and str(cell_a).strip() == "-" and cell_b and "HTTP Header" in str(cell_b):
            # Found headers section, extract headers
            row += 1
            # Skip header row (Name, Type, Length, etc.)
            row += 1
            
            while row <= ws.max_row:
                name_cell = ws[f'B{row}'].value
                type_cell = ws[f'C{row}'].value
                length_cell = ws[f'D{row}'].value
                mandatory_cell = ws[f'E{row}'].value
                sample_cell = ws[f'F{row}'].value
                desc_cell = ws[f'G{row}'].value
                
                # Check for empty row or end of section
                if not name_cell or str(name_cell).strip() == "-":
                    break
                
                api_spec['HTTPHeaders'].append({
                    'Name': str(name_cell).strip(),
                    'Type': str(type_cell).strip() if type_cell else '-',
                    'Length': str(length_cell).strip() if length_cell else '-',
                    'Mandatory': str(mandatory_cell).strip() if mandatory_cell else '-',
                    'Sample Value': str(sample_cell).strip() if sample_cell else '-',
                    'Description': str(desc_cell).strip() if desc_cell else '-'
                })
                row += 1
            break
    
    # Extract HTTP Body
    api_spec['HTTPBody'] = []
    for row in range(7, ws.max_row + 1):
        cell_a = ws[f'A{row}'].value
        cell_b = ws[f'B{row}'].value
        
        if cell_a and str(cell_a).strip() == "-" and cell_b and "HTTP Body" in str(cell_b):
            # Found body section
            row += 1
            # Skip header row
            row += 1
            
            while row <= ws.max_row:
                name_cell = ws[f'B{row}'].value
                type_cell = ws[f'C{row}'].value
                
                # Check for end of section
                if not name_cell or str(name_cell).strip() == "-":
                    break
                
                # Check if we've reached another section
                if str(name_cell).strip().startswith(">") or "Request Parameter" in str(name_cell):
                    break
                
                api_spec['HTTPBody'].append({
                    'Name': str(name_cell).strip(),
                    'Type': str(type_cell).strip() if type_cell else '-',
                    'Length': ws[f'D{row}'].value or '-',
                    'Mandatory': ws[f'E{row}'].value or '-',
                    'Sample Value': ws[f'F{row}'].value or '-',
                    'Description': ws[f'G{row}'].value or '-'
                })
                row += 1
            break
    
    # Extract Request Parameters
    api_spec['RequestParameters'] = []
    for row in range(7, ws.max_row + 1):
        cell_a = ws[f'A{row}'].value
        cell_b = ws[f'B{row}'].value
        
        if cell_a and str(cell_a).strip() == "-" and cell_b and "Request Parameter" in str(cell_b):
            row += 1
            # Skip header row
            row += 1
            
            while row <= ws.max_row:
                name_cell = ws[f'B{row}'].value
                
                if not name_cell or str(name_cell).strip() == "-":
                    break
                
                if "Response" in str(name_cell) or "Error" in str(name_cell):
                    break
                
                api_spec['RequestParameters'].append({
                    'Name': str(name_cell).strip(),
                    'Type': ws[f'C{row}'].value or '-',
                    'Length': ws[f'D{row}'].value or '-',
                    'Mandatory': ws[f'E{row}'].value or '-',
                    'Sample Value': ws[f'F{row}'].value or '-',
                    'Description': ws[f'G{row}'].value or '-'
                })
                row += 1
            break
    
    # Extract Response
    api_spec['Response'] = []
    for row in range(7, ws.max_row + 1):
        cell_a = ws[f'A{row}'].value
        cell_b = ws[f'B{row}'].value
        
        if cell_a and str(cell_a).strip() == "-" and cell_b and "Response" in str(cell_b) and "Error" not in str(cell_b):
            row += 1
            # Skip header row
            row += 1
            
            while row <= ws.max_row:
                name_cell = ws[f'B{row}'].value
                
                if not name_cell or str(name_cell).strip() == "-":
                    break
                
                if "Error" in str(name_cell):
                    break
                
                api_spec['Response'].append({
                    'Name': str(name_cell).strip(),
                    'Type': ws[f'C{row}'].value or '-',
                    'Length': ws[f'D{row}'].value or '-',
                    'Mandatory': ws[f'E{row}'].value or '-',
                    'Sample Value': ws[f'F{row}'].value or '-',
                    'Description': ws[f'G{row}'].value or '-'
                })
                row += 1
            break
    
    # Extract Error Responses
    api_spec['ErrorResponses'] = []
    for row in range(7, ws.max_row + 1):
        cell_a = ws[f'A{row}'].value
        cell_b = ws[f'B{row}'].value
        
        if cell_a and str(cell_a).strip() == "-" and cell_b and "Error" in str(cell_b) and "Response" in str(cell_b):
            row += 1
            # Skip header row
            row += 1
            
            while row <= ws.max_row:
                error_code = ws[f'B{row}'].value
                
                if not error_code or str(error_code).strip() == "-":
                    break
                
                api_spec['ErrorResponses'].append({
                    'Error Code': str(error_code).strip(),
                    'Error Message': ws[f'C{row}'].value or '-',
                    'Description': ws[f'D{row}'].value or '-'
                })
                row += 1
            break
    
    return api_spec

# Load workbook
wb = openpyxl.load_workbook(file_path)
sheet_names = wb.sheetnames

# Skip metadata sheets
skip_sheets = ['README', 'API_Specs_Index', 'Sheet1']
api_sheets = [s for s in sheet_names if s not in skip_sheets]

print("\n" + "=" * 100)
print("COMPREHENSIVE API SPECIFICATIONS EXTRACTION")
print("=" * 100)
print(f"\nFile: {file_path}")
print(f"Total API Endpoints Found: {len(api_sheets)}\n")

# Extract all API endpoints
all_endpoints = {}

for sheet_name in api_sheets:
    ws = wb[sheet_name]
    
    try:
        endpoint = extract_api_endpoint(ws)
        all_endpoints[sheet_name] = endpoint
        
        print(f"\n{'─' * 100}")
        print(f"API ENDPOINT: {sheet_name}")
        print(f"{'─' * 100}")
        print(f"Service Name: {endpoint['Service']}")
        print(f"HTTP Method:  {endpoint['Method']}")
        print(f"URL:          {endpoint['URL']}")
        print(f"Message Type: {endpoint['MessageType']}")
        
        # Print headers
        if endpoint['HTTPHeaders']:
            print(f"\n[HTTP Headers ({len(endpoint['HTTPHeaders'])} total)]")
            for header in endpoint['HTTPHeaders']:
                mandatory = "✓" if header['Mandatory'] == 'Y' else "✗"
                print(f"  • {header['Name']:<20} | Type: {header['Type']:<10} | Mandatory: {mandatory} | Sample: {str(header['Sample Value'])[:30]}")
        
        # Print body parameters
        if endpoint['HTTPBody']:
            print(f"\n[HTTP Body ({len(endpoint['HTTPBody'])} total)]")
            for param in endpoint['HTTPBody']:
                mandatory = "✓" if param['Mandatory'] == 'Y' else "✗"
                print(f"  • {param['Name']:<20} | Type: {param['Type']:<10} | Mandatory: {mandatory}")
        
        # Print request parameters
        if endpoint['RequestParameters']:
            print(f"\n[Request Parameters ({len(endpoint['RequestParameters'])} total)]")
            for param in endpoint['RequestParameters']:
                mandatory = "✓" if param['Mandatory'] == 'Y' else "✗"
                print(f"  • {param['Name']:<20} | Type: {param['Type']:<10} | Mandatory: {mandatory}")
        
        # Print response fields
        if endpoint['Response']:
            print(f"\n[Response Fields ({len(endpoint['Response'])} total)]")
            for field in endpoint['Response'][:5]:  # Show first 5
                print(f"  • {field['Name']:<20} | Type: {field['Type']:<10}")
            if len(endpoint['Response']) > 5:
                print(f"  ... and {len(endpoint['Response']) - 5} more fields")
        
        # Print error responses
        if endpoint['ErrorResponses']:
            print(f"\n[Error Responses ({len(endpoint['ErrorResponses'])} total)]")
            for error in endpoint['ErrorResponses']:
                print(f"  • {error['Error Code']:<10} | {error['Error Message']}")
        
    except Exception as e:
        print(f"\nERROR processing {sheet_name}: {e}")

# Save summary to JSON
output_file = r"artifacts\api_endpoints_summary.json"
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(all_endpoints, f, indent=2, ensure_ascii=False)

print(f"\n\n{'=' * 100}")
print(f"Summary saved to: {output_file}")
print(f"Total endpoints extracted: {len(all_endpoints)}")
print(f"{'=' * 100}")

