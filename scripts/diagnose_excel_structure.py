#!/usr/bin/env python3
"""
Quick diagnostic to check Excel file structure
"""

import openpyxl

EXCEL_FILE = r"input\api\ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx"

wb = openpyxl.load_workbook(EXCEL_FILE)

print("=" * 80)
print("EXCEL FILE STRUCTURE ANALYSIS")
print("=" * 80)
print(f"\nFile: {EXCEL_FILE}")
print(f"\nTotal Sheets: {len(wb.sheetnames)}")
print(f"\nSheet Names:")
for i, sheet in enumerate(wb.sheetnames[:20], 1):
    ws = wb[sheet]
    print(f"  {i:2d}. {sheet:<50} (Rows: {ws.max_row}, Cols: {ws.max_column})")

# Examine first sheet in detail
print("\n" + "=" * 80)
print("FIRST SHEET CONTENT PREVIEW")
print("=" * 80)

ws = wb[wb.sheetnames[0]]
print(f"\nSheet: {wb.sheetnames[0]}")
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns\n")

print("First 20 rows:")
for row_idx in range(1, min(21, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(8, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value if cell.value else ""
        row_data.append(str(value)[:30])
    print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

# Check for sheets that look like API spec sheets
print("\n" + "=" * 80)
print("POTENTIAL API SPECIFICATION SHEETS")
print("=" * 80)

api_keywords = ['get', 'set', 'post', 'put', 'patch', 'delete', 'list', 'create', 'update', 'remove', 'block']
for sheet in wb.sheetnames:
    if any(keyword in sheet.lower() for keyword in api_keywords):
        print(f"✓ {sheet}")

print("\n" + "=" * 80)
