import openpyxl
import json
from pathlib import Path
from typing import Dict, List, Any

# File path
file_path = r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx"

try:
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Get all sheet names
    sheet_names = wb.sheetnames
    print("=" * 80)
    print("EXCEL FILE STRUCTURE ANALYSIS")
    print("=" * 80)
    print(f"\nFile: {file_path}")
    print(f"Total Sheets: {len(sheet_names)}\n")
    
    print("Sheet Names:")
    for i, sheet in enumerate(sheet_names, 1):
        print(f"  {i}. {sheet}")
    
    print("\n" + "=" * 80)
    print("SHEET CONTENT PREVIEW")
    print("=" * 80)
    
    # Preview content of each sheet
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        print(f"\n\n[SHEET: {sheet_name}]")
        print(f"Dimensions: {ws.dimensions}")
        
        # Get max row and column
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"Max Row: {max_row}, Max Column: {max_col}\n")
        
        # Read first 50 rows to understand structure
        print("Content Preview (first 20 rows):")
        print("-" * 160)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, max_row), values_only=True), 1):
            # Print row with limited width
            row_str = " | ".join(str(cell)[:25].ljust(25) if cell is not None else "-".ljust(25) for cell in row[:8])
            print(f"Row {row_idx}: {row_str}")
        
        if max_row > 20:
            print(f"... ({max_row - 20} more rows)")
            
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
