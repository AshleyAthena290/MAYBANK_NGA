#!/usr/bin/env python3
"""
Generate test data Excel file for ECLIPSE Account Dashboard Casa API Test Cases
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import json

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Data"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define columns
headers = [
    "Test Data ID",
    "Related Test Case ID",
    "Description",
    "Input Parameters",
    "Expected Result",
    "Positive/Negative",
    "Remarks"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 30

# Test data entries
test_data = [
    {
        "id": "TD_001",
        "tc_id": "CASA_DASHBOARD_TC_001",
        "desc": "Valid Deposit Account Group Request",
        "params": json.dumps({"accountGroup": "DEPOSIT", "client-id": "pnt-17ht3", "env": "UAT"}),
        "result": "Returns 200 with array of DEPOSIT accounts",
        "type": "Positive",
        "remarks": "Happy path - all deposit accounts returned"
    },
    {
        "id": "TD_002",
        "tc_id": "CASA_DASHBOARD_TC_002",
        "desc": "Valid Investment Account Group Request",
        "params": json.dumps({"accountGroup": "INVESTMENT", "client-id": "pnt-17ht3", "env": "UAT"}),
        "result": "Returns 200 with array of INVESTMENT accounts",
        "type": "Positive",
        "remarks": "MF, BOND, GOLD account types returned"
    },
    {
        "id": "TD_003",
        "tc_id": "CASA_DASHBOARD_TC_003",
        "desc": "Multiple Account Groups Query",
        "params": json.dumps({"accountGroup": ["DEPOSIT", "INVESTMENT"], "client-id": "pnt-17ht3", "env": "UAT"}),
        "result": "Returns 200 with both DEPOSIT and INVESTMENT accounts",
        "type": "Positive",
        "remarks": "Multi-group filtering supported"
    },
    {
        "id": "TD_004",
        "tc_id": "CASA_DASHBOARD_TC_004",
        "desc": "Credit Card Account Type Query",
        "params": json.dumps({"accountType": "CC", "client-id": "pnt-17ht3", "env": "UAT"}),
        "result": "Returns 200 with Credit Card accounts only",
        "type": "Positive",
        "remarks": "Credit card specific filtering"
    },
    {
        "id": "TD_005",
        "tc_id": "CASA_DASHBOARD_TC_005",
        "desc": "Missing Authentication Token",
        "params": json.dumps({"accountGroup": "DEPOSIT", "client-id": "pnt-17ht3", "env": "UAT"}),
        "result": "Returns 401 Unauthorized - no account data",
        "type": "Negative",
        "remarks": "Authorization header missing - security breach"
    },
    {
        "id": "TD_006",
        "tc_id": "CASA_DASHBOARD_TC_006",
        "desc": "Invalid/Expired Authentication Token",
        "params": json.dumps({"auth": "Bearer INVALID_TOKEN_12345", "accountGroup": "DEPOSIT"}),
        "result": "Returns 401 - Invalid or expired token",
        "type": "Negative",
        "remarks": "Malformed token validation"
    },
    {
        "id": "TD_007",
        "tc_id": "CASA_DASHBOARD_TC_007",
        "desc": "Expired Token Test",
        "params": json.dumps({"auth": "Bearer <EXPIRED_TOKEN>", "accountGroup": "DEPOSIT"}),
        "result": "Returns 401 - Token expired",
        "type": "Negative",
        "remarks": "Token expiry validation"
    },
    {
        "id": "TD_008",
        "tc_id": "CASA_DASHBOARD_TC_008",
        "desc": "Missing Mandatory Client-ID Header",
        "params": json.dumps({"auth": "Bearer <VALID>", "env": "UAT", "accountGroup": "DEPOSIT"}),
        "result": "Returns 400 - Missing mandatory header: client-id",
        "type": "Negative",
        "remarks": "Client-ID validation"
    },
    {
        "id": "TD_009",
        "tc_id": "CASA_DASHBOARD_TC_009",
        "desc": "Invalid Client-ID Value",
        "params": json.dumps({"client-id": "invalid-client-12345", "env": "UAT"}),
        "result": "Returns 403 - Client not authorized",
        "type": "Negative",
        "remarks": "Client authorization check"
    },
    {
        "id": "TD_010",
        "tc_id": "CASA_DASHBOARD_TC_010",
        "desc": "Missing Environment Header",
        "params": json.dumps({"client-id": "pnt-17ht3", "api-key": "a7dfa879sd8a7sd8"}),
        "result": "Returns 400 - Missing mandatory header: env",
        "type": "Negative",
        "remarks": "Environment header validation"
    },
    {
        "id": "TD_011",
        "tc_id": "CASA_DASHBOARD_TC_011",
        "desc": "Invalid Environment Value",
        "params": json.dumps({"env": "INVALID", "client-id": "pnt-17ht3"}),
        "result": "Returns 400 - Invalid environment value",
        "type": "Negative",
        "remarks": "Environment value must be UAT or PROD"
    },
    {
        "id": "TD_012",
        "tc_id": "CASA_DASHBOARD_TC_012",
        "desc": "Invalid Account Group Parameter",
        "params": json.dumps({"accountGroup": "INVALID_GROUP"}),
        "result": "Returns 422 - Invalid accountGroup",
        "type": "Negative",
        "remarks": "Unsupported account group value"
    },
    {
        "id": "TD_013",
        "tc_id": "CASA_DASHBOARD_TC_013",
        "desc": "No Accounts Found - Empty Result",
        "params": json.dumps({"accountGroup": "INVESTMENT"}),
        "result": "Returns 200 with empty accounts array",
        "type": "Negative",
        "remarks": "User has no investment accounts"
    },
    {
        "id": "TD_014",
        "tc_id": "CASA_DASHBOARD_TC_014",
        "desc": "Database Connection Timeout",
        "params": json.dumps({"accountGroup": "DEPOSIT"}),
        "result": "Returns 503 - Service Unavailable",
        "type": "Negative",
        "remarks": "Backend timeout scenario"
    },
    {
        "id": "TD_015",
        "tc_id": "CASA_DASHBOARD_TC_015",
        "desc": "Empty Query Parameter Value",
        "params": json.dumps({"accountGroup": ""}),
        "result": "Returns 400 - Invalid or missing accountGroup",
        "type": "Negative",
        "remarks": "Malformed query parameter"
    },
    {
        "id": "TD_016",
        "tc_id": "CASA_DASHBOARD_TC_016",
        "desc": "SQL Injection Attack Attempt",
        "params": json.dumps({"accountGroup": "DEPOSIT'; DROP TABLE accounts; --"}),
        "result": "Returns 400 - Invalid parameter format",
        "type": "Negative",
        "remarks": "Security validation - injection prevention"
    },
    {
        "id": "TD_017",
        "tc_id": "CASA_DASHBOARD_TC_017",
        "desc": "XSS Attack Attempt",
        "params": json.dumps({"accountGroup": "<script>alert('XSS')</script>"}),
        "result": "Returns 400 - Invalid parameter format",
        "type": "Negative",
        "remarks": "Security - XSS prevention"
    },
    {
        "id": "TD_018",
        "tc_id": "CASA_DASHBOARD_TC_018",
        "desc": "Case Sensitivity Check",
        "params": json.dumps({"accountGroup": "deposit"}),
        "result": "Returns 422 or 200 depending on implementation",
        "type": "Positive",
        "remarks": "Verify case-sensitive parameter handling"
    },
    {
        "id": "TD_019",
        "tc_id": "CASA_DASHBOARD_TC_019",
        "desc": "Duplicate Query Parameters",
        "params": json.dumps({"accountGroup": ["DEPOSIT", "INVESTMENT", "CREDIT_CARD"]}),
        "result": "Returns 200 with accounts from all groups",
        "type": "Positive",
        "remarks": "Multiple filters applied correctly"
    },
    {
        "id": "TD_020",
        "tc_id": "CASA_DASHBOARD_TC_020",
        "desc": "Very Long Query String (>2000 chars)",
        "params": json.dumps({"accountGroup": "DEPOSIT", "filter": "x" * 2500}),
        "result": "Returns 414 URI Too Long or 400",
        "type": "Negative",
        "remarks": "URI length validation"
    },
    {
        "id": "TD_021",
        "tc_id": "CASA_DETAIL_TC_001",
        "desc": "Valid Checking Account Detail Request",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526"}),
        "result": "Returns 200 with complete account details",
        "type": "Positive",
        "remarks": "CA account details retrieved successfully"
    },
    {
        "id": "TD_022",
        "tc_id": "CASA_DETAIL_TC_002",
        "desc": "Valid Savings Account Detail Request",
        "params": json.dumps({"accountType": "SA", "accountId": "123456789012"}),
        "result": "Returns 200 with savings details including interest rate",
        "type": "Positive",
        "remarks": "SA specific fields included"
    },
    {
        "id": "TD_023",
        "tc_id": "CASA_DETAIL_TC_003",
        "desc": "Invalid Account ID Format",
        "params": json.dumps({"accountType": "CA", "accountId": "INVALID123"}),
        "result": "Returns 400 - Invalid account ID format",
        "type": "Negative",
        "remarks": "Format validation"
    },
    {
        "id": "TD_024",
        "tc_id": "CASA_DETAIL_TC_004",
        "desc": "Non-Existent Account ID",
        "params": json.dumps({"accountType": "CA", "accountId": "999999999999"}),
        "result": "Returns 404 - Account not found",
        "type": "Negative",
        "remarks": "Not found scenario"
    },
    {
        "id": "TD_025",
        "tc_id": "CASA_DETAIL_TC_005",
        "desc": "Unauthorized Account Access",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "customerRole": "OTHER"}),
        "result": "Returns 403 - Forbidden access",
        "type": "Negative",
        "remarks": "Cross-customer access prevention"
    },
    {
        "id": "TD_026",
        "tc_id": "CASA_DETAIL_TC_006",
        "desc": "Invalid Account Type",
        "params": json.dumps({"accountType": "INVALID", "accountId": "123456789012"}),
        "result": "Returns 422 - Invalid account type",
        "type": "Negative",
        "remarks": "Account type validation"
    },
    {
        "id": "TD_027",
        "tc_id": "CASA_DETAIL_TC_007",
        "desc": "Closed Account Access",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "status": "CLOSED"}),
        "result": "Returns 410 - Account is closed",
        "type": "Negative",
        "remarks": "Closed account handling"
    },
    {
        "id": "TD_028",
        "tc_id": "CASA_DETAIL_TC_008",
        "desc": "Missing Authentication Token",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526"}),
        "result": "Returns 401 - Unauthorized",
        "type": "Negative",
        "remarks": "Auth check for detail endpoint"
    },
    {
        "id": "TD_029",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_001",
        "desc": "Valid Transaction History Retrieval",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "startDate": "2026-06-17", "endDate": "2026-07-17", "page": 1, "limit": 10}),
        "result": "Returns 200 with transactions array (max 10 items)",
        "type": "Positive",
        "remarks": "Last 30 days transaction history"
    },
    {
        "id": "TD_030",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_002",
        "desc": "Pagination - Second Page",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "startDate": "2026-06-01", "endDate": "2026-07-17", "page": 2, "limit": 20}),
        "result": "Returns 200 with page 2 data",
        "type": "Positive",
        "remarks": "Pagination support validated"
    },
    {
        "id": "TD_031",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_003",
        "desc": "Search Transactions by Keyword",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "startDate": "2026-06-17", "endDate": "2026-07-17", "keyword": "transfer"}),
        "result": "Returns 200 with filtered transactions containing 'transfer'",
        "type": "Positive",
        "remarks": "Keyword search functionality"
    },
    {
        "id": "TD_032",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_004",
        "desc": "Missing Start Date Parameter",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "endDate": "2026-07-17"}),
        "result": "Returns 400 - Missing mandatory parameter: startDate",
        "type": "Negative",
        "remarks": "Required parameter validation"
    },
    {
        "id": "TD_033",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_005",
        "desc": "Missing End Date Parameter",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "startDate": "2026-06-17"}),
        "result": "Returns 400 - Missing mandatory parameter: endDate",
        "type": "Negative",
        "remarks": "Required parameter validation"
    },
    {
        "id": "TD_034",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_006",
        "desc": "Invalid Date Format",
        "params": json.dumps({"startDate": "17-06-2026", "endDate": "2026-07-17"}),
        "result": "Returns 400 - Invalid date format. Use YYYY-MM-DD",
        "type": "Negative",
        "remarks": "Date format validation"
    },
    {
        "id": "TD_035",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_007",
        "desc": "Start Date After End Date",
        "params": json.dumps({"startDate": "2026-07-17", "endDate": "2026-06-17"}),
        "result": "Returns 400 - startDate must be before endDate",
        "type": "Negative",
        "remarks": "Date range validation"
    },
    {
        "id": "TD_036",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_008",
        "desc": "Date Range Exceeds 12 Months",
        "params": json.dumps({"startDate": "2024-01-01", "endDate": "2026-07-17"}),
        "result": "Returns 400 - Date range exceeds 12 months",
        "type": "Negative",
        "remarks": "Maximum date range enforcement"
    },
    {
        "id": "TD_037",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_009",
        "desc": "Invalid Page Number (< 1)",
        "params": json.dumps({"page": 0}),
        "result": "Returns 400 - Invalid page number",
        "type": "Negative",
        "remarks": "Page number boundary validation"
    },
    {
        "id": "TD_038",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_010",
        "desc": "Limit Exceeds Maximum (>100)",
        "params": json.dumps({"limit": 1000}),
        "result": "Returns 400 or caps to 100",
        "type": "Negative",
        "remarks": "Limit boundary validation"
    },
    {
        "id": "TD_039",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_011",
        "desc": "No Transactions Found",
        "params": json.dumps({"startDate": "2020-01-01", "endDate": "2020-01-31"}),
        "result": "Returns 200 with empty transactions array",
        "type": "Positive",
        "remarks": "Empty result handling"
    },
    {
        "id": "TD_040",
        "tc_id": "CASA_TRANSACTION_HISTORY_TC_012",
        "desc": "Unauthorized Account Access",
        "params": json.dumps({"accountType": "CA", "accountId": "514013314526", "customerRole": "OTHER"}),
        "result": "Returns 403 - Forbidden",
        "type": "Negative",
        "remarks": "Cross-customer security"
    },
    {
        "id": "TD_041",
        "tc_id": "CASA_TRANSACTION_DETAIL_TC_001",
        "desc": "Valid Transaction Detail Request",
        "params": json.dumps({"transactionReferenceId": "TXN20260717001234"}),
        "result": "Returns 200 with full transaction details",
        "type": "Positive",
        "remarks": "Complete transaction information returned"
    },
    {
        "id": "TD_042",
        "tc_id": "CASA_TRANSACTION_DETAIL_TC_002",
        "desc": "Invalid Transaction ID Format",
        "params": json.dumps({"transactionReferenceId": "INVALID"}),
        "result": "Returns 400 - Invalid transaction ID format",
        "type": "Negative",
        "remarks": "Transaction ID validation"
    },
    {
        "id": "TD_043",
        "tc_id": "CASA_TRANSACTION_DETAIL_TC_003",
        "desc": "Non-Existent Transaction",
        "params": json.dumps({"transactionReferenceId": "TXN99999999999"}),
        "result": "Returns 404 - Transaction not found",
        "type": "Negative",
        "remarks": "Transaction lookup failure"
    },
    {
        "id": "TD_044",
        "tc_id": "CASA_TRANSACTION_DETAIL_TC_004",
        "desc": "Unauthorized Transaction Access",
        "params": json.dumps({"transactionReferenceId": "TXN20260717001234", "customerRole": "OTHER"}),
        "result": "Returns 403 - Forbidden",
        "type": "Negative",
        "remarks": "Cross-customer transaction access"
    },
    {
        "id": "TD_045",
        "tc_id": "CASA_VIEW_RECEIPT_TC_001",
        "desc": "Valid Receipt Retrieval",
        "params": json.dumps({"transactionReferenceId": "TXN20260717001234"}),
        "result": "Returns 200 with PDF/image receipt",
        "type": "Positive",
        "remarks": "Receipt document retrieved"
    },
    {
        "id": "TD_046",
        "tc_id": "CASA_VIEW_RECEIPT_TC_002",
        "desc": "Receipt Not Available",
        "params": json.dumps({"transactionReferenceId": "TXN20260717009999"}),
        "result": "Returns 404 - Receipt not available",
        "type": "Negative",
        "remarks": "Pending transaction - no receipt"
    },
    {
        "id": "TD_047",
        "tc_id": "CASA_VIEW_RECEIPT_TC_003",
        "desc": "Invalid Transaction Reference",
        "params": json.dumps({"transactionReferenceId": "INVALID_ID"}),
        "result": "Returns 400 - Invalid reference format",
        "type": "Negative",
        "remarks": "Receipt request validation"
    },
    {
        "id": "TD_048",
        "tc_id": "CASA_MANAGE_LINKED_CARD_TC_001",
        "desc": "Get Linked Cards for Account",
        "params": json.dumps({"accountId": "900000097152"}),
        "result": "Returns 200 with array of linked cards",
        "type": "Positive",
        "remarks": "All linked cards retrieved"
    },
    {
        "id": "TD_049",
        "tc_id": "CASA_MANAGE_LINKED_CARD_TC_002",
        "desc": "Missing Account ID Parameter",
        "params": json.dumps({}),
        "result": "Returns 400 - Missing mandatory parameter",
        "type": "Negative",
        "remarks": "Account ID is required"
    },
    {
        "id": "TD_050",
        "tc_id": "CASA_MANAGE_LINKED_CARD_TC_003",
        "desc": "Invalid Account ID Format",
        "params": json.dumps({"accountId": "INVALID"}),
        "result": "Returns 400 - Invalid account ID format",
        "type": "Negative",
        "remarks": "Account ID format validation"
    },
    {
        "id": "TD_051",
        "tc_id": "CASA_AUTO_SWEEP_TC_001",
        "desc": "Get Auto-Sweep Settings",
        "params": json.dumps({"accountId": "123456789012"}),
        "result": "Returns 200 with auto-sweep configuration",
        "type": "Positive",
        "remarks": "Current auto-sweep status retrieved"
    },
    {
        "id": "TD_052",
        "tc_id": "CASA_AUTO_SWEEP_TC_002",
        "desc": "Invalid Account ID for Auto-Sweep",
        "params": json.dumps({"accountId": "INVALID"}),
        "result": "Returns 400 - Invalid account ID",
        "type": "Negative",
        "remarks": "Account ID validation for auto-sweep"
    },
    {
        "id": "TD_053",
        "tc_id": "PERSONAL_INFORMATION_TC_001",
        "desc": "Get Authenticated User's Personal Info",
        "params": json.dumps({"auth": "Bearer <VALID_TOKEN>"}),
        "result": "Returns 200 with user personal information",
        "type": "Positive",
        "remarks": "User profile data retrieved"
    },
    {
        "id": "TD_054",
        "tc_id": "LINK_DEBIT_CARD_TC_001",
        "desc": "Link Debit Card Successfully",
        "params": json.dumps({"accountId": "514013314526", "cardNumber": "5140133145267596", "primaryFlag": True}),
        "result": "Returns 202 with S2U confirmation request",
        "type": "Positive",
        "remarks": "S2U required for card linking"
    },
    {
        "id": "TD_055",
        "tc_id": "LINK_DEBIT_CARD_TC_002",
        "desc": "Missing Account ID",
        "params": json.dumps({"cardNumber": "5140133145267596"}),
        "result": "Returns 400 - Missing mandatory field: account.accountId",
        "type": "Negative",
        "remarks": "Account ID is required"
    },
    {
        "id": "TD_056",
        "tc_id": "LINK_DEBIT_CARD_TC_003",
        "desc": "Missing Card Number",
        "params": json.dumps({"accountId": "514013314526"}),
        "result": "Returns 400 - Missing mandatory field: card.linkedCardNumber",
        "type": "Negative",
        "remarks": "Card number is required"
    },
    {
        "id": "TD_057",
        "tc_id": "LINK_DEBIT_CARD_TC_004",
        "desc": "Invalid Card Number Format",
        "params": json.dumps({"accountId": "514013314526", "cardNumber": "1234"}),
        "result": "Returns 400 - Invalid card number format",
        "type": "Negative",
        "remarks": "Card number length validation"
    },
    {
        "id": "TD_058",
        "tc_id": "LINK_DEBIT_CARD_TC_005",
        "desc": "Card Fails Luhn Validation",
        "params": json.dumps({"cardNumber": "5140133145267595"}),
        "result": "Returns 400 - Invalid card number",
        "type": "Negative",
        "remarks": "Luhn algorithm validation"
    },
    {
        "id": "TD_059",
        "tc_id": "LINK_DEBIT_CARD_TC_006",
        "desc": "Card Already Linked",
        "params": json.dumps({"accountId": "514013314526", "cardNumber": "5140133145267596"}),
        "result": "Returns 409 - Card already linked",
        "type": "Negative",
        "remarks": "Duplicate link prevention"
    },
    {
        "id": "TD_060",
        "tc_id": "LINK_DEBIT_CARD_TC_007",
        "desc": "Account Not Found",
        "params": json.dumps({"accountId": "999999999999", "cardNumber": "5140133145267596"}),
        "result": "Returns 404 - Account not found",
        "type": "Negative",
        "remarks": "Account lookup failure"
    },
    {
        "id": "TD_061",
        "tc_id": "LINK_DEBIT_CARD_TC_008",
        "desc": "Cannot Link to Closed Account",
        "params": json.dumps({"accountId": "514013314526", "cardNumber": "5140133145267596", "accountStatus": "CLOSED"}),
        "result": "Returns 422 - Cannot link to closed account",
        "type": "Negative",
        "remarks": "Status validation"
    },
    {
        "id": "TD_062",
        "tc_id": "LINK_DEBIT_CARD_TC_009",
        "desc": "S2U Verification Required",
        "params": json.dumps({"accountId": "514013314526", "cardNumber": "5140133145267596", "requireS2U": True}),
        "result": "Returns 202 with S2U transaction details",
        "type": "Positive",
        "remarks": "Sensitive operation requires S2U"
    },
    {
        "id": "TD_063",
        "tc_id": "LINK_DEBIT_CARD_TC_010",
        "desc": "Malformed JSON Request",
        "params": "{invalid json}",
        "result": "Returns 400 - Invalid JSON format",
        "type": "Negative",
        "remarks": "JSON parsing error"
    },
    {
        "id": "TD_064",
        "tc_id": "LINK_DEBIT_CARD_TC_011",
        "desc": "SQL Injection in Phone Number",
        "params": json.dumps({"phoneNumber": "800225556814'; DROP TABLE customers; --"}),
        "result": "Returns 400 - Invalid phone format",
        "type": "Negative",
        "remarks": "SQL injection prevention"
    },
    {
        "id": "TD_065",
        "tc_id": "BLOCK_ACCOUNT_TC_001",
        "desc": "Block Single Account",
        "params": json.dumps({"accounts": [{"accountId": "183476335222"}]}),
        "result": "Returns 202 with S2U confirmation",
        "type": "Positive",
        "remarks": "Single account blocking"
    },
    {
        "id": "TD_066",
        "tc_id": "BLOCK_ACCOUNT_TC_002",
        "desc": "Block Multiple Accounts",
        "params": json.dumps({"accounts": [{"accountId": "183476335222"}, {"accountId": "898876334567"}]}),
        "result": "Returns 202 with batch S2U confirmation",
        "type": "Positive",
        "remarks": "Batch blocking operation"
    },
    {
        "id": "TD_067",
        "tc_id": "BLOCK_ACCOUNT_TC_003",
        "desc": "Missing Accounts Array",
        "params": json.dumps({}),
        "result": "Returns 400 - Missing mandatory field: accounts",
        "type": "Negative",
        "remarks": "Accounts array required"
    },
    {
        "id": "TD_068",
        "tc_id": "BLOCK_ACCOUNT_TC_004",
        "desc": "Empty Accounts Array",
        "params": json.dumps({"accounts": []}),
        "result": "Returns 400 - Array cannot be empty",
        "type": "Negative",
        "remarks": "At least one account required"
    },
    {
        "id": "TD_069",
        "tc_id": "BLOCK_ACCOUNT_TC_005",
        "desc": "Invalid Account ID in Array",
        "params": json.dumps({"accounts": [{"accountId": "INVALID"}]}),
        "result": "Returns 400 - Invalid account ID format",
        "type": "Negative",
        "remarks": "Format validation for each account"
    },
    {
        "id": "TD_070",
        "tc_id": "BLOCK_ACCOUNT_TC_006",
        "desc": "Non-Existent Account",
        "params": json.dumps({"accounts": [{"accountId": "999999999999"}]}),
        "result": "Returns 404 - Account not found",
        "type": "Negative",
        "remarks": "Account lookup"
    },
    {
        "id": "TD_071",
        "tc_id": "BLOCK_ACCOUNT_TC_007",
        "desc": "Account Already Blocked",
        "params": json.dumps({"accounts": [{"accountId": "183476335222"}], "accountStatus": "BLOCKED"}),
        "result": "Returns 409 - Already blocked",
        "type": "Negative",
        "remarks": "Duplicate block prevention"
    },
    {
        "id": "TD_072",
        "tc_id": "BLOCK_ACCOUNT_TC_008",
        "desc": "Cannot Block Closed Account",
        "params": json.dumps({"accounts": [{"accountId": "183476335222"}], "accountStatus": "CLOSED"}),
        "result": "Returns 422 - Cannot block closed account",
        "type": "Negative",
        "remarks": "Status validation"
    },
    {
        "id": "TD_073",
        "tc_id": "BLOCK_ACCOUNT_TC_009",
        "desc": "Missing S2U Authorization",
        "params": json.dumps({"accounts": [{"accountId": "183476335222"}]}),
        "result": "Returns 401 - S2U required",
        "type": "Negative",
        "remarks": "S2U authentication check"
    },
    {
        "id": "TD_074",
        "tc_id": "BLOCK_ACCOUNT_TC_010",
        "desc": "Duplicate Account IDs",
        "params": json.dumps({"accounts": [{"accountId": "183476335222"}, {"accountId": "183476335222"}]}),
        "result": "Returns 400 or deduplicates",
        "type": "Negative",
        "remarks": "Duplicate handling"
    },
    {
        "id": "TD_075",
        "tc_id": "REACTIVATE_ACCOUNT_TC_001",
        "desc": "Reactivate Inactive Account",
        "params": json.dumps({"accountId": "123456789012", "email": "customer@example.com", "annualIncome": 120000}),
        "result": "Returns 202 with S2U confirmation",
        "type": "Positive",
        "remarks": "Full reactivation workflow"
    },
    {
        "id": "TD_076",
        "tc_id": "REACTIVATE_ACCOUNT_TC_002",
        "desc": "Missing Account ID",
        "params": json.dumps({"email": "customer@example.com"}),
        "result": "Returns 400 - Missing accountId",
        "type": "Negative",
        "remarks": "Required field validation"
    },
    {
        "id": "TD_077",
        "tc_id": "REACTIVATE_ACCOUNT_TC_003",
        "desc": "Missing Email",
        "params": json.dumps({"accountId": "123456789012"}),
        "result": "Returns 400 - Missing email",
        "type": "Negative",
        "remarks": "Email required"
    },
    {
        "id": "TD_078",
        "tc_id": "REACTIVATE_ACCOUNT_TC_004",
        "desc": "Invalid Email Format",
        "params": json.dumps({"accountId": "123456789012", "email": "invalid-email@"}),
        "result": "Returns 400 - Invalid email format",
        "type": "Negative",
        "remarks": "Email validation"
    },
    {
        "id": "TD_079",
        "tc_id": "REACTIVATE_ACCOUNT_TC_005",
        "desc": "Account Not Found",
        "params": json.dumps({"accountId": "999999999999", "email": "test@example.com"}),
        "result": "Returns 404 - Account not found",
        "type": "Negative",
        "remarks": "Account lookup"
    },
    {
        "id": "TD_080",
        "tc_id": "REACTIVATE_ACCOUNT_TC_006",
        "desc": "Account Already Active",
        "params": json.dumps({"accountId": "123456789012", "email": "test@example.com", "accountStatus": "ACTIVE"}),
        "result": "Returns 409 - Already active",
        "type": "Negative",
        "remarks": "Status check"
    },
    {
        "id": "TD_081",
        "tc_id": "REACTIVATE_ACCOUNT_TC_007",
        "desc": "Cannot Reactivate Closed Account",
        "params": json.dumps({"accountId": "123456789012", "email": "test@example.com", "accountStatus": "CLOSED"}),
        "result": "Returns 422 - Cannot reactivate closed",
        "type": "Negative",
        "remarks": "Status validation"
    },
    {
        "id": "TD_082",
        "tc_id": "REACTIVATE_ACCOUNT_TC_008",
        "desc": "Invalid Annual Income (Negative)",
        "params": json.dumps({"accountId": "123456789012", "email": "test@example.com", "annualIncome": -50000}),
        "result": "Returns 400 - Invalid income",
        "type": "Negative",
        "remarks": "Income validation"
    },
    {
        "id": "TD_083",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_001",
        "desc": "Enable Auto-Sweep",
        "params": json.dumps({"accountId": "123456789012", "autoSweepFlag": "Y", "primaryCurrency": "IDR", "primaryLimit": 200000}),
        "result": "Returns 202 with S2U confirmation",
        "type": "Positive",
        "remarks": "Auto-sweep enabled"
    },
    {
        "id": "TD_084",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_002",
        "desc": "Disable Auto-Sweep",
        "params": json.dumps({"accountId": "123456789012", "autoSweepFlag": "N"}),
        "result": "Returns 202 with S2U confirmation",
        "type": "Positive",
        "remarks": "Auto-sweep disabled"
    },
    {
        "id": "TD_085",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_003",
        "desc": "Invalid Sweep Flag",
        "params": json.dumps({"autoSweepFlag": "MAYBE"}),
        "result": "Returns 400 - Invalid flag value",
        "type": "Negative",
        "remarks": "Flag validation"
    },
    {
        "id": "TD_086",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_004",
        "desc": "Invalid Currency Code",
        "params": json.dumps({"primaryCurrency": "XXX"}),
        "result": "Returns 400 - Invalid currency",
        "type": "Negative",
        "remarks": "Currency validation"
    },
    {
        "id": "TD_087",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_005",
        "desc": "Negative Sweep Limit",
        "params": json.dumps({"primaryLimit": -100000}),
        "result": "Returns 400 - Must be positive",
        "type": "Negative",
        "remarks": "Limit validation"
    },
    {
        "id": "TD_088",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_006",
        "desc": "Zero Sweep Limit",
        "params": json.dumps({"primaryLimit": 0}),
        "result": "Returns 400 - Must be > 0",
        "type": "Negative",
        "remarks": "Zero validation"
    },
    {
        "id": "TD_089",
        "tc_id": "UPDATE_AUTO_SWEEP_TC_007",
        "desc": "Sweep Limit Exceeds Balance",
        "params": json.dumps({"primaryLimit": 999999999999}),
        "result": "Returns 400 - Exceeds balance",
        "type": "Negative",
        "remarks": "Balance check"
    },
    {
        "id": "TD_090",
        "tc_id": "MANAGE_CASA_NICKNAME_ADD_TC_001",
        "desc": "Add Account Nickname",
        "params": json.dumps({"accountId": "123456789012", "nickname": "My Savings"}),
        "result": "Returns 200/201 with updated account",
        "type": "Positive",
        "remarks": "Nickname added successfully"
    },
    {
        "id": "TD_091",
        "tc_id": "MANAGE_CASA_NICKNAME_ADD_TC_002",
        "desc": "Missing Account ID",
        "params": json.dumps({"nickname": "My Savings"}),
        "result": "Returns 400 - Missing accountId",
        "type": "Negative",
        "remarks": "Account ID required"
    },
    {
        "id": "TD_092",
        "tc_id": "MANAGE_CASA_NICKNAME_ADD_TC_003",
        "desc": "Missing Nickname",
        "params": json.dumps({"accountId": "123456789012"}),
        "result": "Returns 400 - Missing nickname",
        "type": "Negative",
        "remarks": "Nickname required"
    },
    {
        "id": "TD_093",
        "tc_id": "MANAGE_CASA_NICKNAME_ADD_TC_004",
        "desc": "Empty Nickname",
        "params": json.dumps({"accountId": "123456789012", "nickname": ""}),
        "result": "Returns 400 - Cannot be empty",
        "type": "Negative",
        "remarks": "Empty validation"
    },
    {
        "id": "TD_094",
        "tc_id": "MANAGE_CASA_NICKNAME_ADD_TC_005",
        "desc": "Nickname Too Long (>50 chars)",
        "params": json.dumps({"accountId": "123456789012", "nickname": "x" * 51}),
        "result": "Returns 400 - Exceeds max length",
        "type": "Negative",
        "remarks": "Length validation"
    },
    {
        "id": "TD_095",
        "tc_id": "MANAGE_CASA_NICKNAME_UPDATE_TC_001",
        "desc": "Update Existing Nickname",
        "params": json.dumps({"accountId": "123456789012", "nickname": "Updated Savings"}),
        "result": "Returns 200 with updated nickname",
        "type": "Positive",
        "remarks": "Nickname updated"
    },
    {
        "id": "TD_096",
        "tc_id": "MANAGE_CASA_NICKNAME_DELETE_TC_001",
        "desc": "Delete Account Nickname",
        "params": json.dumps({"accountId": "123456789012"}),
        "result": "Returns 200/204 - Nickname removed",
        "type": "Positive",
        "remarks": "Nickname deleted"
    },
    {
        "id": "TD_097",
        "tc_id": "MANAGE_CASA_NICKNAME_DELETE_TC_002",
        "desc": "Delete Non-Existent Nickname",
        "params": json.dumps({"accountId": "123456789012"}),
        "result": "Returns 404 - Not found",
        "type": "Negative",
        "remarks": "Nickname not found"
    },
    {
        "id": "TD_098",
        "tc_id": "REMOVE_CURRENCY_TC_001",
        "desc": "Remove Currency from Account",
        "params": json.dumps({"accountId": "123456789012", "currency": "USD"}),
        "result": "Returns 200/204 - Currency removed",
        "type": "Positive",
        "remarks": "Currency deletion"
    },
    {
        "id": "TD_099",
        "tc_id": "REMOVE_CURRENCY_TC_002",
        "desc": "Invalid Currency Code",
        "params": json.dumps({"currency": "XXX"}),
        "result": "Returns 400 - Invalid code",
        "type": "Negative",
        "remarks": "Currency validation"
    },
    {
        "id": "TD_100",
        "tc_id": "REMOVE_CURRENCY_TC_003",
        "desc": "Cannot Remove Primary Currency",
        "params": json.dumps({"accountId": "123456789012", "currency": "IDR"}),
        "result": "Returns 422 - Cannot remove primary",
        "type": "Negative",
        "remarks": "Primary currency protection"
    },
    {
        "id": "TD_101",
        "tc_id": "REMOVE_CURRENCY_TC_004",
        "desc": "Currency Not in Account",
        "params": json.dumps({"accountId": "123456789012", "currency": "EUR"}),
        "result": "Returns 404 - Not found",
        "type": "Negative",
        "remarks": "Currency not available"
    },
]

# Write test data
for row_num, data in enumerate(test_data, 2):
    ws.cell(row=row_num, column=1).value = data["id"]
    ws.cell(row=row_num, column=2).value = data["tc_id"]
    ws.cell(row=row_num, column=3).value = data["desc"]
    ws.cell(row=row_num, column=4).value = data["params"]
    ws.cell(row=row_num, column=5).value = data["result"]
    ws.cell(row=row_num, column=6).value = data["type"]
    ws.cell(row=row_num, column=7).value = data["remarks"]
    
    # Apply styles
    for col in range(1, 8):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Alternate row coloring
        if row_num % 2 == 0:
            cell.fill = alt_row_fill

# Set row height for better readability
for row in range(1, len(test_data) + 2):
    ws.row_dimensions[row].height = 30

# Freeze header row
ws.freeze_panes = "A2"

# Save workbook
wb.save("/Users/aahmadkamar/Maybank/FSD Parser/artifacts/Account_Dashboard_API_Test_Data.xlsx")
print("✓ Test data Excel file generated successfully!")
print(f"✓ Total test data entries: {len(test_data)}")
print(f"✓ File saved to: artifacts/Account_Dashboard_API_Test_Data.xlsx")
