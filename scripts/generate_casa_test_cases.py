#!/usr/bin/env python3
"""
Generate comprehensive test cases for ECLIPSE Account Dashboard Casa DDD API
Following the reference format from PT_Local_Transfer_Test_Cases.md
"""

import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Casa API Endpoints Configuration
CASA_API_ENDPOINTS = {
    "getManageDebitCard": {
        "method": "GET",
        "endpoint": "/casa/v1/external/manage/debit-card",
        "description": "Retrieve debit card management details",
        "base_priority": "P1",
        "feature": "Casa Management",
        "source_worksheet": "Casa_API_Specification"
    },
    "getAutoSweep": {
        "method": "GET",
        "endpoint": "/casa/v1/external/auto-sweep",
        "description": "Retrieve auto-sweep settings for an account",
        "base_priority": "P1",
        "feature": "Casa Management",
        "source_worksheet": "Casa_API_Specification"
    },
    "setAutoSweep": {
        "method": "POST",
        "endpoint": "/casa/v1/external/auto-sweep",
        "description": "Update auto-sweep settings",
        "base_priority": "P1",
        "feature": "Casa Management",
        "source_worksheet": "Casa_API_Specification"
    },
    "setLinkedCard": {
        "method": "POST",
        "endpoint": "/casa/v1/external/linked-card",
        "description": "Link a card to an account",
        "base_priority": "P1",
        "feature": "Casa Management",
        "source_worksheet": "Casa_API_Specification"
    },
    "blockAccounts": {
        "method": "POST",
        "endpoint": "/casa/v1/external/accounts/block",
        "description": "Block one or more accounts",
        "base_priority": "P0",
        "feature": "Casa Management",
        "source_worksheet": "Casa_API_Specification"
    },
    "reactivateAccount": {
        "method": "POST",
        "endpoint": "/casa/v1/external/accounts/reactivate",
        "description": "Reactivate a blocked account",
        "base_priority": "P0",
        "feature": "Casa Management",
        "source_worksheet": "Casa_API_Specification"
    }
}

# Test case templates for different scenarios
TEST_SCENARIOS = {
    "POSITIVE": {
        "suffix": "-001",
        "scenario": "Successful API call with valid parameters",
        "test_type": "Positive",
        "priority_offset": 0,
        "expected_status": 200,
        "expected_result": "Request processed successfully"
    },
    "MISSING_AUTH": {
        "suffix": "-002",
        "scenario": "Missing authorization header",
        "test_type": "Negative",
        "priority_offset": 1,
        "expected_status": 401,
        "expected_result": "Unauthorized - Missing or invalid authorization header"
    },
    "INVALID_TOKEN": {
        "suffix": "-003",
        "scenario": "Invalid authorization token",
        "test_type": "Negative",
        "priority_offset": 1,
        "expected_status": 401,
        "expected_result": "Unauthorized - Invalid token signature"
    },
    "MISSING_CLIENT_ID": {
        "suffix": "-004",
        "scenario": "Missing required client-id header",
        "test_type": "Negative",
        "priority_offset": 1,
        "expected_status": 400,
        "expected_result": "Bad Request - Missing client-id header"
    },
    "INVALID_CONTENT_TYPE": {
        "suffix": "-005",
        "scenario": "Invalid Content-Type header (for POST methods)",
        "test_type": "Negative",
        "priority_offset": 1,
        "expected_status": 415,
        "expected_result": "Unsupported Media Type"
    },
    "MISSING_BODY": {
        "suffix": "-006",
        "scenario": "Missing required request body (for POST methods)",
        "test_type": "Negative",
        "priority_offset": 2,
        "expected_status": 400,
        "expected_result": "Bad Request - Missing required parameters"
    },
    "INVALID_DATA_TYPE": {
        "suffix": "-007",
        "scenario": "Invalid data type in request parameters",
        "test_type": "Negative",
        "priority_offset": 2,
        "expected_status": 400,
        "expected_result": "Bad Request - Invalid parameter data type"
    },
    "EMPTY_RESPONSE": {
        "suffix": "-008",
        "scenario": "Valid request with no matching data (GET methods)",
        "test_type": "Positive",
        "priority_offset": 2,
        "expected_status": 200,
        "expected_result": "Empty response array returned"
    },
    "BOUNDARY_VALUE": {
        "suffix": "-009",
        "scenario": "Boundary value testing - maximum allowed values",
        "test_type": "Negative",
        "priority_offset": 2,
        "expected_status": 400,
        "expected_result": "Bad Request - Value exceeds maximum allowed"
    },
    "NULL_VALUES": {
        "suffix": "-010",
        "scenario": "Null values in optional fields",
        "test_type": "Positive",
        "priority_offset": 2,
        "expected_status": 200,
        "expected_result": "Request accepted with null optional fields"
    }
}

class CasaTestCaseGenerator:
    def __init__(self):
        self.base_path = Path(__file__).parent.parent
        self.test_cases: List[Dict[str, Any]] = []
        self.markdown_content = ""
        self.excel_data = []
        self.generated_date = datetime.now().strftime("%Y-%m-%dT%H:%M:%S.000Z")
        self.test_case_counter = 0

    def generate_test_case_id(self, api_name: str, scenario_key: str) -> str:
        """Generate unique test case ID"""
        self.test_case_counter += 1
        scenario_suffix = TEST_SCENARIOS[scenario_key]["suffix"]
        return f"CASA_{api_name.upper()}_{scenario_suffix}"

    def get_priority(self, base_priority: str, scenario_key: str) -> str:
        """Calculate priority based on base priority and scenario"""
        priority_map = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
        base_level = priority_map.get(base_priority, 2)
        offset = TEST_SCENARIOS[scenario_key]["priority_offset"]
        
        new_level = min(base_level + offset, 3)
        level_map = {0: "P0", 1: "P1", 2: "P2", 3: "P3"}
        return level_map[new_level]

    def create_request_headers_list(self) -> List[str]:
        """Create standard request headers for Casa API"""
        return [
            "authorization: Bearer {token}",
            "client-id: mobile-banking",
            "content-type: application/json",
            "env: UAT",
            "x-request-id: {uuid}",
            "user-agent: MaybankMobileApp/v1.0"
        ]

    def create_test_case(self, api_name: str, api_config: Dict, scenario_key: str) -> Dict[str, Any]:
        """Create a single test case"""
        scenario = TEST_SCENARIOS[scenario_key]
        test_case_id = self.generate_test_case_id(api_name, scenario_key)
        priority = self.get_priority(api_config["base_priority"], scenario_key)
        
        test_case = {
            "id": test_case_id,
            "api_name": api_name,
            "feature": api_config["feature"],
            "description": f"{api_name} - {scenario['scenario']}",
            "priority": priority,
            "source_worksheet": api_config["source_worksheet"],
            "generated_date": self.generated_date,
            "test_scenario": scenario["scenario"],
            "test_objective": f"Verify {api_name} handles {scenario['scenario'].lower()}",
            "http_method": api_config["method"],
            "endpoint": api_config["endpoint"],
            "test_type": scenario["test_type"],
            "preconditions": [
                "Service endpoint is reachable",
                "Authentication credentials are available" if scenario_key != "MISSING_AUTH" else "Test environment is accessible"
            ],
            "request_headers": self.create_request_headers_list() if scenario_key != "MISSING_AUTH" else [
                "content-type: application/json",
                "env: UAT"
            ],
            "request_body": self.get_request_body(api_name, scenario_key),
            "test_steps": self.get_test_steps(api_name, scenario_key),
            "expected_status_code": scenario["expected_status"],
            "expected_result": scenario["expected_result"],
            "business_validation": self.get_business_validation(api_name, scenario_key),
            "tags": [api_config["method"].lower(), scenario["test_type"].lower()],
            "environment": ["SIT", "UAT"]
        }
        
        return test_case

    def get_request_body(self, api_name: str, scenario_key: str) -> Dict[str, Any]:
        """Get request body based on API and scenario"""
        if "GET" in CASA_API_ENDPOINTS[api_name]["method"]:
            return {}
        
        body_templates = {
            "setAutoSweep": {
                "accountNumber": "1234567890" if scenario_key != "MISSING_BODY" else None,
                "autoSweepEnabled": True if scenario_key != "INVALID_DATA_TYPE" else "yes",
                "sweepAmount": 1000.00 if scenario_key != "BOUNDARY_VALUE" else 999999999.99,
                "sweepFrequency": "DAILY"
            },
            "setLinkedCard": {
                "accountNumber": "1234567890" if scenario_key != "MISSING_BODY" else None,
                "cardNumber": "4532XXXXXXXX1234" if scenario_key != "INVALID_DATA_TYPE" else 4532,
                "cardType": "DEBIT",
                "isPrimary": True if scenario_key != "INVALID_DATA_TYPE" else "true",
                "linkedCardName": "My Primary Card"
            },
            "blockAccounts": {
                "accountNumbers": ["1234567890"] if scenario_key != "MISSING_BODY" else None,
                "blockReason": "LOST_CARD",
                "blockDurationDays": 30
            },
            "reactivateAccount": {
                "accountNumber": "1234567890" if scenario_key != "MISSING_BODY" else None,
                "verificationCode": "123456" if scenario_key != "INVALID_DATA_TYPE" else 123456,
                "reactivationReason": "FOUND_CARD"
            }
        }
        
        return body_templates.get(api_name, {})

    def get_test_steps(self, api_name: str, scenario_key: str) -> List[str]:
        """Get test steps based on scenario"""
        if scenario_key == "MISSING_AUTH":
            return [
                "1. Prepare request without authorization header",
                "2. Send request to endpoint: " + CASA_API_ENDPOINTS[api_name]["endpoint"],
                "3. Verify response status code is 401",
                "4. Verify error message indicates missing authorization"
            ]
        elif scenario_key == "INVALID_TOKEN":
            return [
                "1. Prepare request with invalid bearer token",
                "2. Send request to endpoint: " + CASA_API_ENDPOINTS[api_name]["endpoint"],
                "3. Verify response status code is 401",
                "4. Verify error message indicates invalid token"
            ]
        elif scenario_key == "MISSING_BODY" and "POST" in CASA_API_ENDPOINTS[api_name]["method"]:
            return [
                "1. Prepare POST request without body",
                "2. Send request to endpoint: " + CASA_API_ENDPOINTS[api_name]["endpoint"],
                "3. Verify response status code is 400",
                "4. Verify error message indicates missing required parameters"
            ]
        else:
            method = CASA_API_ENDPOINTS[api_name]["method"]
            return [
                f"1. Prepare {method} request with valid parameters",
                "2. Add required headers (authorization, client-id, content-type)",
                "3. Send request to endpoint: " + CASA_API_ENDPOINTS[api_name]["endpoint"],
                f"4. Verify response status code is {TEST_SCENARIOS[scenario_key]['expected_status']}",
                "5. Verify response body matches expected schema",
                "6. Verify business logic validation rules are applied"
            ]

    def get_business_validation(self, api_name: str, scenario_key: str) -> List[str]:
        """Get business validation rules"""
        validations = {
            "getManageDebitCard": [
                "Verify user has permission to view debit card details",
                "Verify card status is active or valid",
                "Verify response contains PAN last 4 digits only"
            ],
            "getAutoSweep": [
                "Verify account exists and belongs to user",
                "Verify auto-sweep settings are correctly retrieved",
                "Verify sweep frequency is valid (DAILY/WEEKLY/MONTHLY)"
            ],
            "setAutoSweep": [
                "Verify auto-sweep amount is within acceptable range",
                "Verify sweep frequency is valid",
                "Verify account has sufficient funds for sweep operation"
            ],
            "setLinkedCard": [
                "Verify card number format is valid",
                "Verify card belongs to customer",
                "Verify only one primary card is linked per account",
                "Verify card status is active"
            ],
            "blockAccounts": [
                "Verify block reason is valid",
                "Verify account can be blocked (not already blocked)",
                "Verify user has authorization to block account",
                "Verify blocking does not affect linked transactions"
            ],
            "reactivateAccount": [
                "Verify account is in blocked state",
                "Verify verification code is correct",
                "Verify reactivation reason is valid",
                "Verify user has authorization to reactivate"
            ]
        }
        
        return validations.get(api_name, ["Verify API behavior matches specification"])

    def generate_all_test_cases(self):
        """Generate test cases for all APIs and scenarios"""
        print("Generating test cases for Casa API...")
        
        for api_name, api_config in CASA_API_ENDPOINTS.items():
            print(f"  Processing {api_name}...")
            
            # For GET endpoints: include POSITIVE, EMPTY_RESPONSE, MISSING_AUTH, INVALID_TOKEN, MISSING_CLIENT_ID, NULL_VALUES
            # For POST endpoints: include all scenarios
            
            if "GET" in api_config["method"]:
                scenarios = ["POSITIVE", "EMPTY_RESPONSE", "MISSING_AUTH", "INVALID_TOKEN", "MISSING_CLIENT_ID", "NULL_VALUES"]
            else:
                scenarios = ["POSITIVE", "MISSING_AUTH", "INVALID_TOKEN", "MISSING_CLIENT_ID", 
                           "MISSING_BODY", "INVALID_DATA_TYPE", "BOUNDARY_VALUE", "INVALID_CONTENT_TYPE"]
            
            for scenario_key in scenarios:
                test_case = self.create_test_case(api_name, api_config, scenario_key)
                self.test_cases.append(test_case)

        print(f"Generated {len(self.test_cases)} test cases")

    def generate_markdown(self):
        """Generate Markdown test cases documentation"""
        print("Generating Markdown documentation...")
        
        # Create table of contents
        toc = "# Account Dashboard Casa DDD API - Test Cases\n\n"
        toc += f"**Generated Date:** {self.generated_date}\n"
        toc += f"**Total Test Cases:** {len(self.test_cases)}\n\n"
        toc += "---\n\n## Table of Contents\n\n"
        
        for idx, test_case in enumerate(self.test_cases, 1):
            toc += f"{idx}. [{test_case['id']} - {test_case['http_method']} {test_case['endpoint']}](#{test_case['id'].lower()})\n"
        
        toc += "\n---\n\n"
        
        # Generate test case details
        content = toc
        for idx, test_case in enumerate(self.test_cases, 1):
            content += f"## {idx}. {test_case['id']} - {test_case['http_method']} {test_case['endpoint']} {{#{test_case['id'].lower()}}}\n\n"
            content += f"**Feature:** {test_case['feature']}\n"
            content += f"**Description:** {test_case['description']}\n"
            content += f"**Priority:** {test_case['priority']}\n"
            content += f"**Source Worksheet:** {test_case['source_worksheet']}\n"
            content += f"**Generated Date:** {self.generated_date.split('T')[0]}\n\n"
            
            content += "### Tags\n"
            for tag in test_case['tags']:
                content += f"- `{tag}`\n"
            content += "\n"
            
            content += "### Environment\n"
            for env in test_case['environment']:
                content += f"- {env}\n"
            content += "\n"
            
            content += "### Authentication\n"
            content += "Bearer Token\n\n"
            
            content += "### Preconditions\n"
            for precond in test_case['preconditions']:
                content += f"- {precond}\n"
            content += "\n"
            
            content += "### Test Scenario\n"
            content += f"{test_case['test_scenario']}\n\n"
            
            content += "### Test Objective\n"
            content += f"{test_case['test_objective']}\n\n"
            
            content += "### Request Details\n"
            content += "| Property | Value |\n"
            content += "|----------|-------|\n"
            content += f"| Method | {test_case['http_method']} |\n"
            content += f"| Endpoint | `{test_case['endpoint']}` |\n"
            content += f"| Headers | {len(test_case['request_headers'])} required |\n"
            content += f"| Request Body | {'JSON' if test_case['request_body'] else 'None'} |\n\n"
            
            content += "### Request Headers\n"
            for header in test_case['request_headers']:
                content += f"- {header}\n"
            content += "\n"
            
            if test_case['request_body']:
                content += "### Request Body\n"
                content += "```json\n"
                content += json.dumps(test_case['request_body'], indent=2) + "\n"
                content += "```\n\n"
            
            content += "### Test Steps\n"
            for step in test_case['test_steps']:
                content += f"{step}\n"
            content += "\n"
            
            content += "### Expected Response\n"
            content += f"**Status Code:** `{test_case['expected_status_code']}`\n"
            content += f"**Description:** {test_case['expected_result']}\n\n"
            
            content += "### Business Validation\n"
            for validation in test_case['business_validation']:
                content += f"- {validation}\n"
            content += "\n"
            
            content += "### Test Type\n"
            content += f"{test_case['test_type']}\n\n"
            
            content += "### Cleanup\n"
            content += "- Verify no side effects or orphaned data remain\n"
            content += "- Reset test data if necessary for subsequent test runs\n\n"
            
            content += "---\n\n"
        
        self.markdown_content = content

    def generate_excel(self):
        """Generate Excel test cases documentation"""
        print("Generating Excel workbook...")
        
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        # Create main test cases sheet
        ws_cases = workbook.create_sheet("Test Cases")
        
        # Define headers
        headers = [
            "Test Case ID",
            "API Name",
            "HTTP Method",
            "Endpoint",
            "Test Scenario",
            "Test Type",
            "Priority",
            "Expected Status Code",
            "Expected Result",
            "Test Objective",
            "Preconditions",
            "Request Headers",
            "Request Body",
            "Test Steps",
            "Business Validation",
            "Tags",
            "Environment",
            "Generated Date"
        ]
        
        # Add headers
        ws_cases.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws_cases[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        
        # Add test cases
        for test_case in self.test_cases:
            row = [
                test_case["id"],
                test_case["api_name"],
                test_case["http_method"],
                test_case["endpoint"],
                test_case["test_scenario"],
                test_case["test_type"],
                test_case["priority"],
                test_case["expected_status_code"],
                test_case["expected_result"],
                test_case["test_objective"],
                "; ".join(test_case["preconditions"]),
                "; ".join(test_case["request_headers"]),
                json.dumps(test_case["request_body"]) if test_case["request_body"] else "N/A",
                "; ".join(test_case["test_steps"]),
                "; ".join(test_case["business_validation"]),
                ", ".join(test_case["tags"]),
                ", ".join(test_case["environment"]),
                test_case["generated_date"]
            ]
            ws_cases.append(row)
        
        # Adjust column widths
        column_widths = {
            'A': 20, 'B': 18, 'C': 12, 'D': 35, 'E': 30, 'F': 12, 'G': 10, 
            'H': 15, 'I': 20, 'J': 25, 'K': 20, 'L': 30, 'M': 20, 'N': 30, 
            'O': 30, 'P': 20, 'Q': 15, 'R': 20
        }
        
        for col, width in column_widths.items():
            ws_cases.column_dimensions[col].width = width
        
        # Create summary sheet
        ws_summary = workbook.create_sheet("Summary", 0)
        
        ws_summary['A1'] = "Account Dashboard Casa DDD API - Test Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        ws_summary['A3'] = "Generated Date:"
        ws_summary['B3'] = self.generated_date
        
        ws_summary['A4'] = "Total Test Cases:"
        ws_summary['B4'] = len(self.test_cases)
        
        # Count by type
        test_types = {}
        for tc in self.test_cases:
            test_type = tc["test_type"]
            test_types[test_type] = test_types.get(test_type, 0) + 1
        
        row = 6
        ws_summary['A6'] = "Test Cases by Type:"
        ws_summary['A6'].font = Font(bold=True)
        
        row = 7
        for test_type, count in sorted(test_types.items()):
            ws_summary[f'A{row}'] = test_type
            ws_summary[f'B{row}'] = count
            row += 1
        
        # Count by priority
        row += 2
        ws_summary[f'A{row}'] = "Test Cases by Priority:"
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        priorities = {}
        for tc in self.test_cases:
            priority = tc["priority"]
            priorities[priority] = priorities.get(priority, 0) + 1
        
        row += 1
        for priority in sorted(priorities.keys()):
            ws_summary[f'A{row}'] = priority
            ws_summary[f'B{row}'] = priorities[priority]
            row += 1
        
        # Save workbook
        output_path = self.base_path / "artifacts" / "Test_Case" / "Excel" / "Account_Dashboard_Casa_DDD_Test_Cases.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        print(f"Excel file saved: {output_path}")

    def generate_test_data_excel(self):
        """Generate Excel test data matching reference format"""
        print("Generating test data Excel workbook...")
        
        workbook = openpyxl.Workbook()
        
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        ws = workbook.create_sheet("Test Data")
        
        # Headers for test data
        headers = [
            "Test Case ID",
            "API Name",
            "HTTP Method",
            "Endpoint",
            "Request Header - Authorization",
            "Request Header - Client-ID",
            "Request Header - Content-Type",
            "Request Body - Field 1",
            "Request Body - Field 2",
            "Request Body - Field 3",
            "Expected Status Code",
            "Expected Response - Field 1",
            "Expected Response - Field 2",
            "Notes"
        ]
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        
        # Add test data rows
        for test_case in self.test_cases:
            # Generate test data based on scenario
            auth_header = "Bearer {VALID_TOKEN}" if "MISSING_AUTH" not in test_case["id"] else "OMITTED"
            client_id = "mobile-banking"
            content_type = "application/json" if test_case["request_body"] else "N/A"
            
            request_body = json.dumps(test_case["request_body"]) if test_case["request_body"] else "N/A"
            
            # Extract first 3 fields from request body for display
            body_fields = list(test_case["request_body"].values())[:3] if test_case["request_body"] else ["N/A", "N/A", "N/A"]
            
            row = [
                test_case["id"],
                test_case["api_name"],
                test_case["http_method"],
                test_case["endpoint"],
                auth_header,
                client_id,
                content_type,
                str(body_fields[0]) if len(body_fields) > 0 else "",
                str(body_fields[1]) if len(body_fields) > 1 else "",
                str(body_fields[2]) if len(body_fields) > 2 else "",
                test_case["expected_status_code"],
                test_case["expected_result"],
                test_case["test_type"],
                f"Scenario: {test_case['test_scenario']}"
            ]
            
            ws.append(row)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save workbook
        output_path = self.base_path / "artifacts" / "Test_Case" / "Excel" / "Account_Dashboard_Casa_DDD_Test_Data.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        print(f"Test data file saved: {output_path}")

    def save_markdown(self):
        """Save Markdown file"""
        output_path = self.base_path / "artifacts" / "Test_Case" / "Markdown" / "Account_Dashboard_Casa_DDD_Test_Cases.md"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(self.markdown_content)
        
        print(f"Markdown file saved: {output_path}")

    def generate(self):
        """Main generation method"""
        print("\n" + "="*80)
        print("Casa DDD API Test Case Generator")
        print("="*80 + "\n")
        
        self.generate_all_test_cases()
        self.generate_markdown()
        self.generate_excel()
        self.generate_test_data_excel()
        self.save_markdown()
        
        print("\n" + "="*80)
        print(f"✓ Test case generation completed successfully!")
        print(f"✓ Total test cases generated: {len(self.test_cases)}")
        print("="*80 + "\n")
        
        print("Generated deliverables:")
        print(f"  1. Markdown: artifacts/Test_Case/Markdown/Account_Dashboard_Casa_DDD_Test_Cases.md")
        print(f"  2. Test Cases Excel: artifacts/Test_Case/Excel/Account_Dashboard_Casa_DDD_Test_Cases.xlsx")
        print(f"  3. Test Data Excel: artifacts/Test_Case/Excel/Account_Dashboard_Casa_DDD_Test_Data.xlsx")
        print("\n")


if __name__ == "__main__":
    generator = CasaTestCaseGenerator()
    generator.generate()
