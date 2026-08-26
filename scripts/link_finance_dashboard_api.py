#!/usr/bin/env python3
"""Link Finance and Non Financing Dashboard test cases to ECLIPSE APIs."""

from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
import re

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from zipfile import BadZipFile

ROOT = Path(__file__).resolve().parents[1]
QA_FILE = ROOT / "input" / "Journey_Test_cases" / "Finance and Non Financing Dashboard.xlsx"
API_DIR = ROOT / "input" / "api"
OUTPUT_BASE = ROOT / "artifacts" / "NGA_Batch1_QA_Test_Scenarios_and_Cases_4_Loan_API_Linked.xlsx"

PRIMARY_FILES = [
    ("Credit Card API Design", "ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx", ("credit card", "credit-card")),
    ("Debit Card API Design", "ECLIPSE_Account Dashboard_Debit_Card_DDD_API_Design_v1_Workshop (1).xlsx", ("debit card", "debit-card")),
    ("Insurance API Design", "ECLIPSE_Account Dashboard_Insurance_DDD_API_Design_v1_Workshop.xlsx", ("insurance", "takaful")),
    ("Loan API Design", "ECLIPSE_Account Dashboard_Loan_DDD_API_Design_v1_Workshop.xlsx", ("loan", "financing", "mortgage")),
    ("Casa API Design", "ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx", ("casa", "savings", "current account", "deposit")),
    ("Maintenance API Design", "ECLIPSE_Maintenance_DDD_API_Design_v1.xlsx", ("maintenance", "cms", "translation", "i18n")),
]
FALLBACK_FILE = "ECLIPSE_Dashboard_Common_DDD_API_Design_v1_Workshop.xlsx"
TRACE_HEADERS = [
    "API Source Priority", "API Name", "HTTP Method", "Endpoint", "Microservice Ownership",
    "API Source Workbook", "API Source Sheet", "API Source Reference", "Traceability Link",
    "Mapping Rationale", "Mapping Confidence", "API Validation / Notes",
]
CATALOG_HEADERS = [
    "API ID", "Source Priority", "Journey / Feature", "API Name", "HTTP Method", "Endpoint",
    "Microservice Ownership", "Source Workbook", "Source Sheet", "Source Reference", "Status",
    "Remarks / Limitations", "Mapped Test Case Count",
]
SOURCE_ISSUES = []


def clean(value):
    return "" if value is None else str(value).strip()


def compact(value, limit=900):
    value = re.sub(r"\s+", " ", clean(value))
    return value if len(value) <= limit else value[: limit - 3] + "..."


def terms(value):
    return set(re.findall(r"[a-z0-9]+", clean(value).lower()))


def source_ref(sheet, row, end_col):
    return f"{sheet}!A{row}:{get_column_letter(end_col)}{row}"


def load_primary_operations():
    operations = []
    for priority, filename, domains in PRIMARY_FILES:
        path = API_DIR / filename
        try:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except (BadZipFile, OSError, ValueError) as error:
            SOURCE_ISSUES.append(f"{filename}: source could not be parsed ({error}).")
            continue
        sheet = workbook["API_Specs_Index"]
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            values = list(row)
            api_name = clean(values[2] if len(values) > 2 else "")
            method = clean(values[3] if len(values) > 3 else "").upper()
            owner = clean(values[4] if len(values) > 4 else "")
            endpoint = clean(values[5] if len(values) > 5 else "")
            detail_sheet = clean(values[1] if len(values) > 1 else "")
            status = clean(values[6] if len(values) > 6 else "")
            remarks = clean(values[7] if len(values) > 7 else "")
            if api_name and method and endpoint and detail_sheet:
                operations.append({
                    "priority": priority, "filename": filename, "sheet": detail_sheet,
                    "api_name": api_name, "method": method, "endpoint": endpoint,
                    "owner": owner, "status": status, "remarks": remarks,
                    "index_ref": source_ref("API_Specs_Index", row_number, min(len(values), 8)),
                    "domains": domains,
                })
    return operations


def load_common_operations():
    filename = FALLBACK_FILE
    try:
        workbook = openpyxl.load_workbook(API_DIR / filename, data_only=True, read_only=True)
    except (BadZipFile, OSError, ValueError) as error:
        SOURCE_ISSUES.append(f"{filename}: source could not be parsed ({error}).")
        return []
    operations = []
    for index_sheet_name in ("XP APIs >>", "S2U APIs >>", "ECLIPSE APIs >>"):
        if index_sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[index_sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            values = list(row)
            api_name = clean(values[1] if len(values) > 1 else "")
            method = clean(values[2] if len(values) > 2 else "").upper()
            owner = clean(values[3] if len(values) > 3 else "")
            endpoint = clean(values[4] if len(values) > 4 else "")
            journey = clean(values[5] if len(values) > 5 else "")
            detail_sheet = clean(values[6] if len(values) > 6 else "")
            status = clean(values[7] if len(values) > 7 else "")
            remarks = clean(values[8] if len(values) > 8 else "")
            if api_name and method and endpoint:
                operations.append({
                    "priority": "Common API Design (Fallback)", "filename": filename,
                    "sheet": detail_sheet or index_sheet_name, "api_name": api_name,
                    "method": method, "endpoint": endpoint, "owner": owner,
                    "status": status, "remarks": remarks, "journey": journey,
                    "index_ref": source_ref(index_sheet_name, row_number, min(len(values), 9)),
                    "domains": (),
                })
    return operations


def row_domain(row_values):
    text = " ".join(clean(value).lower() for value in row_values)
    if "credit card" in text or "credit-card" in text:
        return "Credit Card API Design"
    if "debit card" in text or "debit-card" in text:
        return "Debit Card API Design"
    if "insurance" in text or "takaful" in text:
        return "Insurance API Design"
    if any(value in text for value in ("loan", "financing", "mortgage")):
        return "Loan API Design"
    if any(value in text for value in ("casa", "savings", "current account", "deposit")):
        return "Casa API Design"
    return "Maintenance API Design"


def operation_score(case_text, operation, domain):
    api_text = " ".join((operation["api_name"], operation["sheet"], operation["endpoint"], operation.get("journey", "")))
    api_terms = {term for term in terms(api_text) if len(term) > 3}
    score = sum(1 for term in api_terms if term in case_text)
    if operation["priority"] == domain:
        score += 5
    if operation["priority"] == "Common API Design (Fallback)":
        score -= 1
    case_actions = case_text & {"search", "filter", "export", "statement", "receipt", "nickname", "summary", "detail", "transaction", "transactions", "dashboard", "listing", "activation", "block", "unblock", "pin", "primary"}
    score += sum(2 for action in case_actions if action in terms(api_text))
    return score


def choose_operation(case_values, operations, domain):
    case_text = terms(" ".join(clean(value).lower() for value in case_values))
    primary = [operation for operation in operations if operation["priority"] != "Common API Design (Fallback)" and (operation["priority"] == domain or domain == "Maintenance API Design")]
    scored = sorted(((operation_score(case_text, op, domain), op) for op in primary), key=lambda item: item[0], reverse=True)
    if scored and scored[0][0] >= 4:
        return scored[0][1], scored[0][0]
    fallback = sorted(((operation_score(case_text, op, domain), op) for op in operations if op["priority"] == "Common API Design (Fallback)"), key=lambda item: item[0], reverse=True)
    if fallback and fallback[0][0] >= 4:
        return fallback[0][1], fallback[0][0]
    return None, 0


def detail_notes(operation):
    workbook = openpyxl.load_workbook(API_DIR / operation["filename"], data_only=True, read_only=True)
    if operation["sheet"] not in workbook.sheetnames:
        return ""
    sheet = workbook[operation["sheet"]]
    facts = []
    for row in sheet.iter_rows(values_only=True):
        line = " ".join(clean(value) for value in row if clean(value))
        if line and any(word in line.lower() for word in ("mandatory", "validation", "pagination", "minimum", "feature flag", "pending", "closed account", "date format")):
            facts.append(compact(line, 250))
    return "; ".join(dict.fromkeys(facts[:4]))


def copy_sheet_style(source, target):
    for row in source.iter_rows():
        for cell in row:
            target_cell = target[cell.coordinate]
            if cell.has_style:
                target_cell._style = copy(cell._style)
            if cell.number_format:
                target_cell.number_format = cell.number_format
            if cell.alignment:
                target_cell.alignment = copy(cell.alignment)
            if cell.protection:
                target_cell.protection = copy(cell.protection)
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height


def add_readme_section(workbook, counts):
    sheet = workbook["README"] if "README" in workbook.sheetnames else workbook.create_sheet("README")
    start = sheet.max_row + 2
    sheet.cell(start, 1, "API Traceability")
    sheet.cell(start, 1).font = Font(bold=True, size=14, color="FFFFFF")
    sheet.cell(start, 1).fill = PatternFill("solid", fgColor="1F4E78")
    lines = [
        "Primary sources: the applicable Account Dashboard or Maintenance DDD API Design workbook is searched first.",
        "Fallback source: the Common API Design workbook is used only when no defensible primary mapping is available.",
        "Unmapped cases have no defensible API mapping in the supplied sources; confidence should be reviewed before execution or sign-off.",
        f"Counts: {counts['primary']} primary, {counts['fallback']} fallback, {counts['mixed']} mixed, {counts['unmapped']} unmapped; traceability {counts['percentage']:.1f}%.",
    ]
    if SOURCE_ISSUES:
        lines.append("Source gaps: " + " ".join(SOURCE_ISSUES))
    for offset, line in enumerate(lines, 1):
        sheet.cell(start + offset, 1, line)
        sheet.cell(start + offset, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, 120)


def build_workbook():
    SOURCE_ISSUES.clear()
    workbook = openpyxl.load_workbook(QA_FILE)
    test_sheet = workbook["Test Cases"]
    existing_headers = [clean(test_sheet.cell(1, col).value) for col in range(1, test_sheet.max_column + 1)]
    header_map = {header: index for index, header in enumerate(existing_headers, 1) if header}
    operations = load_primary_operations() + load_common_operations()
    catalog = []
    catalog_keys = {}
    for operation in operations:
        key = (operation["priority"], operation["api_name"], operation["method"], operation["endpoint"])
        if key not in catalog_keys:
            catalog_keys[key] = len(catalog) + 1
            catalog.append({**operation, "catalog_id": f"API-{len(catalog) + 1:03d}", "mapped_count": 0})

    api_by_key = {(item["priority"], item["api_name"], item["method"], item["endpoint"]): item for item in catalog}
    mapped_rows = []
    source_counts = Counter()
    confidence_counts = Counter()
    module_counts = Counter()
    for row_number in range(2, test_sheet.max_row + 1):
        values = [test_sheet.cell(row_number, col).value for col in range(1, len(existing_headers) + 1)]
        if not any(clean(value) for value in values):
            continue
        case_id = clean(values[header_map.get("Test Case ID", 1) - 1])
        domain = row_domain(values)
        operation, score = choose_operation(values, operations, domain)
        if operation:
            key = (operation["priority"], operation["api_name"], operation["method"], operation["endpoint"])
            catalog_item = api_by_key[key]
            catalog_item["mapped_count"] += 1
            catalog_item.setdefault("first_test_row", row_number)
            source_counts[operation["priority"]] += 1
            module_counts[clean(values[header_map.get("Module", 3) - 1]) or "Unspecified"] += 1
            confidence = "High" if score >= 9 and operation["priority"] != "Common API Design (Fallback)" else "Medium"
            confidence_counts[confidence] += 1
            notes = "; ".join(filter(None, [operation.get("remarks", ""), detail_notes(operation)]))
            rationale = f"Mapped to {operation['api_name']} because the test case terms align with the {operation['sheet']} journey in the applicable API specification."
            trace = {
                "priority": operation["priority"], "api_name": operation["api_name"], "method": operation["method"],
                "endpoint": operation["endpoint"], "owner": operation["owner"], "filename": operation["filename"],
                "sheet": operation["sheet"], "reference": operation["index_ref"], "catalog_id": catalog_item["catalog_id"],
                "rationale": rationale, "confidence": confidence, "notes": notes or "No additional validation notes were documented in the source index.",
            }
        else:
            source_counts["Not Mapped"] += 1
            confidence_counts["Not Mapped"] += 1
            trace = {"priority": "Not Mapped", "api_name": "", "method": "", "endpoint": "", "owner": "", "filename": "", "sheet": "", "reference": "", "catalog_id": "", "rationale": "No defensible API relationship was identified in the applicable primary or Common API specifications.", "confidence": "Not Mapped", "notes": "Review with the BA or API owner if an API dependency is expected."}
        mapped_rows.append((row_number, case_id, trace))

    start_col = len(existing_headers) + 1
    for offset, header in enumerate(TRACE_HEADERS):
        cell = test_sheet.cell(1, start_col + offset, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    catalog_sheet = workbook.create_sheet("ECLIPSE API Catalog") if "ECLIPSE API Catalog" not in workbook.sheetnames else workbook["ECLIPSE API Catalog"]
    if catalog_sheet.max_row > 1:
        catalog_sheet.delete_rows(1, catalog_sheet.max_row)
    for col, header in enumerate(CATALOG_HEADERS, 1):
        cell = catalog_sheet.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    catalog_row_by_id = {}
    for row_index, item in enumerate(catalog, 2):
        catalog_row_by_id[item["catalog_id"]] = row_index
        values = [item["catalog_id"], item["priority"], item.get("journey", item["sheet"]), item["api_name"], item["method"], item["endpoint"], item["owner"], item["filename"], item["sheet"], item["index_ref"], item.get("status", ""), item.get("remarks", ""), item["mapped_count"]]
        for col, value in enumerate(values, 1):
            catalog_sheet.cell(row_index, col, value)
            catalog_sheet.cell(row_index, col).alignment = Alignment(vertical="top", wrap_text=True)
        catalog_sheet.cell(row_index, 1).hyperlink = f"#'Test Cases'!A1"
        catalog_sheet.cell(row_index, 1).style = "Hyperlink"

    for row_number, _, trace in mapped_rows:
        values = [trace["priority"], trace["api_name"], trace["method"], trace["endpoint"], trace["owner"], trace["filename"], trace["sheet"], trace["reference"], trace["catalog_id"], trace["rationale"], trace["confidence"], trace["notes"]]
        for offset, value in enumerate(values):
            cell = test_sheet.cell(row_number, start_col + offset, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        link_cell = test_sheet.cell(row_number, start_col + 8)
        if trace["catalog_id"]:
            link_cell.hyperlink = f"#'ECLIPSE API Catalog'!A{catalog_row_by_id[trace['catalog_id']]}"
            link_cell.style = "Hyperlink"

    end_col = start_col + len(TRACE_HEADERS) - 1
    test_sheet.freeze_panes = "A2"
    test_sheet.auto_filter.ref = f"A1:{get_column_letter(end_col)}{test_sheet.max_row}"
    widths = {start_col: 26, start_col + 1: 30, start_col + 2: 12, start_col + 3: 48, start_col + 4: 24, start_col + 5: 55, start_col + 6: 28, start_col + 7: 30, start_col + 8: 18, start_col + 9: 58, start_col + 10: 18, start_col + 11: 65}
    for col, width in widths.items():
        test_sheet.column_dimensions[get_column_letter(col)].width = width
    for row in test_sheet.iter_rows(min_row=2, max_row=test_sheet.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    catalog_sheet.freeze_panes = "A2"
    catalog_sheet.auto_filter.ref = f"A1:M{catalog_sheet.max_row}"
    for col, width in enumerate((14, 28, 28, 32, 12, 48, 26, 58, 30, 32, 18, 55, 20), 1):
        catalog_sheet.column_dimensions[get_column_letter(col)].width = width

    summary = workbook.create_sheet("API Link Summary") if "API Link Summary" not in workbook.sheetnames else workbook["API Link Summary"]
    if summary.max_row > 1:
        summary.delete_rows(1, summary.max_row)
    summary["A1"] = "API Link Summary"
    summary["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary.append([])
    summary.append(["Metric", "Value"])
    for cell in summary[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    total = len(mapped_rows)
    primary = sum(count for source, count in source_counts.items() if source.endswith("API Design") and source != "Common API Design (Fallback)")
    fallback = source_counts["Common API Design (Fallback)"]
    unmapped = source_counts["Not Mapped"]
    summary_rows = [
        ("Total test cases", f"=COUNTA('Test Cases'!A2:A{test_sheet.max_row})"),
        ("Test cases linked to one or more APIs", f"=COUNTIF('Test Cases'!{get_column_letter(start_col)}2:{get_column_letter(start_col)}{test_sheet.max_row},\"<>Not Mapped\")"),
        ("Test cases mapped from primary DDD API Design", primary),
        ("Test cases mapped using Common API Design fallback", fallback),
        ("Cases using both primary and fallback sources", 0),
        ("Unmapped test cases", unmapped),
        ("Overall API traceability percentage", f"=IFERROR(B5/B4,0)"),
        ("Mapping count by module", "; ".join(f"{key}: {value}" for key, value in module_counts.items())),
        ("Mapping count by API source priority", "; ".join(f"{key}: {value}" for key, value in source_counts.items())),
        ("Mapping count by confidence", "; ".join(f"{key}: {value}" for key, value in confidence_counts.items())),
        ("Top APIs by mapped test-case count", "; ".join(f"{item['api_name']}: {item['mapped_count']}" for item in sorted(catalog, key=lambda x: x['mapped_count'], reverse=True)[:10] if item["mapped_count"])),
    ]
    for row in summary_rows:
        summary.append(list(row))
    summary["B10"].number_format = "0.0%"
    summary.column_dimensions["A"].width = 62
    summary.column_dimensions["B"].width = 110
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    counts = {"primary": primary, "fallback": fallback, "mixed": 0, "unmapped": unmapped, "percentage": (primary + fallback) * 100 / total if total else 0}
    add_readme_section(workbook, counts)
    output = OUTPUT_BASE
    version = 2
    while output.exists():
        output = OUTPUT_BASE.with_name(f"{OUTPUT_BASE.stem}_v{version}{OUTPUT_BASE.suffix}")
        version += 1
    workbook.save(output)
    return output, total, counts, len(catalog), source_counts


if __name__ == "__main__":
    output, total, counts, catalog_count, source_counts = build_workbook()
    print(f"Output: {output}")
    print(f"Total test cases: {total}")
    print(f"Primary: {counts['primary']}; fallback: {counts['fallback']}; mixed: {counts['mixed']}; unmapped: {counts['unmapped']}")
    print(f"Mapping percentage: {counts['percentage']:.1f}%")
    print(f"Catalog APIs: {catalog_count}")
    print(f"Source counts: {dict(source_counts)}")
