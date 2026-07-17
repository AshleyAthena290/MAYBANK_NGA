#!/usr/bin/env python3
"""
Comprehensive API Specification Extraction for Test Case Generation
Extracts from: ECLIPSE Account Dashboard Credit Card DDD API Design v1
"""

import openpyxl
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

EXCEL_FILE = r"input\api\ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx"
OUTPUT_FILE = r"artifacts\CreditCard_API_Comprehensive_Analysis.md"

class CreditCardAPIExtractor:
    def __init__(self, excel_path: str):
        """Initialize the extractor"""
        self.wb = openpyxl.load_workbook(excel_path)
        self.apis = []
        
    @staticmethod
    def clean_value(val) -> Optional[str]:
        """Clean cell value"""
        if val is None:
            return None
        return str(val).strip()
    
    def get_cell_value(self, ws, row: int, col: int) -> Optional[str]:
        """Get cell value safely"""
        try:
            cell = ws.cell(row=row, column=col)
            return self.clean_value(cell.value) if cell.value else None
        except:
            return None
    
    def extract_index(self) -> List[Dict[str, Any]]:
        """Extract API index from API_Specs_Index sheet"""
        if 'API_Specs_Index' not in self.wb.sheetnames:
            return []
        
        ws = self.wb['API_Specs_Index']
        apis = []
        
        # Skip header row (row 1)
        for row_idx in range(2, ws.max_row + 1):
            no = self.get_cell_value(ws, row_idx, 1)
            sheet_name = self.get_cell_value(ws, row_idx, 2)
            api_name = self.get_cell_value(ws, row_idx, 3)
            method = self.get_cell_value(ws, row_idx, 4)
            microservice = self.get_cell_value(ws, row_idx, 5)
            endpoint = self.get_cell_value(ws, row_idx, 6)
            status = self.get_cell_value(ws, row_idx, 8)
            remarks = self.get_cell_value(ws, row_idx, 9)
            
            if api_name and sheet_name and sheet_name in self.wb.sheetnames:
                apis.append({
                    'number': no,
                    'sheet_name': sheet_name,
                    'api_name': api_name,
                    'http_method': method or 'GET',
                    'microservice': microservice,
                    'endpoint': endpoint,
                    'status': status,
                    'remarks': remarks
                })
        
        return apis
    
    def extract_section_details(self, ws, marker: str, start_row: int = 1) -> Dict[str, Any]:
        """Extract details for a specific section"""
        section_data = {
            'headers': [],
            'rows': []
        }
        
        # Find the section
        found = False
        header_row = None
        
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = self.get_cell_value(ws, row_idx, 1)
            
            if cell_value and marker.lower() in cell_value.lower():
                found = True
                header_row = row_idx + 1
                # Extract headers
                for col_idx in range(2, 12):
                    header = self.get_cell_value(ws, header_row, col_idx)
                    if header:
                        section_data['headers'].append(header)
                    else:
                        break
                break
        
        if not found or not header_row:
            return section_data
        
        # Extract rows until we hit another section or empty rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_a = self.get_cell_value(ws, row_idx, 1)
            
            # Stop at next section
            if cell_a and any(x.lower() in cell_a.lower() for x in 
                            ['request', 'response', 'error', 'validation', 
                             'business', 'sample', 'auth', 'security']):
                if marker.lower() not in cell_a.lower():
                    break
            
            # Check if row has data
            row_data = []
            has_data = False
            for col_idx in range(2, 12):
                val = self.get_cell_value(ws, row_idx, col_idx)
                row_data.append(val)
                if val:
                    has_data = True
            
            if has_data:
                section_data['rows'].append(row_data)
            elif row_data[0] is None and len(row_data) > 1 and any(row_data):
                # Empty first column but data in other columns
                section_data['rows'].append(row_data)
        
        return section_data
    
    def parse_table_section(self, section: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Parse extracted section into structured data"""
        if not section['headers'] or not section['rows']:
            return []
        
        parsed = []
        for row in section['rows']:
            item = {}
            for i, header in enumerate(section['headers']):
                if i < len(row):
                    value = row[i]
                    if value:
                        item[header.lower().replace(' ', '_')] = value
            if item:
                parsed.append(item)
        
        return parsed
    
    def extract_api_detail(self, sheet_name: str, api_info: Dict[str, Any]) -> Dict[str, Any]:
        """Extract all details from an API sheet"""
        if sheet_name not in self.wb.sheetnames:
            return {}
        
        ws = self.wb[sheet_name]
        api = {
            'sheet_name': sheet_name,
            'api_name': api_info.get('api_name'),
            'http_method': self.get_cell_value(ws, 2, 2) or api_info.get('http_method'),
            'url': self.get_cell_value(ws, 3, 2),
            'message_type': self.get_cell_value(ws, 4, 2),
            'microservice': api_info.get('microservice'),
            'endpoint': api_info.get('endpoint'),
            'request_headers': [],
            'path_parameters': [],
            'query_parameters': [],
            'request_parameters': [],
            'request_body': [],
            'response_codes': {},
            'response_structure': [],
            'business_rules': [],
            'error_scenarios': [],
            'validations': [],
            'authentication': [],
            'authorization': [],
            'samples': {
                'request': None,
                'response': None
            }
        }
        
        # Extract HTTP Headers
        headers_section = self.extract_section_details(ws, 'HTTP Header')
        api['request_headers'] = self.parse_table_section(headers_section)
        
        # Extract Path Variables
        path_section = self.extract_section_details(ws, 'Path Variable', 
                                                   start_row=headers_section['rows'].__len__() + 10 if headers_section['rows'] else 1)
        api['path_parameters'] = self.parse_table_section(path_section)
        
        # Extract Request Parameters
        param_section = self.extract_section_details(ws, 'Request Parameter')
        api['request_parameters'] = self.parse_table_section(param_section)
        
        # Extract Request Body
        body_section = self.extract_section_details(ws, 'HTTP Body')
        api['request_body'] = self.parse_table_section(body_section)
        
        # Extract Response Structure
        response_section = self.extract_section_details(ws, 'Response')
        api['response_structure'] = self.parse_table_section(response_section)
        
        # Extract error codes and HTTP status codes
        for row_idx in range(1, ws.max_row + 1):
            cell_value = self.get_cell_value(ws, row_idx, 1)
            
            # Response Codes from Response section
            if cell_value and 'response' in cell_value.lower():
                for check_idx in range(row_idx + 1, min(row_idx + 20, ws.max_row + 1)):
                    code_val = self.get_cell_value(ws, check_idx, 2)
                    desc_val = self.get_cell_value(ws, check_idx, 3)
                    if code_val and desc_val:
                        api['response_codes'][code_val] = desc_val
            
            # Error Scenarios
            if cell_value and 'error' in cell_value.lower() and 'scenario' in cell_value.lower():
                for check_idx in range(row_idx + 1, min(row_idx + 20, ws.max_row + 1)):
                    code_val = self.get_cell_value(ws, check_idx, 2)
                    desc_val = self.get_cell_value(ws, check_idx, 3)
                    if code_val and desc_val:
                        api['error_scenarios'].append({
                            'code': code_val,
                            'description': desc_val,
                            'http_status': self.get_cell_value(ws, check_idx, 4)
                        })
            
            # Business Rules
            if cell_value and 'business rule' in cell_value.lower():
                for check_idx in range(row_idx + 1, min(row_idx + 15, ws.max_row + 1)):
                    rule_val = self.get_cell_value(ws, check_idx, 2)
                    if rule_val:
                        api['business_rules'].append(rule_val)
            
            # Validation Rules
            if cell_value and 'validation' in cell_value.lower():
                for check_idx in range(row_idx + 1, min(row_idx + 15, ws.max_row + 1)):
                    val_val = self.get_cell_value(ws, check_idx, 2)
                    if val_val:
                        api['validations'].append(val_val)
            
            # Authentication Requirements
            if cell_value and 'authentication' in cell_value.lower():
                for check_idx in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                    auth_val = self.get_cell_value(ws, check_idx, 2)
                    if auth_val:
                        api['authentication'].append(auth_val)
            
            # Authorization Requirements
            if cell_value and 'authorization' in cell_value.lower():
                for check_idx in range(row_idx + 1, min(row_idx + 10, ws.max_row + 1)):
                    authz_val = self.get_cell_value(ws, check_idx, 2)
                    if authz_val:
                        api['authorization'].append(authz_val)
            
            # Request Sample
            if cell_value == 'Request Sample':
                api['samples']['request'] = self.get_cell_value(ws, row_idx, 2)
            
            # Response Sample
            if cell_value == 'Response Sample':
                api['samples']['response'] = self.get_cell_value(ws, row_idx, 2)
        
        return api
    
    def extract_all_apis(self) -> List[Dict[str, Any]]:
        """Extract all APIs from the workbook"""
        api_index = self.extract_index()
        print(f"Found {len(api_index)} APIs in index")
        
        for api_info in api_index:
            sheet_name = api_info['sheet_name']
            print(f"Extracting: {api_info['api_name']} from sheet '{sheet_name}'")
            
            details = self.extract_api_detail(sheet_name, api_info)
            if details and details.get('api_name'):
                api_info['details'] = details
                self.apis.append(api_info)
        
        return self.apis
    
    def generate_markdown_report(self, output_file: str):
        """Generate comprehensive markdown report"""
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"# ECLIPSE Account Dashboard - Credit Card API Specifications\n")
            f.write(f"## Comprehensive Analysis for Test Case Generation\n")
            f.write(f"\n**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"**Excel Source:** ECLIPSE_Account Dashboard_Credit_Card_DDD_API_Design_v1_Workshop.xlsx\n\n")
            
            # Summary
            f.write(f"## Executive Summary\n---\n\n")
            f.write(f"- **Total APIs Analyzed:** {len(self.apis)}\n\n")
            
            # Count by method
            methods = {}
            for api in self.apis:
                method = api.get('http_method', 'UNKNOWN')
                methods[method] = methods.get(method, 0) + 1
            
            f.write(f"- **API Distribution by Method:**\n")
            for method in sorted(methods.keys()):
                f.write(f"  - {method}: {methods[method]}\n")
            f.write(f"\n")
            
            # API List
            f.write(f"## 1. API Endpoints Summary\n---\n\n")
            f.write(f"| No. | API Name | Method | Endpoint | Microservice | Status |\n")
            f.write(f"|-----|----------|--------|----------|--------------|--------|\n")
            
            for idx, api in enumerate(self.apis, 1):
                endpoint = api.get('endpoint', '')[:40]
                if len(api.get('endpoint', '')) > 40:
                    endpoint += '...'
                f.write(f"| {idx} | {api.get('api_name', '')} | {api.get('http_method', '')} | {endpoint} | {api.get('microservice', '')} | {api.get('status', '')} |\n")
            
            f.write(f"\n")
            
            # Detailed Specifications
            f.write(f"## 2. Detailed API Specifications\n---\n\n")
            
            for idx, api in enumerate(self.apis, 1):
                details = api.get('details', {})
                
                f.write(f"### 2.{idx} {api.get('api_name', 'Unknown')}\n\n")
                
                # Basic Info
                f.write(f"#### Basic Information\n")
                f.write(f"- **HTTP Method:** `{api.get('http_method', 'N/A')}`\n")
                f.write(f"- **Endpoint URL:** `{details.get('url', 'N/A')}`\n")
                f.write(f"- **Endpoint Path:** `{api.get('endpoint', 'N/A')}`\n")
                f.write(f"- **Microservice:** {api.get('microservice', 'N/A')}\n")
                f.write(f"- **Message Type:** {details.get('message_type', 'JSON')}\n")
                if api.get('remarks'):
                    f.write(f"- **Remarks:** {api.get('remarks')}\n")
                f.write(f"\n")
                
                # Request Headers
                if details.get('request_headers'):
                    f.write(f"#### Request Headers\n")
                    f.write(f"| Field | Type | Length | Mandatory | Sample Value | Description |\n")
                    f.write(f"|-------|------|--------|-----------|--------------|-------------|\n")
                    for header in details['request_headers']:
                        f.write(f"| {header.get('name', '')} | {header.get('type', '')} | {header.get('length', '')} | {header.get('mandatory', '')} | {header.get('sample_value', '')} | {header.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Path Parameters
                if details.get('path_parameters'):
                    f.write(f"#### Path Parameters\n")
                    f.write(f"| Parameter | Type | Length | Mandatory | Sample Value | Description |\n")
                    f.write(f"|-----------|------|--------|-----------|--------------|-------------|\n")
                    for param in details['path_parameters']:
                        f.write(f"| {param.get('name', '')} | {param.get('type', '')} | {param.get('length', '')} | {param.get('mandatory', '')} | {param.get('sample_value', '')} | {param.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Query Parameters
                if details.get('query_parameters'):
                    f.write(f"#### Query Parameters\n")
                    f.write(f"| Parameter | Type | Mandatory | Sample Value | Description |\n")
                    f.write(f"|-----------|------|-----------|--------------|-------------|\n")
                    for param in details['query_parameters']:
                        f.write(f"| {param.get('name', '')} | {param.get('type', '')} | {param.get('mandatory', '')} | {param.get('sample_value', '')} | {param.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Request Body
                if details.get('request_body'):
                    f.write(f"#### Request Body\n")
                    f.write(f"| Field | Type | Length | Mandatory | Sample Value | Description |\n")
                    f.write(f"|-------|------|--------|-----------|--------------|-------------|\n")
                    for body_field in details['request_body']:
                        f.write(f"| {body_field.get('name', '')} | {body_field.get('type', '')} | {body_field.get('length', '')} | {body_field.get('mandatory', '')} | {body_field.get('sample_value', '')} | {body_field.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Response Codes
                if details.get('response_codes'):
                    f.write(f"#### HTTP Response Codes\n")
                    f.write(f"| Status Code | Description |\n")
                    f.write(f"|-------------|-------------|\n")
                    for code in sorted(details['response_codes'].keys()):
                        f.write(f"| {code} | {details['response_codes'][code]} |\n")
                    f.write(f"\n")
                
                # Response Structure
                if details.get('response_structure'):
                    f.write(f"#### Response Structure\n")
                    f.write(f"| Field | Type | Description |\n")
                    f.write(f"|-------|------|-------------|\n")
                    for resp_field in details['response_structure']:
                        f.write(f"| {resp_field.get('name', '')} | {resp_field.get('type', '')} | {resp_field.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Business Rules
                if details.get('business_rules'):
                    f.write(f"#### Business Rules\n")
                    for rule in details['business_rules']:
                        f.write(f"- {rule}\n")
                    f.write(f"\n")
                
                # Validations
                if details.get('validations'):
                    f.write(f"#### Field Validations\n")
                    for val in details['validations']:
                        f.write(f"- {val}\n")
                    f.write(f"\n")
                
                # Error Scenarios
                if details.get('error_scenarios'):
                    f.write(f"#### Error Scenarios\n")
                    f.write(f"| Error Code | HTTP Status | Description |\n")
                    f.write(f"|------------|-------------|-------------|\n")
                    for error in details['error_scenarios']:
                        f.write(f"| {error.get('code', '')} | {error.get('http_status', '')} | {error.get('description', '')} |\n")
                    f.write(f"\n")
                
                # Authentication
                if details.get('authentication'):
                    f.write(f"#### Authentication Requirements\n")
                    for auth in details['authentication']:
                        f.write(f"- {auth}\n")
                    f.write(f"\n")
                
                # Authorization
                if details.get('authorization'):
                    f.write(f"#### Authorization Requirements\n")
                    for authz in details['authorization']:
                        f.write(f"- {authz}\n")
                    f.write(f"\n")
                
                # Samples
                if details.get('samples', {}).get('request'):
                    f.write(f"#### Request Sample\n")
                    f.write(f"```json\n{details['samples']['request']}\n```\n\n")
                
                if details.get('samples', {}).get('response'):
                    f.write(f"#### Response Sample\n")
                    f.write(f"```json\n{details['samples']['response']}\n```\n\n")
                
                f.write(f"---\n\n")
            
            # Test Case Generation Guide
            f.write(f"## 3. Test Case Generation Framework\n---\n\n")
            f.write(self._generate_test_framework())
        
        return output_file
    
    def _generate_test_framework(self) -> str:
        """Generate test case framework"""
        content = []
        
        content.append("### 3.1 Positive Test Cases (Happy Path)\n")
        content.append("Expected: Valid requests succeed with HTTP 200/201\n")
        content.append("- Test each API with all required parameters\n")
        content.append("- Verify response structure matches specification\n")
        content.append("- Verify response data types\n")
        content.append("- Verify response codes are as documented\n")
        content.append("\n")
        
        content.append("### 3.2 Negative Test Cases\n")
        content.append("**Missing Parameters:**\n")
        content.append("- Remove each mandatory parameter → Expect HTTP 400/422\n")
        content.append("- Verify error message identifies missing field\n")
        content.append("\n")
        
        content.append("**Invalid Parameter Values:**\n")
        content.append("- Use wrong data type (string instead of number, etc.)\n")
        content.append("- Use invalid values (e.g., invalid status code)\n")
        content.append("- Use out-of-range values\n")
        content.append("- Expect HTTP 400/422 with descriptive error\n")
        content.append("\n")
        
        content.append("**Authentication/Authorization:**\n")
        content.append("- Missing Authorization header → HTTP 401\n")
        content.append("- Invalid/expired Bearer token → HTTP 401\n")
        content.append("- Invalid client-id → HTTP 401\n")
        content.append("- Cross-customer access attempt → HTTP 403\n")
        content.append("- Insufficient permissions → HTTP 403\n")
        content.append("\n")
        
        content.append("**Resource Not Found:**\n")
        content.append("- Non-existent card number → HTTP 404\n")
        content.append("- Non-existent account → HTTP 404\n")
        content.append("- Non-existent resource ID → HTTP 404\n")
        content.append("\n")
        
        content.append("### 3.3 Boundary Value Testing\n")
        content.append("For numeric fields:\n")
        content.append("- Minimum value (0, -1, minimum allowed)\n")
        content.append("- Maximum value (max allowed)\n")
        content.append("- Just beyond boundaries\n")
        content.append("\n")
        
        content.append("For string fields:\n")
        content.append("- Empty string ''\n")
        content.append("- Single character\n")
        content.append("- Maximum length\n")
        content.append("- Beyond maximum length (if applicable)\n")
        content.append("- Null value\n")
        content.append("\n")
        
        content.append("For date fields:\n")
        content.append("- Minimum date\n")
        content.append("- Maximum date\n")
        content.append("- Invalid dates (e.g., Feb 30)\n")
        content.append("- Past/future dates (as applicable)\n")
        content.append("\n")
        
        content.append("### 3.4 Field Validation Testing\n")
        content.append("For each documented field validation rule:\n")
        content.append("- Test data that satisfies the rule\n")
        content.append("- Test data that violates the rule → Expect error\n")
        content.append("- Test edge cases\n")
        content.append("- Verify error message identifies violated field\n")
        content.append("\n")
        
        content.append("### 3.5 Business Logic Testing\n")
        content.append("For each documented business rule:\n")
        content.append("- Test scenarios where rule allows operation\n")
        content.append("- Test scenarios where rule blocks operation\n")
        content.append("- Test rule combinations\n")
        content.append("- Verify business error codes\n")
        content.append("\n")
        
        content.append("### 3.6 Security Testing\n")
        content.append("SQL Injection:\n")
        content.append("- Test with SQL injection payloads in string fields\n")
        content.append("- Example: `'; DROP TABLE--`, `1' OR '1'='1`\n")
        content.append("\n")
        
        content.append("XSS Injection:\n")
        content.append("- Test with XSS payloads in string fields\n")
        content.append("- Example: `<script>alert('XSS')</script>`, `<img src=x onerror=alert(1)>`\n")
        content.append("\n")
        
        content.append("Cross-Customer Access:\n")
        content.append("- Attempt to access another customer's data\n")
        content.append("- Attempt to modify another customer's data\n")
        content.append("- Expect HTTP 403 or 404\n")
        content.append("\n")
        
        content.append("### 3.7 Performance Testing\n")
        content.append("- Measure response time for typical requests\n")
        content.append("- Test with large payloads\n")
        content.append("- Test batch operations with multiple records\n")
        content.append("- Verify response time is within SLA\n")
        content.append("\n")
        
        content.append("### 3.8 Integration Testing\n")
        content.append("- Verify data consistency across dependent APIs\n")
        content.append("- Verify state propagation after operations\n")
        content.append("- Test cascading operations\n")
        content.append("- Test rollback/error scenarios\n")
        content.append("\n")
        
        content.append("### 3.9 Error Code Mapping\n")
        content.append("For each documented error scenario:\n")
        content.append("- Trigger the error condition\n")
        content.append("- Verify correct error code is returned\n")
        content.append("- Verify correct HTTP status code\n")
        content.append("- Verify error message is present and descriptive\n")
        content.append("- Verify error response structure matches specification\n")
        content.append("\n")
        
        content.append("### 3.10 Test Data Requirements\n")
        content.append("- Valid customer IDs (numeric, non-zero)\n")
        content.append("- Valid credit card numbers (real or test format)\n")
        content.append("- Valid account numbers\n")
        content.append("- Valid currency codes (IDR, USD, SGD, MYR, EUR)\n")
        content.append("- Valid status values (ACTIVE, BLOCKED, INACTIVE, etc.)\n")
        content.append("- Valid date formats (YYYY-MM-DD or per spec)\n")
        content.append("- Valid Bearer tokens (test/mock tokens)\n")
        content.append("- Valid client IDs\n")
        content.append("- Test environment values (UAT, DEV, PROD)\n")
        
        return ''.join(content)


def main():
    """Main execution"""
    try:
        print("=" * 80)
        print("ECLIPSE Account Dashboard - Credit Card API Analysis")
        print("=" * 80)
        print(f"Source: {EXCEL_FILE}")
        print(f"Output: {OUTPUT_FILE}")
        print("")
        
        # Initialize extractor
        extractor = CreditCardAPIExtractor(EXCEL_FILE)
        
        # Extract all APIs
        print("Extracting API specifications...")
        apis = extractor.extract_all_apis()
        print(f"✓ Extracted {len(apis)} APIs")
        print("")
        
        # Generate report
        print("Generating comprehensive report...")
        report_path = extractor.generate_markdown_report(OUTPUT_FILE)
        
        print("")
        print("=" * 80)
        print("ANALYSIS COMPLETE")
        print("=" * 80)
        print(f"✓ Total APIs Analyzed: {len(apis)}")
        print(f"✓ Report Generated: {report_path}")
        print("")
        
    except Exception as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
