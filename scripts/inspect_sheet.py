import openpyxl

file_path = r"input\api\ECLIPSE_Account Dashboard_Casa_DDD_API_Design_v1_Workshop.xlsx"
wb = openpyxl.load_workbook(file_path)

# Check the setAutoSweep sheet
ws = wb['setAutoSweep']

print("setAutoSweep Sheet - First 40 rows:\n")
print("=" * 200)

for row_idx in range(1, 41):
    cells = []
    for col_idx in range(1, 12):  # First 11 columns
        cell = ws.cell(row=row_idx, column=col_idx)
        value = str(cell.value)[:25] if cell.value is not None else "-"
        cells.append(value.ljust(25))
    print(f"Row {row_idx:2d}: {' | '.join(cells)}")

print("\n" + "=" * 200)
print("\nLet's also check one of the problematic sheets:")
ws2 = wb['1.2 Account Listing']
print(f"\n1.2 Account Listing Sheet - First 30 rows:\n")
for row_idx in range(1, 31):
    cells = []
    for col_idx in range(1, 6):
        cell = ws2.cell(row=row_idx, column=col_idx)
        value = str(cell.value)[:30] if cell.value is not None else "-"
        cells.append(value.ljust(30))
    print(f"Row {row_idx:2d}: {' | '.join(cells)}")

