#!/usr/bin/env python3
"""
Convert P&T_Local_Transfer_API_Test_Cases_Detailed.md to Excel format
Better parser for markdown tables
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path

# Configuration
md_file = 'artifacts/Test_Case/P&T_Local_Transfer_API_Test_Cases_Detailed.md'
excel_file = 'artifacts/Test_Case/P&T_Local_Transfer_API_Test_Cases.xlsx'

# Read markdown file
print(f"Reading {md_file}...")
with open(md_file, 'r', encoding='utf-8') as f:
    md_content = f.read()

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
positive_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
negative_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
high_priority_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
high_priority_font = Font(bold=True, color="FFFFFF")
medium_priority_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
medium_priority_font = Font(bold=True, color="FFFFFF")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Parse markdown and extract test cases
print("Parsing markdown content...")
test_cases = []

# Split by API sections
api_pattern = r'### API (\d+): ([^\n]+)'
api_matches = re.finditer(api_pattern, md_content)
api_list = list(api_matches)

for api_idx, api_match in enumerate(api_list):
    api_num = api_match.group(1)
    api_name = api_match.group(2).strip()
    
    # Get content between this API and next API (or end of file)
    start_pos = api_match.end()
    if api_idx + 1 < len(api_list):
        end_pos = api_list[api_idx + 1].start()
    else:
        end_pos = len(md_content)
    
    api_section = md_content[start_pos:end_pos]
    
    # Extract metadata
    method_match = re.search(r'\*\*Method:\*\* `([^`]+)`', api_section)
    endpoint_match = re.search(r'\*\*Endpoint:\*\* `([^`]+)`', api_section)
    url_match = re.search(r'\*\*URL:\*\* `([^`]+)`', api_section)
    description_match = re.search(r'\*\*Description:\*\* ([^\n]+)', api_section)
    
    method = method_match.group(1) if method_match else ''
    endpoint = endpoint_match.group(1) if endpoint_match else ''
    url = url_match.group(1) if url_match else ''
    description = description_match.group(1).strip() if description_match else ''
    
    # Find test cases table
    # Pattern: | TC_XX_XXX | ... | ... | ... | ... | ... | ... |
    test_case_lines = re.findall(r'\| TC_\d+_[PN]\d+ \| .+ \| .+ \| .+ \| .+ \| .+ \| .+ \|', api_section)
    
    for line in test_case_lines:
        # Split by | and clean up
        cells = [cell.strip() for cell in line.split('|')[1:-1]]  # Remove empty first/last
        
        if len(cells) >= 7:
            tc = {
                'api_num': api_num,
                'api_name': api_name,
                'method': method,
                'endpoint': endpoint,
                'url': url,
                'description': description,
                'tc_id': cells[0],
                'title': cells[1],
                'type': cells[2],  # Positive or Negative
                'category': cells[3],
                'preconditions': cells[4],
                'expected_result': cells[5],
                'priority': cells[6] if len(cells) > 6 else 'Medium'
            }
            test_cases.append(tc)

print(f"✓ Extracted {len(test_cases)} test cases")

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Add title
ws.merge_cells('A1:H1')
title_cell = ws['A1']
title_cell.value = 'P&T Local Transfer API - Test Cases'
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = header_fill
title_cell.alignment = center_align
ws.row_dimensions[1].height = 25

# Add metadata row
ws.merge_cells('A2:H2')
metadata_cell = ws['A2']
metadata_cell.value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Total Test Cases: {len(test_cases)}'
metadata_cell.font = Font(size=10, italic=True)
metadata_cell.alignment = left_align
ws.row_dimensions[2].height = 20

# Add headers
row = 4
headers = ['TC ID', 'API Name', 'Test Title', 'Test Type', 'Category', 'Pre-conditions', 'Expected Result', 'Priority']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

ws.row_dimensions[row].height = 25

# Add test cases
row = 5
for tc in test_cases:
    ws.cell(row=row, column=1, value=tc['tc_id'])
    ws.cell(row=row, column=2, value=tc['api_name'])
    ws.cell(row=row, column=3, value=tc['title'])
    ws.cell(row=row, column=4, value=tc['type'])
    ws.cell(row=row, column=5, value=tc['category'])
    ws.cell(row=row, column=6, value=tc['preconditions'])
    ws.cell(row=row, column=7, value=tc['expected_result'])
    ws.cell(row=row, column=8, value=tc['priority'])
    
    # Apply row formatting
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = left_align
        
        # Color coding for test type
        if tc['type'] == 'Positive':
            cell.fill = positive_fill
        elif tc['type'] == 'Negative':
            cell.fill = negative_fill
        
        # Priority highlighting
        if col == 8:  # Priority column
            if tc['priority'] == 'High':
                cell.fill = high_priority_fill
                cell.font = high_priority_font
            elif tc['priority'] == 'Medium':
                cell.fill = medium_priority_fill
                cell.font = medium_priority_font
            cell.alignment = center_align
        elif col == 4:  # Type column
            cell.alignment = center_align
    
    ws.row_dimensions[row].height = 40
    row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 35
ws.column_dimensions['H'].width = 12

# Create Summary sheet
print("Creating summary sheet...")
summary_ws = wb.create_sheet('Summary')

# Summary title
summary_ws.merge_cells('A1:D1')
summary_title = summary_ws['A1']
summary_title.value = 'Test Execution Summary'
summary_title.font = Font(bold=True, size=14, color="FFFFFF")
summary_title.fill = header_fill
summary_title.alignment = center_align
summary_ws.row_dimensions[1].height = 25

# Summary statistics
row = 3
positive_count = len([tc for tc in test_cases if tc['type'] == 'Positive'])
negative_count = len([tc for tc in test_cases if tc['type'] == 'Negative'])
high_count = len([tc for tc in test_cases if tc['priority'] == 'High'])
medium_count = len([tc for tc in test_cases if tc['priority'] == 'Medium'])
unique_apis = len(set(tc['api_name'] for tc in test_cases))

summary_data = [
    ('Metric', 'Count'),
    ('Total APIs', unique_apis),
    ('Total Test Cases', len(test_cases)),
    ('Positive Tests', positive_count),
    ('Negative Tests', negative_count),
    ('High Priority', high_count),
    ('Medium Priority', medium_count),
]

for label, value in summary_data:
    label_cell = summary_ws.cell(row=row, column=1, value=label)
    value_cell = summary_ws.cell(row=row, column=2, value=value)
    
    if row == 3:  # Header row
        label_cell.fill = header_fill
        label_cell.font = header_font
        value_cell.fill = header_fill
        value_cell.font = header_font
    
    label_cell.border = border
    value_cell.border = border
    label_cell.font = Font(bold=True)
    value_cell.font = Font(bold=True, size=12)
    value_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1

# APIs by count
summary_ws.merge_cells('A10:D10')
api_title = summary_ws['A10']
api_title.value = 'Test Cases by API'
api_title.font = Font(bold=True, size=12, color="FFFFFF")
api_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
api_title.alignment = center_align

row = 11
api_counts = {}
for tc in test_cases:
    api_counts[tc['api_name']] = api_counts.get(tc['api_name'], 0) + 1

# API header
api_name_header = summary_ws.cell(row=row, column=1, value='API Name')
api_count_header = summary_ws.cell(row=row, column=2, value='Test Cases')
api_name_header.fill = header_fill
api_count_header.fill = header_fill
api_name_header.font = header_font
api_count_header.font = header_font
api_name_header.border = border
api_count_header.border = border
api_name_header.alignment = center_align
api_count_header.alignment = center_align

row += 1
for api_name, count in sorted(api_counts.items()):
    api_cell = summary_ws.cell(row=row, column=1, value=api_name)
    count_cell = summary_ws.cell(row=row, column=2, value=count)
    
    api_cell.border = border
    count_cell.border = border
    count_cell.alignment = center_align
    count_cell.font = Font(bold=True)
    
    row += 1

summary_ws.column_dimensions['A'].width = 40
summary_ws.column_dimensions['B'].width = 15
summary_ws.column_dimensions['C'].width = 15

# Create Category Analysis sheet
print("Creating category analysis sheet...")
category_ws = wb.create_sheet('Category Analysis')

# Title
category_ws.merge_cells('A1:C1')
category_title = category_ws['A1']
category_title.value = 'Test Cases by Category'
category_title.font = Font(bold=True, size=14, color="FFFFFF")
category_title.fill = header_fill
category_title.alignment = center_align
category_ws.row_dimensions[1].height = 25

# Calculate categories
category_counts = {}
for tc in test_cases:
    cat = tc['category']
    category_counts[cat] = category_counts.get(cat, 0) + 1

row = 3
headers = ['Category', 'Count', 'Percentage']
for col, header in enumerate(headers, 1):
    cell = category_ws.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

category_ws.row_dimensions[row].height = 20

row = 4
total = len(test_cases)
for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
    category_ws.cell(row=row, column=1, value=category)
    category_ws.cell(row=row, column=2, value=count)
    percentage = (count / total) * 100
    category_ws.cell(row=row, column=3, value=f'{percentage:.1f}%')
    
    for col in range(1, 4):
        cell = category_ws.cell(row=row, column=col)
        cell.border = border
        if col > 1:
            cell.alignment = center_align
            cell.font = Font(bold=True)
    
    row += 1

category_ws.column_dimensions['A'].width = 25
category_ws.column_dimensions['B'].width = 12
category_ws.column_dimensions['C'].width = 15

# Save workbook
wb.save(excel_file)
print(f"✓ Excel file created successfully!")
print(f"\nFile: {excel_file}")
print(f"File Size: {Path(excel_file).stat().st_size / 1024:.1f} KB")

print(f"\n{'='*60}")
print(f"Workbook Structure:")
print(f"{'='*60}")
print(f"Sheet 1: Test Cases")
print(f"  - Total Rows: {len(test_cases)}")
print(f"  - Positive Tests: {positive_count}")
print(f"  - Negative Tests: {negative_count}")
print(f"\nSheet 2: Summary")
print(f"  - Statistics: Total APIs, Test Cases, Priority Breakdown")
print(f"  - API Coverage: {unique_apis} APIs covered")
print(f"\nSheet 3: Category Analysis")
print(f"  - Test Distribution by Category")
print(f"  - {len(category_counts)} different categories")
print(f"\nFormatting Applied:")
print(f"  ✓ Color-coded test types (Positive: Green, Negative: Orange)")
print(f"  ✓ Priority highlighting (High: Red, Medium: Yellow)")
print(f"  ✓ Auto-adjusted column widths")
print(f"  ✓ Wrapped text for better readability")
print(f"\n✓ Ready for use!")

