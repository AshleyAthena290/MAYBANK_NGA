#!/usr/bin/env python3
"""Generate and format API validation test cases as Markdown and Excel"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Load test case metadata
with open('test_cases_metadata.json', 'r') as f:
    data = json.load(f)

apis = data['apis']
test_cases = data['test_cases']

print("=" * 100)
print("GENERATING MARKDOWN AND EXCEL OUTPUTS")
print("=" * 100)

# ===== MARKDOWN OUTPUT =====
markdown_output = """# API Validation Test Cases for DEP_App_Dashboard_DDD_API_Design_v1

**Generated Date:** """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """

---

## 1. TEST SUMMARY

### Overview
- **Total Number of APIs Covered:** """ + str(len(apis)) + """
- **Total Number of Test Cases:** """ + str(len(test_cases)) + """
  - **Positive Test Cases:** """ + str(len([tc for tc in test_cases if tc['type'] == 'Positive'])) + """
  - **Negative Test Cases:** """ + str(len([tc for tc in test_cases if tc['type'] == 'Negative'])) + """

### APIs Identified
| # | Service Name | HTTP Method | Endpoint | Module |
|---|---|---|---|---|
"""

for i, api in enumerate(apis, 1):
    markdown_output += f"| {i} | {api['service']} | {api['method']} | {api['url']} | Channel Layer |\n"

markdown_output += """
### Key Validation Areas Covered
- **Field Validation:** Testing mandatory/optional fields, data types, field lengths
- **Authentication & Authorization:** Missing tokens, invalid tokens, expired tokens, unauthorized access
- **Business Rule Validation:** Request/response format, API response structure, error handling
- **Data Type Checks:** String, Boolean, Integer, Object, Array types as per spec
- **Boundary Conditions:** Min/max length, min/max numeric ranges
- **Error Handling:** Malformed JSON, invalid formats, missing mandatory fields
- **Security:** SQL Injection, XSS payloads, special character handling
- **HTTP Protocol:** Correct HTTP method usage, status code validation

### Assumptions Made Where DDD is Ambiguous
- **Response Error Messages:** Test cases assume standard HTTP status codes (400, 401, 405, etc.) and that error responses include descriptive messages
- **Rate Limiting:** DDD does not explicitly define rate limiting; test cases can be extended if rate limiting is implemented
- **Token Expiry:** Assumed token expiry time is standard; specific timeout values not explicitly defined in DDD
- **Data Sanitization:** Assumed API implements input sanitization for security; test cases verify safe handling
- **Backward Compatibility:** Assumed API maintains compatibility with optional fields; removing optional fields should not cause errors

---

## 2. DETAILED TEST CASES

### Test Case Format
- **Test Case ID:** Unique identifier (TC_API##_###)
- **Service:** The API being tested
- **Endpoint:** Full API endpoint URL
- **HTTP Method:** GET, POST, PUT, DELETE, etc.
- **Test Type:** Positive (valid inputs) or Negative (invalid inputs)
- **Test Category:** Type of validation being tested
- **Priority:** High/Medium/Low based on criticality

---

"""

# Group test cases by API
test_cases_by_api = {}
for tc in test_cases:
    api_key = tc['service']
    if api_key not in test_cases_by_api:
        test_cases_by_api[api_key] = []
    test_cases_by_api[api_key].append(tc)

# Generate markdown for each API
for api in apis:
    api_name = api['service']
    api_tcs = test_cases_by_api.get(api_name, [])
    
    markdown_output += f"""### {api_name}
**Endpoint:** `{api['method']} {api['url']}`  
**Description:** {api['description']}

"""
    
    # Summary table for this API
    markdown_output += f"""#### {api_name} - Test Case Summary
| ID | Title | Type | Category | Priority |
|---|---|---|---|---|
"""
    
    for tc in api_tcs:
        markdown_output += f"| {tc['id']} | {tc['title']} | {tc['type']} | {tc['category']} | {tc['priority']} |\n"
    
    markdown_output += "\n"
    
    # Detailed test cases
    markdown_output += f"#### {api_name} - Detailed Test Cases\n\n"
    
    for tc in api_tcs:
        markdown_output += f"""**Test Case ID:** `{tc['id']}`  
**Title:** {tc['title']}  
**Type:** {tc['type']}  
**Category:** {tc['category']}  
**Priority:** {tc['priority']}  
**Endpoint:** `{tc['method']} {tc['api_endpoint']}`  

**Pre-conditions:**  
{tc['preconditions']}

**Test Steps:**  
"""
        for step in (tc['steps'] if isinstance(tc['steps'], list) else [tc['steps']]):
            markdown_output += f"{step}  \n"
        
        markdown_output += f"""
**Input Data / Request Payload:**  
```
{tc['input_data']}
```

**Expected Result:**  
{tc['expected_result']}

---

"""

# Save markdown
with open('DEP_App_Dashboard_API_Test_Cases.md', 'w', encoding='utf-8') as f:
    f.write(markdown_output)

print("✓ Markdown file generated: DEP_App_Dashboard_API_Test_Cases.md")

# ===== EXCEL OUTPUT =====
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Create Summary sheet
ws_summary = wb.create_sheet("Summary", 0)
ws_summary['A1'] = "API VALIDATION TEST CASES - SUMMARY"
ws_summary['A1'].font = Font(bold=True, size=14)

ws_summary['A3'] = "Report Generated Date:"
ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

ws_summary['A5'] = "Total Number of APIs:"
ws_summary['B5'] = len(apis)
ws_summary['B5'].font = Font(bold=True)

ws_summary['A6'] = "Total Number of Test Cases:"
ws_summary['B6'] = len(test_cases)
ws_summary['B6'].font = Font(bold=True)

ws_summary['A7'] = "Positive Test Cases:"
ws_summary['B7'] = len([tc for tc in test_cases if tc['type'] == 'Positive'])

ws_summary['A8'] = "Negative Test Cases:"
ws_summary['B8'] = len([tc for tc in test_cases if tc['type'] == 'Negative'])

ws_summary['A10'] = "APIs Covered"
ws_summary['A10'].font = Font(bold=True, size=12)

ws_summary['A11'] = "No"
ws_summary['B11'] = "Service Name"
ws_summary['C11'] = "HTTP Method"
ws_summary['D11'] = "Endpoint"
ws_summary['E11'] = "Module"

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_summary[f'{col}11'].fill = header_fill
    ws_summary[f'{col}11'].font = header_font

for i, api in enumerate(apis, 1):
    ws_summary[f'A{11+i}'] = i
    ws_summary[f'B{11+i}'] = api['service']
    ws_summary[f'C{11+i}'] = api['method']
    ws_summary[f'D{11+i}'] = api['url']
    ws_summary[f'E{11+i}'] = "Channel Layer"

# Set column widths
ws_summary.column_dimensions['A'].width = 5
ws_summary.column_dimensions['B'].width = 30
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 40
ws_summary.column_dimensions['E'].width = 20

# Create Key Validation Areas sheet
ws_validation = wb.create_sheet("Validation Areas")
ws_validation['A1'] = "KEY VALIDATION AREAS COVERED"
ws_validation['A1'].font = Font(bold=True, size=12)

validation_areas = [
    "Field Validation - Testing mandatory/optional fields, data types, field lengths",
    "Authentication & Authorization - Missing tokens, invalid tokens, expired tokens, unauthorized access",
    "Business Rule Validation - Request/response format, API response structure, error handling",
    "Data Type Checks - String, Boolean, Integer, Object, Array types as per specification",
    "Boundary Conditions - Min/max length, min/max numeric ranges",
    "Error Handling - Malformed JSON, invalid formats, missing mandatory fields",
    "Security - SQL Injection, XSS payloads, special character handling",
    "HTTP Protocol - Correct HTTP method usage, status code validation"
]

for i, area in enumerate(validation_areas, 1):
    ws_validation[f'A{i+2}'] = area
    ws_validation.row_dimensions[i+2].height = 30
    ws_validation[f'A{i+2}'].alignment = Alignment(wrap_text=True, vertical='top')

ws_validation.column_dimensions['A'].width = 80

# Create Test Cases sheets (grouped by API)
for api in apis:
    api_name = api['service']
    ws_api = wb.create_sheet(api_name[:31])  # Excel sheet name limit is 31 characters
    
    # Headers
    headers = ['Test Case ID', 'Test Case Title', 'Test Type', 'Test Category', 'HTTP Method', 
               'Endpoint', 'Pre-conditions', 'Test Steps', 'Input Data', 'Expected Result', 'Priority']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_api.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add test cases for this API
    api_tcs = test_cases_by_api.get(api_name, [])
    
    for row_idx, tc in enumerate(api_tcs, 2):
        ws_api.cell(row=row_idx, column=1).value = tc['id']
        ws_api.cell(row=row_idx, column=2).value = tc['title']
        ws_api.cell(row=row_idx, column=3).value = tc['type']
        ws_api.cell(row=row_idx, column=4).value = tc['category']
        ws_api.cell(row=row_idx, column=5).value = tc['method']
        ws_api.cell(row=row_idx, column=6).value = tc['api_endpoint']
        ws_api.cell(row=row_idx, column=7).value = tc['preconditions']
        
        # Format steps
        steps_text = '\n'.join(tc['steps']) if isinstance(tc['steps'], list) else tc['steps']
        ws_api.cell(row=row_idx, column=8).value = steps_text
        
        ws_api.cell(row=row_idx, column=9).value = tc['input_data']
        ws_api.cell(row=row_idx, column=10).value = tc['expected_result']
        ws_api.cell(row=row_idx, column=11).value = tc['priority']
        
        # Apply styling
        for col_idx in range(1, 12):
            cell = ws_api.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Set column widths
    ws_api.column_dimensions['A'].width = 12
    ws_api.column_dimensions['B'].width = 40
    ws_api.column_dimensions['C'].width = 12
    ws_api.column_dimensions['D'].width = 18
    ws_api.column_dimensions['E'].width = 10
    ws_api.column_dimensions['F'].width = 30
    ws_api.column_dimensions['G'].width = 25
    ws_api.column_dimensions['H'].width = 30
    ws_api.column_dimensions['I'].width = 35
    ws_api.column_dimensions['J'].width = 40
    ws_api.column_dimensions['K'].width = 10
    
    # Freeze header row
    ws_api.freeze_panes = 'A2'

# Create Assumptions sheet
ws_assumptions = wb.create_sheet("Assumptions")
ws_assumptions['A1'] = "ASSUMPTIONS MADE WHERE DDD IS AMBIGUOUS OR INCOMPLETE"
ws_assumptions['A1'].font = Font(bold=True, size=12)

assumptions = [
    ("Response Error Messages", "Test cases assume standard HTTP status codes (400, 401, 405, etc.) and that error responses include descriptive messages with error codes and keys for i18n support."),
    ("Rate Limiting & Throttling", "DDD does not explicitly define rate limiting or throttling behavior. Test cases can be extended if rate limiting (e.g., X-RateLimit-Limit header) is implemented in future versions."),
    ("Token Expiry & Refresh", "Assumed standard JWT token expiry handling. Specific timeout values and token refresh mechanisms are not explicitly defined in DDD; test cases use generic token validation."),
    ("Data Sanitization & Input Validation", "Assumed API implements input sanitization for security. Test cases verify safe handling of special characters, SQL injection payloads, and XSS attempts."),
    ("Backward Compatibility", "Assumed API maintains compatibility with optional fields. Removing optional fields from requests should not cause errors; API should process requests successfully."),
    ("Content Negotiation", "Assumed API supports application/json content type. Other formats (XML, CSV) are not tested as DDD specifies JSON only."),
    ("Request Timeout", "No explicit timeout duration specified in DDD. Standard HTTP timeout (e.g., 30 seconds) is assumed for test case validation."),
    ("Caching Behavior", "getDashboardState mentions caching with 5/10 min TTL for pre-login snapshots. Test cases do not explicitly validate cache behavior but can be extended for cache testing."),
    ("Device Binding", "X-Device-Bind header is optional. Test cases assume API handles both bound and unbound devices gracefully."),
    ("Internationalization (i18n)", "Response includes message object with i18n keys. Test cases assume proper key mapping for different locales based on X-APP-LOCALE header.")
]

row = 3
for assumption_title, assumption_desc in assumptions:
    ws_assumptions[f'A{row}'] = assumption_title
    ws_assumptions[f'A{row}'].font = Font(bold=True)
    ws_assumptions[f'B{row}'] = assumption_desc
    ws_assumptions.row_dimensions[row].height = 40
    ws_assumptions[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 2

ws_assumptions.column_dimensions['A'].width = 30
ws_assumptions.column_dimensions['B'].width = 80

# Save Excel file
excel_filename = 'DEP_App_Dashboard_API_Test_Cases.xlsx'
wb.save(excel_filename)
print(f"✓ Excel file generated: {excel_filename}")

print("\n" + "=" * 100)
print("TEST CASE GENERATION COMPLETE")
print("=" * 100)
print(f"\nGenerated Files:")
print(f"  1. DEP_App_Dashboard_API_Test_Cases.md (Markdown format)")
print(f"  2. DEP_App_Dashboard_API_Test_Cases.xlsx (Excel format with multiple sheets)")
print(f"  3. test_cases_metadata.json (Raw test case data)")
print(f"\nTest Case Statistics:")
print(f"  Total Test Cases: {len(test_cases)}")
print(f"  Positive Test Cases: {len([tc for tc in test_cases if tc['type'] == 'Positive'])}")
print(f"  Negative Test Cases: {len([tc for tc in test_cases if tc['type'] == 'Negative'])}")
print(f"  APIs Covered: {len(apis)}")
