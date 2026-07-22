#!/usr/bin/env python3
"""Extract structure from DEP_App_Dashboard_DDD_API_Design_v1.xlsx"""

import openpyxl
import json

# Load the DDD file
ddd_file = r'input\api\DEP_App_Dashboard_DDD_API_Design_v1.xlsx'
wb = openpyxl.load_workbook(ddd_file)

# Print INDEX sheet to understand API list
print("=" * 100)
print("API INDEX")
print("=" * 100)
ws_index = wb['INDEX']
for row in ws_index.iter_rows(values_only=True):
    print(row)

print("\n" + "=" * 100)
print("FIRST API SHEET STRUCTURE: Get_Onboarding_Accounts")
print("=" * 100)
ws = wb['Get_Onboarding_Accounts']
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 30:
        print(f"Row {i}: {row}")
    else:
        break

# Check another API sheet
print("\n" + "=" * 100)
print("SECOND API SHEET STRUCTURE: getDashboardState")
print("=" * 100)
ws = wb['getDashboardState']
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 30:
        print(f"Row {i}: {row}")
    else:
        break
