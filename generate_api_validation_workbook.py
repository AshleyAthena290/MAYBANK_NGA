#!/usr/bin/env python3
"""Generate a formatted API validation workbook from generated BDD YAML files."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Test Case ID", "Test Case Title", "Test Type", "Test Category",
    "HTTP Method", "Endpoint", "Pre-conditions", "Test Steps",
    "Input Data", "Expected Result", "Priority",
]
API_SHEET_COLUMNS = [8, 34, 18, 24, 12, 52, 38, 48, 58, 68, 12]
HEADER_FILL = PatternFill("solid", fgColor="17365D")
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
PRIORITY_FILLS = {
    "High": PatternFill("solid", fgColor="F4CCCC"),
    "Medium": PatternFill("solid", fgColor="FFF2CC"),
    "Low": PatternFill("solid", fgColor="D9EAD3"),
}


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, indent=2, ensure_ascii=True, default=str)
    return str(value)


def yaml_block(value: Any) -> str:
    if value in (None, {}, [], ""):
        return "{}"
    return yaml.safe_dump(value, sort_keys=False, allow_unicode=False, default_flow_style=False).strip()


def priority(value: Any) -> str:
    return {"P1": "High", "P2": "Medium", "P3": "Low"}.get(str(value).upper(), "Medium")


def test_type(tags: list[str], title: str) -> str:
    values = {str(tag).lower() for tag in tags}
    if "edge-case" in values or "resilience" in values or any(x in title.lower() for x in ("timeout", "downstream")):
        return "Edge/Resilience"
    if "boundary" in values:
        return "Boundary"
    if "negative" in values:
        return "Negative"
    return "Positive"


def category(tags: list[str], title: str) -> str:
    values = {str(tag).lower() for tag in tags}
    title_lower = title.lower()
    if values & {"authentication", "security"}:
        return "Authentication" if "authentication" in values else "Security"
    if "authorization" in values or "role" in title_lower:
        return "Authorization"
    if "rate" in title_lower or "thrott" in title_lower:
        return "Rate Limiting"
    if "content-type" in title_lower or "http" in values:
        return "HTTP Protocol"
    if values & {"boundary", "validation"}:
        return "Boundary Testing" if "boundary" in values else "Field Validation"
    if "duplicate" in title_lower or "timeout" in title_lower or "empty" in title_lower:
        return "Error Handling"
    return "Business Rule" if "business" in values else "API Contract"


def steps_for(case: dict[str, Any]) -> str:
    request = case.get("request") or {}
    return "\n".join([
        "1. Prepare the pre-conditions and test data described for this scenario.",
        f"2. Send the {request.get('method', 'specified')} request to the endpoint with the listed headers, parameters, and body.",
        "3. Capture the HTTP status code and response body.",
        "4. Compare the response with every expected result and assertion.",
    ])


def expected_result(case: dict[str, Any]) -> str:
    response = case.get("response") or {}
    lines: list[str] = []
    if response.get("successStatusCode") is not None:
        lines.append(f"HTTP status: {response['successStatusCode']}")
    if response.get("successDescription"):
        lines.append(f"Response description: {response['successDescription']}")
    for assertion in case.get("assertions") or []:
        lines.append(f"Assertion: {as_text(assertion)}")
    for field in ("schema", "bodySchema", "expectedFields"):
        if response.get(field):
            lines.append(f"{field}:\n{yaml_block(response[field])}")
    if response.get("errorScenarios"):
        lines.append(f"Error scenarios:\n{yaml_block(response['errorScenarios'])}")
    if response.get("errorStatusCodes"):
        lines.append(f"Error status codes:\n{yaml_block(response['errorStatusCodes'])}")
    if not lines:
        lines.append("Assumption: The YAML does not define an expected status code or response body.")
    return "\n".join(lines)


def make_row(case: dict[str, Any], api_folder: str, source_yaml: Path) -> list[str]:
    request = case.get("request") or {}
    tags = case.get("tags") or []
    title = as_text(case.get("title") or case.get("id"))
    input_data = {
        "headers": request.get("headers") or {},
        "pathParams": request.get("pathParams") or {},
        "queryParams": request.get("queryParams") or {},
        "body": request.get("body") or {},
    }
    return [
        as_text(case.get("id")), title, test_type(tags, title), category(tags, title),
        as_text(request.get("method") or case.get("method")),
        as_text(request.get("url") or request.get("endpoint") or case.get("endpoint")),
        yaml_block(case.get("preconditions") or []), steps_for(case), yaml_block(input_data),
        expected_result(case), priority(case.get("priority")),
    ]


def load_cases(input_root: Path) -> tuple[dict[str, list[tuple[list[str], Path, dict[str, Any]]]], list[dict[str, str]]]:
    grouped: dict[str, list[tuple[list[str], Path, dict[str, Any]]]] = defaultdict(list)
    inventory: list[dict[str, str]] = []
    for api_dir in sorted(path for path in input_root.iterdir() if path.is_dir() and path.name.lower() != "test cases"):
        cases = []
        for yaml_path in sorted(api_dir.rglob("*.yaml")):
            if yaml_path.name == "_index.yaml":
                continue
            with yaml_path.open("r", encoding="utf-8") as stream:
                case = yaml.safe_load(stream) or {}
            if not isinstance(case, dict) or not case.get("id"):
                continue
            cases.append((make_row(case, api_dir.name, yaml_path), yaml_path, case))
        if cases:
            first = cases[0][2]
            request = first.get("request") or {}
            references = first.get("references") or {}
            inventory.append({
                "api": api_dir.name,
                "method": as_text(request.get("method") or first.get("method")),
                "endpoint": as_text(request.get("url") or request.get("endpoint") or first.get("endpoint")),
                "sourceWorksheet": as_text(references.get("sourceWorksheet")) or "Assumption: not present in YAML",
                "apiSpecFile": as_text(references.get("apiSpecFile")) or "Assumption: not present in YAML",
            })
            grouped[api_dir.name] = cases
    return grouped, inventory


def style_case_sheet(ws, rows: list[tuple[list[str], Path, dict[str, Any]]]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(1, len(rows) + 1)}"
    for index, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, index, header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number, (values, yaml_path, case) in enumerate(rows, 2):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_number, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row_number, 1).comment = Comment(
            "Source YAML: " + str(yaml_path) + "\nSource worksheet: " +
            as_text((case.get("references") or {}).get("sourceWorksheet")) +
            "\nAPI spec file: " + as_text((case.get("references") or {}).get("apiSpecFile")),
            "QA/API Test Generator",
        )
        ws.cell(row_number, 11).fill = PRIORITY_FILLS[values[10]]
        ws.row_dimensions[row_number].height = 92
    for column, width in enumerate(API_SHEET_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[1].height = 34


def build_workbook(grouped: dict[str, list[tuple[list[str], Path, dict[str, Any]]]], inventory: list[dict[str, str]], output: Path) -> None:
    all_rows = [item for api_rows in grouped.values() for item in api_rows]
    counts = Counter(item[0][2] for item in all_rows)
    categories = Counter(item[0][3] for item in all_rows)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "ECLIPSE App Dashboard API Validation Test Cases"
    summary["A1"].font = Font(bold=True, size=16, color="17365D")
    summary["A3"] = "Generated date"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    metrics = [("Total APIs covered", len(grouped)), ("Total test cases", len(all_rows)),
               ("Positive", counts["Positive"]), ("Negative", counts["Negative"]),
               ("Boundary", counts["Boundary"]), ("Edge/Resilience", counts["Edge/Resilience"])]
    for row, (label, value) in enumerate(metrics, 5):
        summary.cell(row, 1, label).font = Font(bold=True)
        summary.cell(row, 2, value).font = Font(bold=True)
    summary["A13"] = "Test category breakdown"
    summary["A13"].font = Font(bold=True, size=12)
    summary.append([])
    summary.cell(14, 1, "Category").fill = HEADER_FILL
    summary.cell(14, 2, "Count").fill = HEADER_FILL
    for cell in summary[14][:2]:
        cell.font = Font(bold=True, color="FFFFFF")
    for row, (label, count) in enumerate(sorted(categories.items()), 15):
        summary.cell(row, 1, label)
        summary.cell(row, 2, count)
    inventory_row = 15 + len(categories) + 2
    summary.cell(inventory_row, 1, "API inventory").font = Font(bold=True, size=12)
    columns = ["API folder", "HTTP Method", "Endpoint", "Source worksheet", "API spec file"]
    for column, label in enumerate(columns, 1):
        cell = summary.cell(inventory_row + 1, column, label)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
    for row, item in enumerate(inventory, inventory_row + 2):
        for column, key in enumerate(("api", "method", "endpoint", "sourceWorksheet", "apiSpecFile"), 1):
            summary.cell(row, column, item[key]).alignment = Alignment(vertical="top", wrap_text=True)
    summary.column_dimensions["A"].width = 28
    for column in "BCDE":
        summary.column_dimensions[column].width = 42
    summary.freeze_panes = "A15"

    validation = wb.create_sheet("Validation Areas")
    validation["A1"] = "KEY VALIDATION AREAS COVERED"
    validation["A1"].font = Font(bold=True, size=14, color="17365D")
    areas = [
        "Field validation: mandatory, optional, null, and omitted fields from YAML scenarios.",
        "Authentication: missing and invalid authentication credentials.",
        "Authorization: role and permission checks where represented by scenario tags or titles.",
        "Business rules: scenario assertions, duplicate submissions, and contract behavior.",
        "Data types and formats: schema, enum, date, numeric, and format assertions from YAML.",
        "Boundary conditions: minimum, maximum, null, omitted, and invalid boundary scenarios.",
        "Error handling: malformed or empty requests, downstream failures, and expected errors.",
        "Security: security-tagged authentication and invalid-input scenarios.",
        "HTTP protocol: method, endpoint, content type, and expected status code checks.",
        "Rate limiting/performance: included when represented in the generated YAML; otherwise marked as not defined.",
    ]
    validation["A3"] = "Area"
    validation["B3"] = "Coverage"
    for cell in validation[3][:2]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
    for row, area in enumerate(areas, 4):
        validation.cell(row, 1, area.split(":", 1)[0])
        validation.cell(row, 2, area).alignment = Alignment(wrap_text=True, vertical="top")
        validation.row_dimensions[row].height = 34
    validation.column_dimensions["A"].width = 30
    validation.column_dimensions["B"].width = 110

    for api_name, rows in grouped.items():
        ws = wb.create_sheet(api_name[:31])
        style_case_sheet(ws, rows)
    all_sheet = wb.create_sheet("All Test Cases")
    style_case_sheet(all_sheet, all_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Generated API YAML root")
    parser.add_argument("--output", type=Path, required=True, help="Output .xlsx path")
    args = parser.parse_args()
    grouped, inventory = load_cases(args.input)
    if not grouped:
        raise SystemExit(f"No executable YAML scenarios found under {args.input}")
    build_workbook(grouped, inventory, args.output)
    print(f"Generated {args.output} with {len(grouped)} APIs and {sum(map(len, grouped.values()))} test cases")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())