#!/usr/bin/env python3
"""
Generate comprehensive API validation test cases from DDD Excel file
"""

import openpyxl
import json
import os
from datetime import datetime
from pathlib import Path

xlsx_path = 'input/api/P&T_Local_Transfer_DDD_API_Spec_v1.xlsx'
output_dir = 'artifacts/Test_Case'

# Create output directory
Path(output_dir).mkdir(parents=True, exist_ok=True)

# Load workbook
wb = openpyxl.load_workbook(xlsx_path)

# Extract API index
index_sheet = wb['API_Specs_Index']
apis = []
for row in index_sheet.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:
        apis.append({
            'no': int(row[0]) if isinstance(row[0], (int, float)) else row[0],
            'name': str(row[1]).strip() if row[1] else '',
            'description': str(row[2]).strip() if row[2] else '',
            'method': str(row[3]).strip() if row[3] else '',
            'endpoint': str(row[4]).strip() if row[4] else '',
            'ownership': str(row[5]).strip() if row[5] else ''
        })

print(f"✓ Found {len(apis)} APIs")

# Extract detailed specs for each API
api_specs = {}
for api in apis:
    sheet_name = api['name']
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        spec = {
            'name': sheet_name,
            'method': api['method'],
            'endpoint': api['endpoint'],
            'description': api['description'],
            'ownership': api['ownership'],
            'request_params': [],
            'response_fields': [],
            'headers': [],
            'error_codes': []
        }
        
        # Extract all data from the sheet
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(row)
        
        # Parse structure
        current_section = None
        for i, row in enumerate(all_rows):
            if not any(row):  # Skip empty rows
                continue
            
            cell_value = str(row[0]).strip() if row[0] else ''
            
            # Detect sections
            if 'request' in cell_value.lower() or 'parameter' in cell_value.lower():
                current_section = 'request'
            elif 'response' in cell_value.lower():
                current_section = 'response'
            elif 'header' in cell_value.lower():
                current_section = 'headers'
            elif 'error' in cell_value.lower() or 'status' in cell_value.lower():
                current_section = 'error'
            elif current_section == 'request' and row[0] and row[1]:
                spec['request_params'].append({
                    'name': str(row[0]).strip() if row[0] else '',
                    'type': str(row[1]).strip() if row[1] else '',
                    'required': str(row[2]).strip() if len(row) > 2 and row[2] else 'No',
                    'description': str(row[3]).strip() if len(row) > 3 and row[3] else ''
                })
            elif current_section == 'response' and row[0] and row[1]:
                spec['response_fields'].append({
                    'name': str(row[0]).strip() if row[0] else '',
                    'type': str(row[1]).strip() if row[1] else '',
                    'description': str(row[2]).strip() if len(row) > 2 and row[2] else ''
                })
            elif current_section == 'error' and row[0] and row[1]:
                spec['error_codes'].append({
                    'code': str(row[0]).strip() if row[0] else '',
                    'description': str(row[1]).strip() if row[1] else ''
                })
        
        api_specs[sheet_name] = spec
        print(f"  ✓ {sheet_name}: {len(spec['request_params'])} request params, {len(spec['response_fields'])} response fields")

print(f"\n✓ Extracted specifications for {len(api_specs)} APIs")

# Generate test cases
test_summary = {
    'total_apis': len(apis),
    'total_test_cases': 0,
    'apis_covered': [api['name'] for api in apis],
    'validation_areas': [
        'Field Validation',
        'Authentication',
        'Business Rules',
        'Boundary Conditions',
        'Error Handling',
        'Data Type Checks',
        'Required Fields'
    ]
}

# Calculate expected test cases
positive_per_api = 2
negative_per_api = 10
test_summary['total_test_cases'] = len(apis) * (positive_per_api + negative_per_api)

# Create markdown output
md_content = f"""# P&T Local Transfer API - Comprehensive Test Cases

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---

## TEST SUMMARY

### Overview
- **Total APIs Covered:** {test_summary['total_apis']}
- **Total Test Cases:** {test_summary['total_test_cases']} (Positive: {len(apis) * positive_per_api}, Negative: {len(apis) * negative_per_api})
- **Test Case Format:** Positive + Negative scenarios per API
- **Coverage Ratio:** ~{positive_per_api + negative_per_api} test cases per API

### APIs Covered

"""

for api in apis:
    md_content += f"| {api['no']:2d} | {api['name']:40s} | {api['method']:6s} | {api['endpoint']:50s} |\n"
    if api['no'] == 1:
        md_content = md_content.replace(f"| {api['no']:2d}", "| No | API Name | Method | Endpoint |\n|----|----|--------|---------|\n" + f"| {api['no']:2d}")

md_content += """

### Key Validation Areas
- **Field Validation:** Mandatory/optional fields, data types, format validation
- **Authentication:** Invalid/missing tokens, expired credentials
- **Business Rules:** Duplicate submissions, business logic validation
- **Boundary Conditions:** Min/max values, string length, numeric ranges
- **Error Handling:** Error codes, error messages, status codes
- **Data Type Checks:** Type mismatches, invalid formats
- **Required Fields:** Missing mandatory parameters

### Assumptions
1. All APIs require proper authentication tokens in headers
2. Request/Response format is JSON
3. Standard HTTP status codes apply (200, 400, 401, 404, 500)
4. Validation rules follow REST API best practices
5. Rate limiting/throttling may apply (as per service configuration)

---

## DETAILED TEST CASES

"""

# Generate test cases for each API
test_case_counter = 1
for api_idx, api in enumerate(apis, 1):
    api_name = api['name']
    spec = api_specs.get(api_name, {})
    
    md_content += f"\n### API {api_idx}: {api_name}\n\n"
    md_content += f"**Method:** `{api['method']}`  \n"
    md_content += f"**Endpoint:** `{api['endpoint']}`  \n"
    md_content += f"**Description:** {api['description']}  \n"
    md_content += f"**Ownership:** {api['ownership']}\n\n"
    
    # Request parameters info
    if spec.get('request_params'):
        md_content += f"**Request Parameters ({len(spec['request_params'])}):**\n\n"
        md_content += "| Parameter | Type | Required | Description |\n"
        md_content += "|-----------|------|----------|-------------|\n"
        for param in spec['request_params'][:10]:  # Limit to first 10 for readability
            md_content += f"| {param.get('name', 'N/A')} | {param.get('type', 'N/A')} | {param.get('required', 'No')} | {param.get('description', '')[:50]} |\n"
        md_content += "\n"
    
    # Test Cases Table
    md_content += "#### Test Cases\n\n"
    md_content += "| TC ID | Test Title | Type | Category | Pre-conditions | Test Steps | Input/Payload | Expected Result | Priority |\n"
    md_content += "|-------|-----------|------|----------|---|---|---|---|----------|\n"
    
    # Positive Test Cases
    md_content += f"| TC_API{api_idx:02d}_P01 | Valid request with all mandatory fields | Positive | Field Validation | Valid auth token | Send valid request | All mandatory params populated correctly | HTTP 200, Valid response object | High |\n"
    md_content += f"| TC_API{api_idx:02d}_P02 | Valid request with optional fields | Positive | Field Validation | Valid auth token | Send request with optional params | Mandatory + optional params | HTTP 200, Response includes all fields | High |\n"
    
    # Negative Test Cases
    md_content += f"| TC_API{api_idx:02d}_N01 | Missing mandatory field | Negative | Field Validation | Valid auth token | Omit required parameter | Remove one mandatory field | HTTP 400, Error message indicates missing field | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N02 | Invalid data type | Negative | Data Type Check | Valid auth token | Send string in numeric field | String value in numeric parameter | HTTP 400, Type validation error | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N03 | Invalid format | Negative | Field Validation | Valid auth token | Send incorrectly formatted data | Invalid format (e.g., wrong date) | HTTP 400, Format validation error | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N04 | Missing authentication token | Negative | Authentication | No auth token | Send request without token | Request without Authorization header | HTTP 401, Unauthorized error | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N05 | Invalid authentication token | Negative | Authentication | Invalid token | Send expired/invalid token | Authorization: Bearer invalid_token | HTTP 401, Invalid token error | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N06 | Boundary value - below minimum | Negative | Boundary Condition | Valid auth token | Send value below minimum | Min value - 1 | HTTP 400, Out of range error | Medium |\n"
    md_content += f"| TC_API{api_idx:02d}_N07 | Boundary value - above maximum | Negative | Boundary Condition | Valid auth token | Send value above maximum | Max value + 1 | HTTP 400, Out of range error | Medium |\n"
    md_content += f"| TC_API{api_idx:02d}_N08 | Malformed JSON | Negative | Error Handling | Valid auth token | Send invalid JSON | Incomplete/broken JSON structure | HTTP 400, Malformed JSON error | Medium |\n"
    md_content += f"| TC_API{api_idx:02d}_N09 | SQL Injection attempt | Negative | Security | Valid auth token | Inject SQL in string field | '; DROP TABLE--; | HTTP 400, Invalid input error | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N10 | XSS payload | Negative | Security | Valid auth token | Inject XSS payload | <script>alert('xss')</script> | HTTP 400, Invalid input error | High |\n"
    md_content += f"| TC_API{api_idx:02d}_N11 | Empty payload | Negative | Error Handling | Valid auth token | Send empty request body | {{}} (empty object) | HTTP 400, Missing required fields error | Medium |\n"
    md_content += f"| TC_API{api_idx:02d}_N12 | Invalid HTTP method | Negative | Error Handling | Valid auth token | Use wrong HTTP method | Use POST instead of {api['method']} | HTTP 405, Method not allowed | Medium |\n"

md_content += """

---

## TEST EXECUTION SUMMARY

### Test Case Distribution
- **Total Positive Test Cases:** """ + str(len(apis) * 2) + """
- **Total Negative Test Cases:** """ + str(len(apis) * 10) + """
- **Total Test Cases:** """ + str(test_summary['total_test_cases']) + """

### Priority Breakdown
- **High Priority:** """ + str(len(apis) * 8) + """ (Field validation, Auth, Security)
- **Medium Priority:** """ + str(len(apis) * 4) + """ (Boundary conditions, Error handling)
- **Low Priority:** 0

### Recommended Test Execution Order
1. Positive test cases (validation of happy path)
2. High-priority negative test cases (auth, security, required fields)
3. Medium-priority negative test cases (boundary, error handling)

---

## NOTES AND ASSUMPTIONS

### General Assumptions
- All endpoints require valid Bearer token authentication
- Request/Response format is JSON
- Content-Type: application/json for all requests
- Standard HTTP status codes (200, 400, 401, 404, 500) apply
- API responses include error codes and descriptive error messages
- No specific rate limiting mentioned in DDD - standard practices assumed

### Field Validation Rules
- Mandatory fields must be present in request
- Data types must match specification
- String fields follow specified length constraints
- Numeric fields follow specified min/max ranges
- Date fields follow ISO 8601 format (YYYY-MM-DD)
- Enum fields accept only specified values

### Test Environment Setup
1. Ensure test environment is isolated from production
2. Use test credentials with proper permissions
3. Clear any cached data before each test run
4. Validate both positive and negative cases
5. Verify error messages are clear and actionable

### Known Limitations
- Test cases are template-based; specific payloads should be populated during test execution
- Boundary values should be confirmed from actual API implementation
- Error codes may vary based on backend implementation
- Response time SLAs not specified in test cases

---

**Document Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  
**Source Document:** P&T_Local_Transfer_DDD_API_Spec_v1.xlsx  
**Total Pages:** Generated from {len(apis)} API specifications
"""

# Save markdown file
md_file = f"{output_dir}/P&T_Local_Transfer_API_Test_Cases.md"
with open(md_file, 'w', encoding='utf-8') as f:
    f.write(md_content)

print(f"\n✓ Generated test cases markdown: {md_file}")

# Generate JSON summary
summary_json = {
    'generation_date': datetime.now().isoformat(),
    'source_file': 'P&T_Local_Transfer_DDD_API_Spec_v1.xlsx',
    'total_apis': test_summary['total_apis'],
    'total_test_cases': test_summary['total_test_cases'],
    'apis': [
        {
            'no': api['no'],
            'name': api['name'],
            'method': api['method'],
            'endpoint': api['endpoint'],
            'description': api['description'],
            'test_cases': 12,
            'positive': 2,
            'negative': 10
        }
        for api in apis
    ],
    'validation_areas': test_summary['validation_areas']
}

json_file = f"{output_dir}/test_cases_summary.json"
with open(json_file, 'w', encoding='utf-8') as f:
    json.dump(summary_json, f, indent=2)

print(f"✓ Generated summary JSON: {json_file}")
print(f"\n✓ Test case generation complete!")
print(f"\nGenerated Files:")
print(f"  - {md_file}")
print(f"  - {json_file}")

