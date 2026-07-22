#!/usr/bin/env python3
"""Generate comprehensive API validation test cases from DEP_App_Dashboard_DDD_API_Design_v1.xlsx"""

import openpyxl
import json
from datetime import datetime
from typing import Dict, List, Any, Tuple

class APISpecExtractor:
    def __init__(self, excel_file):
        self.wb = openpyxl.load_workbook(excel_file)
        self.apis = []
        
    def extract_all_apis(self) -> List[Dict]:
        """Extract all API specifications from the workbook"""
        # Get sheet names, skip README and INDEX
        api_sheets = [s for s in self.wb.sheetnames if s not in ['README', 'INDEX']]
        
        for sheet_name in api_sheets:
            api_spec = self.extract_api_spec(sheet_name)
            if api_spec:
                self.apis.append(api_spec)
        
        return self.apis
    
    def extract_api_spec(self, sheet_name: str) -> Dict:
        """Extract API specification from a single sheet"""
        ws = self.wb[sheet_name]
        spec = {
            'sheet_name': sheet_name,
            'service': None,
            'description': None,
            'business_rule': None,
            'method': None,
            'url': None,
            'message_type': None,
            'headers': [],
            'body_params': [],
            'query_params': [],
            'response_fields': [],
            'request_sample': None,
        }
        
        current_section = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not any(row):
                continue
                
            col_a = str(row[0]).strip() if row[0] else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            # Extract basic info
            if col_a == 'Service':
                spec['service'] = col_b
            elif col_a == 'Description':
                spec['description'] = col_b
            elif col_a == 'Business Rule':
                spec['business_rule'] = col_b
            elif col_a == 'Method':
                spec['method'] = col_b
            elif col_a == 'URL':
                spec['url'] = col_b
            elif col_a == 'Message Type':
                spec['message_type'] = col_b
            elif col_a == 'Request Sample':
                spec['request_sample'] = col_b
            
            # Parse sections
            if col_a == 'Request' and col_b == 'HTTP Header':
                current_section = 'headers'
            elif col_a == 'HTTP Body' or (current_section == 'headers' and col_a and col_a not in ['HTTP Body', 'Request Parameter']):
                if col_a == 'HTTP Body':
                    current_section = 'body'
            elif col_a == 'Request Parameter':
                current_section = 'query'
            elif col_a == 'Response':
                current_section = 'response'
            elif current_section and row_idx > 1:
                # Parse parameter rows
                if len(row) >= 7 and row[1] and col_a in ['', None]:  # Parameter row
                    param_name = str(row[1]).strip()
                    if param_name and param_name not in ['Name', 'Type', 'Length']:
                        param = {
                            'name': param_name,
                            'type': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                            'length': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                            'mandatory': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                            'sample': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                            'description': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                        }
                        
                        if current_section == 'headers':
                            spec['headers'].append(param)
                        elif current_section == 'body':
                            spec['body_params'].append(param)
                        elif current_section == 'query':
                            spec['query_params'].append(param)
                        elif current_section == 'response':
                            spec['response_fields'].append(param)
        
        return spec if spec['service'] else None

class TestCaseGenerator:
    def __init__(self, apis: List[Dict]):
        self.apis = apis
        self.test_cases = []
        self.test_case_id_counter = 1
        
    def generate_all_test_cases(self) -> List[Dict]:
        """Generate test cases for all APIs"""
        for api in self.apis:
            self.generate_test_cases_for_api(api)
        return self.test_cases
    
    def generate_test_cases_for_api(self, api: Dict):
        """Generate comprehensive test cases for a single API"""
        api_num = len([a for a in self.apis[:self.apis.index(api)+1]])
        
        # Extract mandatory headers
        mandatory_headers = [h for h in api['headers'] if h['mandatory'].upper() == 'Y']
        optional_headers = [h for h in api['headers'] if h['mandatory'].upper() != 'Y']
        
        # Extract body/query parameters
        all_params = api['body_params'] + api['query_params']
        mandatory_params = [p for p in all_params if p['mandatory'].upper() == 'Y']
        optional_params = [p for p in all_params if p['mandatory'].upper() != 'Y']
        
        # POSITIVE TEST CASES
        self._add_positive_test_case(api, api_num, "Valid request with all mandatory headers",
                                     True, False, False)
        
        if optional_headers:
            self._add_positive_test_case(api, api_num, "Valid request with optional headers included",
                                         True, True, False)
        
        if mandatory_params:
            self._add_positive_test_case(api, api_num, "Valid request with all mandatory parameters",
                                         True, False, False)
        
        # NEGATIVE TEST CASES
        if mandatory_headers:
            for header in mandatory_headers:
                self._add_negative_test_case(api, api_num, f"Missing mandatory header: {header['name']}",
                                             "Missing Header", header)
        
        if mandatory_params:
            for param in mandatory_params:
                self._add_negative_test_case(api, api_num, f"Missing mandatory parameter: {param['name']}",
                                             "Missing Parameter", param)
        
        # Invalid data type test cases
        if all_params:
            self._add_negative_test_case(api, api_num, "Invalid data type in request",
                                         "Invalid Data Type", all_params[0])
        
        # Invalid format test cases
        self._add_negative_test_case(api, api_num, "Malformed JSON in request body",
                                     "Invalid Format", None)
        
        # Authentication test cases
        self._add_negative_test_case(api, api_num, "Missing authentication token",
                                     "Authentication Error", None)
        self._add_negative_test_case(api, api_num, "Invalid/expired authentication token",
                                     "Authentication Error", None)
        
        # Unauthorized access
        self._add_negative_test_case(api, api_num, "Request with invalid authorization token format",
                                     "Authorization Error", None)
        
        # Boundary test cases
        if mandatory_params:
            self._add_negative_test_case(api, api_num, "Request with out-of-range values",
                                         "Boundary Violation", mandatory_params[0])
        
        # Special characters/XSS test cases
        if all_params:
            self._add_negative_test_case(api, api_num, "SQL injection attempt in request parameter",
                                         "Security", all_params[0])
            self._add_negative_test_case(api, api_num, "XSS payload in request parameter",
                                         "Security", all_params[0])
        
        # Invalid HTTP method
        self._add_negative_test_case(api, api_num, "Incorrect HTTP method (POST instead of GET)",
                                     "Invalid HTTP Method", None)
        
        # Empty payload
        if api['method'].upper() == 'POST':
            self._add_negative_test_case(api, api_num, "Empty or null request body",
                                         "Invalid Payload", None)
    
    def _add_positive_test_case(self, api: Dict, api_num: int, title: str, 
                                with_headers: bool, with_optional: bool, with_params: bool):
        """Add a positive test case"""
        test_case = {
            'id': f'TC_API{api_num:02d}_{self.test_case_id_counter:03d}',
            'title': title,
            'type': 'Positive',
            'category': 'Field Validation / Request Format',
            'api_endpoint': api['url'],
            'method': api['method'],
            'service': api['service'],
            'preconditions': 'User is authenticated; Valid bearer token available; Required headers configured',
            'steps': [
                f"1. Prepare valid request with {api['method']} method to {api['url']}",
                "2. Include all mandatory headers with valid values",
                "3. Set Content-Type to application/json",
                "4. Send the API request",
                "5. Verify the response is received"
            ],
            'input_data': self._generate_sample_request(api, with_headers, with_optional, with_params),
            'expected_result': f"HTTP 200 OK; Response body contains success: true; Response includes message object with code and key; Data object returned with expected schema as per API specification",
            'priority': 'High',
        }
        self.test_cases.append(test_case)
        self.test_case_id_counter += 1
    
    def _add_negative_test_case(self, api: Dict, api_num: int, title: str, 
                                category: str, affected_param: Dict = None):
        """Add a negative test case"""
        preconditions = "User is authenticated; Valid bearer token available"
        steps = [f"1. Prepare {api['method']} request to {api['url']}"]
        
        if category == 'Missing Header':
            steps.extend([
                f"2. Exclude the mandatory header: {affected_param['name']}",
                "3. Include other required headers",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = f"HTTP 400/401 Bad Request; Response should include error message indicating missing header: {affected_param['name']}"
        
        elif category == 'Missing Parameter':
            steps.extend([
                f"2. Exclude the mandatory parameter: {affected_param['name']}",
                "3. Include other required parameters",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = f"HTTP 400 Bad Request; Response should include error message indicating missing required field: {affected_param['name']}"
        
        elif category == 'Invalid Data Type':
            steps.extend([
                f"2. Set parameter '{affected_param['name']}' (expected type: {affected_param['type']}) with string value '\"invalid\"'",
                "3. Include other required parameters with valid values",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = f"HTTP 400 Bad Request; Response indicates invalid data type for {affected_param['name']}"
        
        elif category == 'Invalid Format':
            steps.extend([
                "2. Send malformed JSON (e.g., missing closing bracket)",
                "3. Ensure Content-Type is application/json",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = "HTTP 400 Bad Request; Response indicates malformed request body"
        
        elif category == 'Authentication Error':
            if 'invalid/expired' in title.lower():
                steps.extend([
                    "2. Set Authorization header to 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2MDAwMDAwMDB9.invalid'",
                    "3. Include other required headers",
                    "4. Send the API request",
                    "5. Verify the error response"
                ])
            else:
                steps.extend([
                    "2. Omit the Authorization header",
                    "3. Include other required headers",
                    "4. Send the API request",
                    "5. Verify the error response"
                ])
            expected = "HTTP 401 Unauthorized; Response indicates missing or invalid authentication token"
        
        elif category == 'Authorization Error':
            steps.extend([
                "2. Set Authorization header to 'Bearer invalid_format_token'",
                "3. Include other required headers",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = "HTTP 401 Unauthorized; Response indicates invalid token format"
        
        elif category == 'Boundary Violation':
            steps.extend([
                f"2. Set parameter '{affected_param['name']}' to value exceeding maximum limit or below minimum",
                "3. Include other required parameters with valid values",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = f"HTTP 400 Bad Request; Response indicates out-of-range value for {affected_param['name']}"
        
        elif category == 'Security':
            if 'SQL' in title:
                payload = "'; DROP TABLE users; --"
                steps[0] = f"1. Prepare {api['method']} request with SQL injection payload: {payload}"
            else:
                payload = "<script>alert('XSS')</script>"
                steps[0] = f"1. Prepare {api['method']} request with XSS payload: {payload}"
            
            steps.extend([
                f"2. Include the malicious payload in request parameter",
                "3. Send the API request",
                "4. Verify the API handles the input safely"
            ])
            expected = "HTTP 400/200; Response should not execute the malicious code; Input should be sanitized or rejected"
        
        elif category == 'Invalid HTTP Method':
            steps.extend([
                f"2. Send POST request instead of {api['method']}",
                "3. Include valid request body and headers",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = "HTTP 405 Method Not Allowed; Response indicates unsupported HTTP method"
        
        elif category == 'Invalid Payload':
            steps.extend([
                "2. Send empty request body or null value",
                "3. Include required headers",
                "4. Send the API request",
                "5. Verify the error response"
            ])
            expected = "HTTP 400 Bad Request; Response indicates empty or invalid request body"
        
        else:
            expected = "Error response received with appropriate status code"
        
        test_case = {
            'id': f'TC_API{api_num:02d}_{self.test_case_id_counter:03d}',
            'title': title,
            'type': 'Negative',
            'category': category,
            'api_endpoint': api['url'],
            'method': api['method'],
            'service': api['service'],
            'preconditions': preconditions,
            'steps': steps,
            'input_data': f"Invalid/Incomplete request payload for {api['url']}",
            'expected_result': expected,
            'priority': 'High' if category in ['Missing Header', 'Missing Parameter', 'Authentication Error'] else 'Medium',
        }
        self.test_cases.append(test_case)
        self.test_case_id_counter += 1
    
    def _generate_sample_request(self, api: Dict, with_headers: bool, with_optional: bool, with_params: bool) -> str:
        """Generate a sample valid request"""
        request = f"Method: {api['method']}\n"
        request += f"URL: {api['url']}\n"
        request += f"Headers:\n"
        
        # Mandatory headers
        request += "  Authorization: Bearer <valid_access_token>\n"
        request += "  X-APP-PLATFORM: IOS\n"
        request += "  X-APP-OS-VERSION: 26.1\n"
        request += "  X-APP-LOCALE: en_MY\n"
        
        if with_optional and any(h['mandatory'].upper() != 'Y' for h in api['headers']):
            request += "  Accept: application/json\n"
        
        if api['method'].upper() == 'POST' and api['body_params']:
            request += "Request Body (JSON):\n{\n"
            for param in api['body_params'][:2]:
                request += f'  "{param["name"]}": {param["sample"]},\n'
            request = request.rstrip(',\n') + "\n}\n"
        
        return request

# Main execution
if __name__ == '__main__':
    print("=" * 100)
    print("EXTRACTING API SPECIFICATIONS FROM DDD FILE")
    print("=" * 100)
    
    extractor = APISpecExtractor(r'input\api\DEP_App_Dashboard_DDD_API_Design_v1.xlsx')
    apis = extractor.extract_all_apis()
    
    print(f"\nFound {len(apis)} API specifications")
    print("\nAPIs extracted:")
    for i, api in enumerate(apis, 1):
        print(f"{i}. {api['service']} ({api['method']} {api['url']})")
        print(f"   Headers: {len(api['headers'])}, Body Params: {len(api['body_params'])}, Query Params: {len(api['query_params'])}")
    
    print("\n" + "=" * 100)
    print("GENERATING TEST CASES")
    print("=" * 100)
    
    generator = TestCaseGenerator(apis)
    test_cases = generator.generate_all_test_cases()
    
    print(f"\nGenerated {len(test_cases)} test cases")
    print("\nTest cases by API:")
    current_api = None
    tc_count = 0
    for tc in test_cases:
        if tc['service'] != current_api:
            current_api = tc['service']
            tc_count = 0
            print(f"\n{current_api}:")
        tc_count += 1
        print(f"  {tc_count}. {tc['title']} ({tc['type']})")
    
    # Save results to JSON for further processing
    with open('test_cases_metadata.json', 'w') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'total_apis': len(apis),
            'total_test_cases': len(test_cases),
            'apis': apis,
            'test_cases': test_cases
        }, f, indent=2)
    
    print(f"\n✓ Test case metadata saved to test_cases_metadata.json")
    print(f"\nSUMMARY:")
    print(f"  Total APIs: {len(apis)}")
    print(f"  Total Test Cases: {len(test_cases)}")
    print(f"  Positive Test Cases: {len([tc for tc in test_cases if tc['type'] == 'Positive'])}")
    print(f"  Negative Test Cases: {len([tc for tc in test_cases if tc['type'] == 'Negative'])}")
